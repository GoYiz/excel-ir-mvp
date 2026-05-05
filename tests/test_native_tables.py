from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'src'
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))


class ExcelIRNativeTableTests(unittest.TestCase):
    def make_native_table_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = 'NativeTable'
        ws.append(['Region', 'Revenue', 'Cost'])
        ws.append(['North', 100, 70])
        ws.append(['South', 120, 80])
        tab = Table(displayName='SalesTable', ref='A1:C3')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
        wb.save(path)

    def test_single_row_native_table_roundtrips_as_native(self):
        from excel_ir_mvp.excel_ir_plus import diff_workbooks_plus, parse_workbook_plus, rebuild_workbook_plus
        src = ROOT / 'native_table_fixture.xlsx'
        rebuilt = ROOT / 'native_table_rebuilt.xlsx'
        self.make_native_table_workbook(src)
        ir = parse_workbook_plus(str(src))
        table = ir['workbook']['sheets'][0]['extra']['tables'][0]
        self.assertEqual(table.get('table_kind'), 'native')
        self.assertTrue(table.get('native_table_supported'))
        rebuild_workbook_plus(ir, str(rebuilt))
        wb = load_workbook(rebuilt)
        self.assertIn('SalesTable', wb['NativeTable'].tables)
        reparsed = parse_workbook_plus(str(rebuilt))
        rebuilt_table = reparsed['workbook']['sheets'][0]['extra']['tables'][0]
        self.assertEqual(rebuilt_table.get('table_kind'), 'native')
        self.assertEqual(diff_workbooks_plus(str(src), str(rebuilt))['diff_count'], 0)

    def test_semantic_table_kind_is_honored_even_without_legacy_flag(self):
        from excel_ir_mvp.excel_ir_plus import parse_workbook_plus, rebuild_workbook_plus
        src = ROOT / 'native_table_fixture_semantic_override.xlsx'
        rebuilt = ROOT / 'native_table_semantic_override_rebuilt.xlsx'
        self.make_native_table_workbook(src)
        ir = parse_workbook_plus(str(src))
        table = ir['workbook']['sheets'][0]['extra']['tables'][0]
        table['table_kind'] = 'semantic'
        table.pop('native_table_supported', None)
        rebuild_workbook_plus(ir, str(rebuilt))
        wb = load_workbook(rebuilt)
        self.assertNotIn('SalesTable', wb['NativeTable'].tables)


if __name__ == '__main__':
    unittest.main()

from __future__ import annotations

import json
import os
import subprocess
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'src'
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from tests.fixtures_loader import fixture_path


class ExcelIRMVPTests(unittest.TestCase):
    def run_cmd(self, args, env=None):
        p = subprocess.run(args, cwd=ROOT, env=env, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr[-1200:] + p.stdout[-1200:])
        return p

    def module_env(self):
        env = os.environ.copy()
        env['PYTHONPATH'] = str(SRC) + (os.pathsep + env['PYTHONPATH'] if env.get('PYTHONPATH') else '')
        return env

    def test_package_import(self):
        import excel_ir_mvp
        self.assertEqual(excel_ir_mvp.__version__, '2.0.0a16')
        self.assertTrue(callable(excel_ir_mvp.parse_workbook_plus))
        self.assertIn('openpyxl', excel_ir_mvp.available_engines())
        self.assertTrue(callable(excel_ir_mvp.parse))
        public = set(excel_ir_mvp.__all__)
        self.assertIn('parse', public)
        self.assertIn('rebuild', public)
        self.assertIn('header_edit', public)

    def test_module_cli_smoke(self):
        self.run_cmd(['python3', '-m', 'excel_ir_mvp.excel_ir_cli', 'doctor'], env=self.module_env())

    def test_package_module_main_smoke(self):
        self.run_cmd(['python3', '-m', 'excel_ir_mvp', 'doctor'], env=self.module_env())

    def test_api_parse_diff_validate(self):
        from excel_ir_mvp.excel_ir_plus import parse_workbook_plus, rebuild_workbook_plus, diff_workbooks_plus
        from excel_ir_mvp.validate_ir import validate_basic_types, validate_json_schema
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        self.assertFalse(validate_json_schema(ir, 'ir.schema.json'))
        self.assertFalse(validate_basic_types(ir))
        out = ROOT / 'api_rebuilt.xlsx'
        rebuild_workbook_plus(ir, str(out))
        diff = diff_workbooks_plus(str(fixture_path('complex_report.xlsx')), str(out))
        self.assertEqual(diff['diff_count'], 0)

    def test_excel_ir_plus_small_helpers(self):
        from excel_ir_mvp import excel_ir_plus as xp
        self.assertEqual(xp.formula_ref_to_range("'Sheet 1'!$A$1:$B$2")['ref'], 'A1:B2')
        self.assertEqual(xp.formula_ref_to_range('$C$3')['min_row'], 3)
        self.assertIsNone(xp.formula_ref_to_range(None))
        self.assertEqual(xp.chart_title_text(None), None)
        self.assertEqual(xp.table_style_to_dict(None), {})
        self.assertIsNone(xp.table_style_from_dict({}))
        self.assertEqual(xp._strip_none({'a': 1, 'b': None, 'c': [1, None]}), {'a': 1, 'c': [1]})

    def test_models_validation_errors(self):
        from excel_ir_mvp.models import CellIR, SheetIR, TableIR, WorkbookIR, validate_basic_types
        self.assertTrue(CellIR(row=0, col=1).validate('A1'))
        self.assertTrue(TableIR(name='', ref='bad', field_map={'x': '1'}).validate())
        self.assertTrue(SheetIR(name='', cells={'bad': CellIR(row=1, col=1)}).validate())
        self.assertTrue(WorkbookIR(schema_version='', sheets=[]).validate())
        self.assertTrue(validate_basic_types({'schema_version': 'x', 'workbook': {'sheets': [{'name': '', 'cells': {'BAD': {'row': 0, 'col': 0}}}]}}))

    def test_reports_and_review_render(self):
        from excel_ir_mvp import audit_report, diff_report
        from excel_ir_mvp.field_map_review_app import main as review_main
        diff = {'diff_count': 1, 'truncated': False, 'diffs': [{'sheet': 'S', 'coord': 'A1', 'type': 'value', 'a': 1, 'b': 2}]}
        plan = {'validation': [], 'actions': [{'op': 'set_cell', 'sheet': 'S', 'coord': 'A1', 'value': 2}]}
        tx = {'ok': True, 'impact': {'cells_changed': 1}, 'impact_graph': {'nodes': {}}, 'actions': [{'index': 1, 'op': 'set_cell', 'cell_diffs_sample': [{'sheet': 'S', 'coord': 'A1'}]}]}
        html = diff_report.render(diff, 'Diff Title', plan, tx)
        self.assertIn('Diff Title', html)
        self.assertIn('Patch Dry-run Plan', html)
        self.assertIn('Transaction Apply Log', html)
        audit_html = audit_report.render(tx, 'Audit Title')
        self.assertIn('Audit Title', audit_html)
        out = ROOT / 'review_api.html'
        review_main([str(fixture_path('complex_ir_v07.json')), str(out)])
        self.assertIn('Field Map Review App', out.read_text(encoding='utf-8'))

    def test_corpus_runner_api(self):
        import corpus_runner
        config = corpus_runner.load_config('tests/fixtures/corpus_config.json')
        self.assertGreaterEqual(len(config['samples']), 4)
        config = dict(config)
        config['output_dir'] = 'corpus_results_api'
        summary = corpus_runner.run_corpus(config)
        self.assertTrue(summary['ok'])
        self.assertTrue((ROOT / summary['report_html']).exists())
        self.assertIn('synthetic_complex', summary['categories'])
        self.assertEqual(summary['results'][0]['diff_count'], 0)

    def test_bench_budget_api(self):
        import bench
        summary = bench.run_bench(max_total_seconds=180.0)
        self.assertTrue(summary['ok'])
        self.assertLessEqual(summary['total_seconds'], 180.0)

    def test_validate_load_and_schema_errors(self):
        from excel_ir_mvp.validate_ir import load, validate_json_schema
        self.assertIn('required', load('ir.schema.json'))
        errors = validate_json_schema({'actions': []}, 'patch.schema.json')
        self.assertEqual(errors, [])
        errors = validate_json_schema({}, 'patch.schema.json')
        self.assertTrue(any(e.get('level') == 'error' for e in errors))

    def test_roundtrip_diff_zero(self):
        self.run_cmd(['python3', 'excel_ir_cli.py', 'parse', str(fixture_path('complex_report.xlsx')), 'unittest_ir.json'])
        self.run_cmd(['python3', 'excel_ir_cli.py', 'rebuild', 'unittest_ir.json', 'unittest_rebuilt.xlsx'])
        self.run_cmd(['python3', 'excel_ir_cli.py', 'diff', str(fixture_path('complex_report.xlsx')), 'unittest_rebuilt.xlsx', 'unittest_diff.json'])
        diff = json.loads((ROOT / 'unittest_diff.json').read_text(encoding='utf-8'))
        self.assertEqual(diff['diff_count'], 0)

    def test_patch_log_has_impact(self):
        self.run_cmd(['python3', 'excel_ir_cli.py', 'patch', str(fixture_path('complex_ir_v07.json')), str(fixture_path('v08_patch.json')), 'unittest_patched_ir.json', '--log', 'unittest_tx.json'])
        log = json.loads((ROOT / 'unittest_tx.json').read_text(encoding='utf-8'))
        self.assertTrue(log['ok'])
        self.assertIn('impact_graph', log)
        self.assertTrue(any(a.get('cell_diffs_sample') for a in log['actions']))

    def test_validate(self):
        self.run_cmd(['python3', 'excel_ir_cli.py', 'validate', 'ir', str(fixture_path('complex_ir_v07.json'))])
        self.run_cmd(['python3', 'excel_ir_cli.py', 'validate', 'patch', str(fixture_path('v08_patch.json'))])
    def test_package_corpus_runner_helpers(self):
        from excel_ir_mvp import corpus_runner as pkg_corpus
        results = [
            {'category': 'synthetic_complex', 'parse_ok': True, 'rebuild_ok': True, 'diff_count': 0},
            {'category': 'semantic_table', 'parse_ok': False, 'rebuild_ok': False, 'diff_count': 2},
        ]
        cats = pkg_corpus._category_summary(results)
        self.assertEqual(cats['synthetic_complex']['ok'], 1)
        self.assertEqual(cats['semantic_table']['failed'], 1)
        self.assertTrue(pkg_corpus.load_config('corpus_config.json')['samples'])
        listed = pkg_corpus.list_samples(pkg_corpus.load_config('corpus_config.json'))
        self.assertIn('metadata_roundtrip', listed['categories'])
        self.assertIn('native_table', listed['categories'])
        self.assertIn('semantic_table', listed['categories'])
        html = pkg_corpus.render_summary_html({'ok': True, 'categories': cats, 'results': results})
        self.assertIn('Excel IR Corpus Report', html)

    def test_validate_ir_error_branches(self):
        from excel_ir_mvp import validate_ir
        self.assertTrue(any(e['level'] == 'error' for e in validate_ir.validate_basic_types({'schema_version': 'x'})))
        self.assertTrue(any(e['level'] == 'error' for e in validate_ir.validate_json_schema({'bad': True}, 'patch.schema.json')))

    def test_cli_main_metadata_diff_inprocess(self):
        import contextlib, io
        from excel_ir_mvp.excel_ir_cli import main as cli_main
        from excel_ir_mvp.excel_ir_plus import collect_semantic_metadata, parse_workbook_plus, save_json, load_json
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        a = ROOT / 'cli_main_meta_a.json'
        b = ROOT / 'cli_main_meta_b.json'
        out = ROOT / 'cli_main_meta_diff.json'
        save_json(collect_semantic_metadata(ir), str(a))
        save_json(collect_semantic_metadata(ir), str(b))
        saved_argv = sys.argv[:]
        buf = io.StringIO()
        try:
            sys.argv = ['excel-ir', 'metadata', 'diff', str(a), str(b), str(out)]
            with contextlib.redirect_stdout(buf):
                cli_main()
        finally:
            sys.argv = saved_argv
        self.assertEqual(load_json(str(out))['diff_count'], 0)
    def test_cli_main_corpus_list_report_inprocess(self):
        import contextlib, io
        from excel_ir_mvp.excel_ir_cli import main as cli_main
        saved_argv = sys.argv[:]
        buf = io.StringIO()
        try:
            sys.argv = ['excel-ir', 'corpus', 'list', '--config', 'corpus_config.json']
            with contextlib.redirect_stdout(buf):
                cli_main()
        finally:
            sys.argv = saved_argv
        self.assertIn('synthetic_complex', json.loads(buf.getvalue())['categories'])

        summary = {'ok': True, 'categories': {'synthetic_complex': {'count': 1, 'ok': 1, 'failed': 0, 'diff_count_total': 0}}, 'results': [{'name': 's', 'category': 'synthetic_complex', 'parse_ok': True, 'rebuild_ok': True, 'diff_count': 0}]}
        summary_path = ROOT / 'cli_corpus_summary.json'
        html_path = ROOT / 'cli_corpus_report.html'
        summary_path.write_text(json.dumps(summary), encoding='utf-8')
        saved_argv = sys.argv[:]
        try:
            sys.argv = ['excel-ir', 'corpus', 'report', str(summary_path), str(html_path)]
            with contextlib.redirect_stdout(io.StringIO()):
                cli_main()
        finally:
            sys.argv = saved_argv
        self.assertIn('Excel IR Corpus Report', html_path.read_text(encoding='utf-8'))
    def test_compare_ir_cli_and_metadata_strip(self):
        from excel_ir_mvp.excel_ir_plus import compare_ir_files, load_json, parse_workbook_plus, rebuild_workbook_plus
        ir_a = ROOT / 'alpha9_a.ir.json'
        ir_b = ROOT / 'alpha9_b.ir.json'
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        ir_a.write_text(json.dumps(ir, ensure_ascii=False, indent=2), encoding='utf-8')
        ir_b.write_text(json.dumps(ir, ensure_ascii=False, indent=2), encoding='utf-8')
        self.assertTrue(compare_ir_files(str(ir_a), str(ir_b))['ok'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'compare-ir', str(ir_a), str(ir_b), 'alpha9_ir_diff.json'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(load_json(str(ROOT / 'alpha9_ir_diff.json'))['diff_count'], 0)
        with_metadata = ROOT / 'alpha9_with_metadata.xlsx'
        stripped = ROOT / 'alpha9_stripped.xlsx'
        rebuild_workbook_plus(ir, str(with_metadata))
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'strip', str(stripped), '--from-xlsx', str(with_metadata)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertTrue(json.loads(p.stdout)['removed'])
        overview = json.loads(subprocess.run(['python3', 'excel_ir_cli.py', 'inspect', str(stripped)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).stdout)
        self.assertFalse(overview['hidden_metadata']['present'])
    def test_inprocess_auxiliary_entrypoints_for_coverage(self):
        import contextlib, io, runpy
        from excel_ir_mvp.excel_ir_cli import main as cli_main
        import excel_ir_mvp.audit_report as audit_mod
        import excel_ir_mvp.bench as pkg_bench
        import excel_ir_mvp.field_map_review as fmr

        saved_argv = sys.argv[:]
        try:
            # Cover python -m excel_ir_mvp in-process.
            sys.argv = ['excel-ir', 'doctor']
            with contextlib.redirect_stdout(io.StringIO()):
                runpy.run_module('excel_ir_mvp.__main__', run_name='__main__')

            # Cover compare-ir CLI in-process.
            out = ROOT / 'inprocess_compare_ir.json'
            sys.argv = ['excel-ir', 'compare-ir', str(fixture_path('complex_ir_v07.json')), str(fixture_path('complex_ir_v07.json')), str(out)]
            with contextlib.redirect_stdout(io.StringIO()):
                cli_main()
            self.assertEqual(json.loads(out.read_text(encoding='utf-8'))['diff_count'], 0)

            # Cover legacy field_map_review main, including usage branch.
            review = ROOT / 'legacy_field_map_review.html'
            sys.argv = ['field_map_review.py', str(fixture_path('complex_ir_v07.json')), str(review)]
            with contextlib.redirect_stdout(io.StringIO()):
                fmr.main()
            self.assertIn('Field Map Review', review.read_text(encoding='utf-8'))
            sys.argv = ['field_map_review.py']
            with contextlib.redirect_stderr(io.StringIO()):
                with self.assertRaises(SystemExit):
                    fmr.main()

            # Cover audit_report main and usage branch.
            tx = ROOT / 'inprocess_audit_tx.json'
            tx.write_text(json.dumps({'impact': {'cells_changed': 1}, 'impact_graph': {}, 'actions': [{'index': 1, 'op': 'set_cell', 'impact': {}, 'cell_diffs_sample': [{'sheet': 'S', 'coord': 'A1'}]}]}), encoding='utf-8')
            audit_html = ROOT / 'inprocess_audit.html'
            sys.argv = ['audit_report.py', str(tx), str(audit_html), 'Aux Audit']
            with contextlib.redirect_stdout(io.StringIO()):
                audit_mod.main()
            self.assertIn('Aux Audit', audit_html.read_text(encoding='utf-8'))
            sys.argv = ['audit_report.py']
            with contextlib.redirect_stderr(io.StringIO()):
                with self.assertRaises(SystemExit):
                    audit_mod.main()

            # Cover package bench without spawning slow subprocesses.
            old_run = pkg_bench.run
            try:
                pkg_bench.run = lambda cmd: {'cmd': cmd, 'returncode': 0, 'seconds': 0.01}
                summary = pkg_bench.run_bench(max_total_seconds=1.0)
                self.assertTrue(summary['ok'])
                sys.argv = ['bench.py']
                with contextlib.redirect_stdout(io.StringIO()):
                    pkg_bench.main()
                self.assertTrue((ROOT / 'bench_results.json').exists())
            finally:
                pkg_bench.run = old_run
        finally:
            sys.argv = saved_argv
    def test_alpha10_anonymize_status_and_compare_modes(self):
        from excel_ir_mvp.excel_ir_plus import anonymize_workbook_xlsx, metadata_status_xlsx, parse_workbook_plus, rebuild_workbook_plus, compare_ir_files
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        rebuilt = ROOT / 'alpha10_rebuilt.xlsx'
        rebuild_workbook_plus(ir, str(rebuilt))
        status = metadata_status_xlsx(str(rebuilt))
        self.assertTrue(status['present'])
        self.assertTrue(status['checksum_ok'])
        anon = ROOT / 'alpha10_anonymized.xlsx'
        result = anonymize_workbook_xlsx(str(rebuilt), str(anon))
        self.assertTrue(result['ok'])
        self.assertGreater(result['cells_changed'], 0)
        self.assertFalse(metadata_status_xlsx(str(anon))['present'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'status', str(rebuilt)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertTrue(json.loads(p.stdout)['present'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'anonymize', str(fixture_path('complex_report.xlsx')), str(ROOT / 'alpha10_cli_anon.xlsx')], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        ir_a = ROOT / 'alpha10_a.ir.json'
        ir_b = ROOT / 'alpha10_b.ir.json'
        ir2 = json.loads(json.dumps(ir, ensure_ascii=False))
        ir2['workbook']['sheets'][0]['extra']['tables'][0]['table_kind'] = 'native'
        ir_a.write_text(json.dumps(ir, ensure_ascii=False), encoding='utf-8')
        ir_b.write_text(json.dumps(ir2, ensure_ascii=False), encoding='utf-8')
        self.assertFalse(compare_ir_files(str(ir_a), str(ir_b), mode='semantic')['ok'])
        self.assertTrue(compare_ir_files(str(ir_a), str(ir_b), mode='structural')['ok'])
        with self.assertRaises(ValueError):
            compare_ir_files(str(ir_a), str(ir_b), mode='bad')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'compare-ir', '--semantic-only', str(ir_a), str(ir_b)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertNotEqual(p.returncode, 0)
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'compare-ir', '--structural-only', str(ir_a), str(ir_b)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
    def test_stream_edit_and_formula_computed_values(self):
        from openpyxl import load_workbook
        from excel_ir_mvp.excel_ir import stream_find_cell_xlsx, stream_update_first_match_xlsx
        from excel_ir_mvp.excel_ir_plus import parse_workbook_plus
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        formula_cells = [c for s in ir['workbook']['sheets'] for c in s['cells'].values() if c.get('data_type') == 'f']
        self.assertTrue(formula_cells)
        self.assertTrue(all('computed_value' in c for c in formula_cells))
        self.assertTrue(all('computed_value_source' not in c for c in formula_cells))
        found_left = stream_find_cell_xlsx(str(fixture_path('complex_report.xlsx')), '总计', start='left')
        self.assertEqual(found_left['coord'], 'A12')
        self.assertLess(found_left['visited_cells'], 200)
        found_right = stream_find_cell_xlsx(str(fixture_path('complex_report.xlsx')), '备注', start='right')
        self.assertEqual(found_right['coord'], 'L6')
        self.assertLess(found_right['visited_cells'], found_left['visited_cells'])
        out = ROOT / 'stream_edit_out.xlsx'
        result = stream_update_first_match_xlsx(str(fixture_path('complex_report.xlsx')), str(out), '总计', '合计', start='left')
        self.assertTrue(result['updated'])
        self.assertEqual(result['coord'], 'A12')
        self.assertEqual(load_workbook(out)['经营驾驶舱']['A12'].value, '合计')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'stream-edit', str(fixture_path('complex_report.xlsx')), str(ROOT / 'stream_edit_cli.xlsx'), '--match', '总计', '--value', '合计'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertLess(json.loads(p.stdout)['visited_cells'], 200)
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'stream-edit', str(fixture_path('complex_report.xlsx')), str(ROOT / 'stream_edit_missing.xlsx'), '--match', 'NOT_FOUND', '--value', 'x', '--max-cells', '5'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertNotEqual(p.returncode, 0)
        self.assertEqual(json.loads(p.stdout)['stopped_reason'], 'max_cells')

    def test_alpha12_stream_edit_preview_all_and_offsets(self):
        from openpyxl import load_workbook
        from excel_ir_mvp.excel_ir import stream_find_cell_xlsx, stream_update_first_match_xlsx
        src = str(fixture_path('complex_report.xlsx'))
        preview_out = ROOT / 'stream_edit_preview_should_not_exist.xlsx'
        if preview_out.exists():
            preview_out.unlink()
        result = stream_update_first_match_xlsx(src, str(preview_out), '云业务', '云产品线', offset_col=1, preview=True)
        self.assertTrue(result['found'])
        self.assertFalse(result['updated'])
        self.assertTrue(result['preview'])
        self.assertEqual(result['target_coord'], 'B7')
        self.assertEqual(result['old_value'], '华东')
        self.assertFalse(preview_out.exists())
        out = ROOT / 'stream_edit_offset.xlsx'
        result = stream_update_first_match_xlsx(src, str(out), '云业务', '华东改', offset_col=1)
        self.assertTrue(result['updated'])
        self.assertEqual(load_workbook(out)['经营驾驶舱']['B7'].value, '华东改')
        self.assertEqual(load_workbook(out)['经营驾驶舱']['A7'].value, '云业务')
        out_all = ROOT / 'stream_edit_all.xlsx'
        result = stream_update_first_match_xlsx(src, str(out_all), '云业务', '云事业部', update_all=True)
        self.assertEqual(result['changed_count'], 3)
        ws = load_workbook(out_all)['经营驾驶舱']
        self.assertEqual(ws['A7'].value, '云事业部')
        self.assertEqual(ws['A8'].value, '云事业部')
        self.assertEqual(ws['B16'].value, '云事业部')
        found = stream_find_cell_xlsx(src, '业务线', offset_row=1, offset_col=2)
        self.assertEqual(found['coord'], 'A5')
        self.assertEqual(found['target']['coord'], 'C6')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'stream-edit', src, str(ROOT / 'stream_edit_cli_preview.xlsx'), '--match', '业务线', '--value', '收入本月', '--offset-row', '1', '--offset-col', '2', '--preview'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        data = json.loads(p.stdout)
        self.assertTrue(data['preview'])
        self.assertFalse(data['updated'])
        self.assertEqual(data['target_coord'], 'C6')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'stream-edit', src, str(ROOT / 'stream_edit_cli_all.xlsx'), '--match', '云业务', '--value', '云事业部', '--all'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(json.loads(p.stdout)['changed_count'], 3)
    def test_alpha13_backend_registry_and_engine_cli(self):
        from excel_ir_mvp import available_engines, engine_status
        from excel_ir_mvp.backends import BackendUnavailableError, resolve_engine
        from excel_ir_mvp.excel_ir_plus import inspect_workbook, parse_workbook_plus
        self.assertIn('openpyxl', available_engines())
        status = engine_status()
        self.assertTrue(status['openpyxl']['available'])
        self.assertIn('wolfxl', status)
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')), engine='openpyxl')
        self.assertEqual(ir['workbook']['engine']['engine'], 'openpyxl')
        overview = inspect_workbook(str(fixture_path('complex_report.xlsx')), engine='auto')
        self.assertEqual(overview['engine'], 'openpyxl')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'inspect', str(fixture_path('complex_report.xlsx')), '--engine', 'openpyxl'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(json.loads(p.stdout)['engine'], 'openpyxl')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'engines'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertIn('openpyxl', json.loads(p.stdout)['available'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'parse', str(fixture_path('complex_report.xlsx')), str(ROOT / 'alpha13_engine.ir.json'), '--engine', 'openpyxl'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(json.loads((ROOT / 'alpha13_engine.ir.json').read_text(encoding='utf-8'))['workbook']['engine']['engine'], 'openpyxl')
        if not status['wolfxl']['available']:
            with self.assertRaises(BackendUnavailableError):
                resolve_engine('wolfxl')
            p = subprocess.run(['python3', 'excel_ir_cli.py', 'parse', str(fixture_path('complex_report.xlsx')), str(ROOT / 'alpha13_wolfxl.ir.json'), '--engine', 'wolfxl'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            self.assertNotEqual(p.returncode, 0)
            self.assertIn('wolfxl', p.stderr + p.stdout)

    def test_alpha14_multi_header_locate_and_edit(self):
        from openpyxl import Workbook, load_workbook
        from excel_ir_mvp import multi_header_columns_xlsx, locate_cell_by_multi_header_xlsx, update_cell_by_multi_header_xlsx
        path = ROOT / 'tests' / 'fixtures' / 'multi_header_dates.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = '日期表'
        ws['A1'] = '项目'
        ws.merge_cells('B1:D1'); ws['B1'] = 2026
        ws.merge_cells('E1:G1'); ws['E1'] = 2027
        ws.merge_cells('B2:D2'); ws['B2'] = 5
        ws.merge_cells('E2:G2'); ws['E2'] = 5
        for idx, day in enumerate([7, 8, 9, 7, 8, 9], start=2):
            ws.cell(3, idx).value = day
        ws['A4'] = '门店A'; ws['B4'] = 10; ws['C4'] = 20; ws['D4'] = 30; ws['E4'] = 40; ws['F4'] = 50; ws['G4'] = 60
        ws['A5'] = '门店B'; ws['B5'] = 11; ws['C5'] = 21; ws['D5'] = 31; ws['E5'] = 41; ws['F5'] = 51; ws['G5'] = 61
        wb.save(path)
        cols = multi_header_columns_xlsx(str(path), sheet='日期表', header_start_row=1, header_end_row=3, min_col=2)['columns']
        c_col = [c for c in cols if c['letter'] == 'C'][0]
        self.assertEqual(c_col['path'], ['2026', '5', '8'])
        located = locate_cell_by_multi_header_xlsx(str(path), ['2026', '5', '8'], sheet='日期表', header_start_row=1, header_end_row=3, row_match='门店A')
        self.assertTrue(located['found'])
        self.assertEqual(located['target']['coord'], 'C4')
        preview = update_cell_by_multi_header_xlsx(str(path), str(ROOT / 'multi_header_preview.xlsx'), ['2026', '5', '8'], 999, sheet='日期表', header_start_row=1, header_end_row=3, row_match='门店A', preview=True)
        self.assertTrue(preview['preview'])
        self.assertFalse(preview['updated'])
        self.assertEqual(preview['target_coord'], 'C4')
        out = ROOT / 'multi_header_edit.xlsx'
        result = update_cell_by_multi_header_xlsx(str(path), str(out), ['2026', '5', '8'], 999, sheet='日期表', header_start_row=1, header_end_row=3, row_match='门店A')
        self.assertTrue(result['updated'])
        self.assertEqual(load_workbook(out)['日期表']['C4'].value, 999)
        self.assertEqual(load_workbook(out)['日期表']['F4'].value, 50)
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'header-edit', str(path), str(ROOT / 'multi_header_cli_preview.xlsx'), '--sheet', '日期表', '--headers', '2026/5/8', '--row-match', '门店A', '--value', '888', '--as-number', '--preview'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        cli_data = json.loads(p.stdout)
        self.assertTrue(cli_data['preview'])
        self.assertEqual(cli_data['target_coord'], 'C4')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'header-edit', str(path), str(ROOT / 'multi_header_cli_edit.xlsx'), '--sheet', '日期表', '--headers', '["2026","5","8"]', '--row-match', '门店A', '--value', '777', '--as-number'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(load_workbook(ROOT / 'multi_header_cli_edit.xlsx')['日期表']['C4'].value, 777)

    def test_alpha15_selective_sheets_and_compact_ir(self):
        from openpyxl import Workbook, load_workbook
        from excel_ir_mvp.excel_ir_plus import parse_workbook_plus, rebuild_workbook_plus
        path = ROOT / 'alpha15_multi_sheet.xlsx'
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Keep'
        ws1['A1'] = '保留'
        ws1['B1'] = '=1+1'
        ws2 = wb.create_sheet('Skip')
        ws2['A1'] = '跳过'
        ws2.row_dimensions[1].hidden = False
        ws2.column_dimensions['A'].hidden = False
        wb.save(path)
        full = parse_workbook_plus(str(path), infer_logic=False)
        selected = parse_workbook_plus(str(path), infer_logic=False, sheet_names=['Keep'])
        self.assertEqual([s['name'] for s in selected['workbook']['sheets']], ['Keep'])
        self.assertEqual(selected['workbook']['selected_sheets'], ['Keep'])
        self.assertLess(len(json.dumps(selected, ensure_ascii=False)), len(json.dumps(full, ensure_ascii=False)))
        cells = selected['workbook']['sheets'][0]['cells']
        self.assertIn('computed_value', cells['B1'])
        self.assertNotIn('computed_value_source', cells['B1'])
        style = selected['workbook']['styles'][cells['A1']['style_id']]
        self.assertNotEqual(style.get('protection'), {'locked': True, 'hidden': False})
        for sheet in selected['workbook']['sheets']:
            self.assertNotIn('Skip', sheet['name'])
        out = ROOT / 'alpha15_selected_rebuilt.xlsx'
        rebuild_workbook_plus(full, str(out), sheet_names=['Keep'])
        self.assertEqual(load_workbook(out).sheetnames, ['Keep'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'parse', str(path), str(ROOT / 'alpha15_cli_selected.ir.json'), '--sheet', 'Keep'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        parsed = json.loads((ROOT / 'alpha15_cli_selected.ir.json').read_text(encoding='utf-8'))
        self.assertEqual([s['name'] for s in parsed['workbook']['sheets']], ['Keep'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'rebuild', str(ROOT / 'alpha15_cli_selected.ir.json'), str(ROOT / 'alpha15_cli_selected.xlsx'), '--sheet', 'Keep'], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(load_workbook(ROOT / 'alpha15_cli_selected.xlsx').sheetnames, ['Keep'])
    def test_alpha16_public_api_facade(self):
        from openpyxl import load_workbook
        import excel_ir_mvp as xir
        src = fixture_path('complex_report.xlsx')
        ir = xir.parse(src, sheets='经营驾驶舱')
        self.assertEqual([s['name'] for s in ir['workbook']['sheets']], ['经营驾驶舱'])
        out = ROOT / 'alpha16_api_rebuilt.xlsx'
        xir.rebuild(ir, out, sheets='经营驾驶舱')
        self.assertEqual([s for s in load_workbook(out).sheetnames if not s.startswith('_excel_ir_metadata')], ['经营驾驶舱'])
        overview = xir.inspect(src)
        self.assertTrue(overview['ok'])
        self.assertIn('openpyxl', xir.engines()['available'])
        preview = xir.stream_edit(src, ROOT / 'alpha16_stream_preview.xlsx', match='业务线', value='收入本月', options=xir.StreamEditOptions(offset_row=1, offset_col=2, preview=True))
        self.assertFalse(preview['updated'])
        self.assertEqual(preview['target_coord'], 'C6')
        header_preview = xir.header_edit(fixture_path('multi_header_dates.xlsx'), ROOT / 'alpha16_header_preview.xlsx', headers=['2026', '5', '8'], value=123, options=xir.HeaderEditOptions(row_match='门店A', preview=True))
        self.assertTrue(header_preview['found'])
        self.assertEqual(header_preview['target_coord'], 'C4')


if __name__ == '__main__':
    unittest.main()

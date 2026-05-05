from __future__ import annotations

import json
import subprocess
import sys
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'src'
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from tests.fixtures_loader import fixture_path


class ExcelIRMetadataTests(unittest.TestCase):
    def test_hidden_metadata_sheet_roundtrip(self):
        from excel_ir_mvp.excel_ir_plus import (
            METADATA_SHEET_NAME,
            collect_semantic_metadata,
            diff_workbooks_plus,
            parse_workbook_plus,
            rebuild_workbook_plus,
            verify_metadata_checksum,
            verify_semantic_metadata,
            verify_semantic_metadata_xlsx,
        )
        src_ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        metadata = collect_semantic_metadata(src_ir)
        self.assertEqual(metadata['kind'], 'excel_ir_semantic_metadata')
        self.assertIn('checksum', metadata)
        self.assertTrue(verify_metadata_checksum(metadata))
        self.assertTrue(verify_semantic_metadata(metadata)['ok'])
        self.assertEqual(metadata['sheets'][0]['tables'][0]['table_kind'], 'semantic')
        out = ROOT / 'metadata_hidden_rebuilt.xlsx'
        rebuild_workbook_plus(src_ir, str(out))
        wb = load_workbook(out)
        self.assertTrue(verify_semantic_metadata_xlsx(str(out))['ok'])
        self.assertIn(METADATA_SHEET_NAME, wb.sheetnames)
        self.assertEqual(wb[METADATA_SHEET_NAME].sheet_state, 'veryHidden')
        reparsed = parse_workbook_plus(str(out))
        self.assertNotIn(METADATA_SHEET_NAME, [s['name'] for s in reparsed['workbook']['sheets']])
        table = reparsed['workbook']['sheets'][0]['extra']['tables'][0]
        self.assertEqual(table.get('table_kind'), 'semantic')
        self.assertIn('field_map_candidates', table.get('ir', {}))
        self.assertEqual(diff_workbooks_plus(str(fixture_path('complex_report.xlsx')), str(out))['diff_count'], 0)

    def test_metadata_export_import_api_and_cli(self):
        from excel_ir_mvp.excel_ir_plus import (
            apply_semantic_metadata,
            collect_semantic_metadata,
            export_semantic_metadata_from_ir,
            load_json,
            parse_workbook_plus,
            save_json,
            semantic_metadata_diff,
            verify_metadata_checksum,
            verify_semantic_metadata,
        )
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        metadata_path = ROOT / 'metadata_export_api.json'
        metadata = export_semantic_metadata_from_ir(ir, str(metadata_path))
        self.assertTrue(verify_metadata_checksum(metadata))
        self.assertEqual(metadata['sheets'][0]['tables'][0]['table_kind'], 'semantic')
        stripped = json.loads(json.dumps(ir, ensure_ascii=False))
        table = stripped['workbook']['sheets'][0]['extra']['tables'][0]
        table.pop('table_kind', None)
        table.pop('ir', None)
        apply_semantic_metadata(stripped, metadata)
        restored = stripped['workbook']['sheets'][0]['extra']['tables'][0]
        self.assertEqual(restored.get('table_kind'), 'semantic')
        self.assertIn('field_map_candidates', restored.get('ir', {}))
        stripped_path = ROOT / 'metadata_stripped.ir.json'
        imported_path = ROOT / 'metadata_imported.ir.json'
        cli_meta_path = ROOT / 'metadata_export_cli.json'
        save_json(stripped, str(stripped_path))
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'export', str(stripped_path), str(cli_meta_path)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertTrue(load_json(str(cli_meta_path))['sheets'])
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'import', str(stripped_path), str(metadata_path), str(imported_path)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        imported = load_json(str(imported_path))
        imported_table = imported['workbook']['sheets'][0]['extra']['tables'][0]
        self.assertEqual(imported_table.get('table_kind'), 'semantic')
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'verify', str(metadata_path)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertTrue(json.loads(p.stdout)['ok'])
        import contextlib, io
        from excel_ir_mvp.excel_ir_cli import main as cli_main
        saved_argv = sys.argv[:]
        buf = io.StringIO()
        try:
            sys.argv = ['excel-ir', 'metadata', 'verify', str(metadata_path)]
            with contextlib.redirect_stdout(buf):
                cli_main()
        finally:
            sys.argv = saved_argv
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'verify', '--from-xlsx', str(ROOT / 'metadata_hidden_rebuilt.xlsx')], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(json.loads(p.stdout)['source'], 'xlsx')
        extracted_path = ROOT / 'metadata_extracted_from_xlsx.json'
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'extract', str(extracted_path), '--from-xlsx', str(ROOT / 'metadata_hidden_rebuilt.xlsx')], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertTrue(load_json(str(extracted_path))['checksum'])
        self.assertEqual(semantic_metadata_diff(metadata, load_json(str(cli_meta_path)))['diff_count'], 0)
        diff_path = ROOT / 'metadata_cli_diff.json'
        p = subprocess.run(['python3', 'excel_ir_cli.py', 'metadata', 'diff', str(metadata_path), str(cli_meta_path), str(diff_path)], cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self.assertEqual(p.returncode, 0, msg=p.stderr + p.stdout)
        self.assertEqual(load_json(str(diff_path))['diff_count'], 0)

    def test_confirmed_field_map_is_persisted_to_metadata(self):
        from excel_ir_mvp.excel_ir_plus import collect_semantic_metadata
        from excel_ir_mvp.ir_patch import apply_patch_with_log
        from tests.fixtures_loader import load_json_fixture
        ir = load_json_fixture('complex_ir_v07.json')
        patch = load_json_fixture('v08_patch.json')
        out, log = apply_patch_with_log(ir, patch)
        self.assertTrue(log['ok'])
        metadata = collect_semantic_metadata(out)
        table = metadata['sheets'][0]['tables'][0]
        fmap = table.get('ir', {}).get('field_map', {})
        self.assertEqual(fmap.get('本月收入'), 'C')
        self.assertEqual(fmap.get('评级'), 'K')
        self.assertEqual(table.get('table_kind'), 'semantic')

    def test_corrupted_hidden_metadata_checksum_is_ignored(self):
        from excel_ir_mvp.excel_ir_plus import METADATA_CELL, METADATA_SHEET_NAME, parse_workbook_plus, rebuild_workbook_plus
        ir = parse_workbook_plus(str(fixture_path('complex_report.xlsx')))
        out = ROOT / 'metadata_corrupt_rebuilt.xlsx'
        rebuild_workbook_plus(ir, str(out))
        wb = load_workbook(out)
        raw = wb[METADATA_SHEET_NAME][METADATA_CELL].value
        data = json.loads(raw)
        data['sheets'][0]['tables'][0]['table_kind'] = 'native'
        # Keep the old checksum, making the payload invalid.
        wb[METADATA_SHEET_NAME][METADATA_CELL] = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        wb.save(out)
        reparsed = parse_workbook_plus(str(out))
        self.assertNotIn(METADATA_SHEET_NAME, [s['name'] for s in reparsed['workbook']['sheets']])
        self.assertFalse(reparsed['workbook']['sheets'][0].get('extra', {}).get('tables'))


if __name__ == '__main__':
    unittest.main()

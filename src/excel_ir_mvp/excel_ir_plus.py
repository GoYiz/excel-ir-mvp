from __future__ import annotations

import argparse
import base64
import copy
import hashlib
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

try:
    from . import excel_ir
    from .backends import available_engines, engine_status, resolve_engine, BackendUnavailableError
    from .excel_ir import (
        stream_find_cell_xlsx, stream_update_first_match_xlsx,
        multi_header_columns_xlsx, locate_cell_by_multi_header_xlsx, update_cell_by_multi_header_xlsx,
    )
except ImportError:  # flat-source dev fallback
    import excel_ir
    from backends import available_engines, engine_status, resolve_engine, BackendUnavailableError
    from excel_ir import (
        stream_find_cell_xlsx, stream_update_first_match_xlsx,
        multi_header_columns_xlsx, locate_cell_by_multi_header_xlsx, update_cell_by_multi_header_xlsx,
    )

def _load_workbook(path: str, *, engine: str | None = None, for_modify: bool = False, **kwargs: Any) -> Any:
    backend = resolve_engine(engine)
    if backend.name == "wolfxl" and for_modify:
        kwargs.setdefault("modify", True)
    return backend.module().load_workbook(path, **kwargs)


def _engine_info(engine: str | None = None) -> Dict[str, Any]:
    backend = resolve_engine(engine)
    status = engine_status().get(backend.name, {})
    return {"engine": backend.name, "engine_version": status.get("version")}


SCHEMA_VERSION = "0.7"
METADATA_SHEET_NAME = "_excel_ir_metadata"
METADATA_CELL = "A1"
METADATA_KIND = "excel_ir_semantic_metadata"
METADATA_VERSION = "2"
TABLE_METADATA_KEYS = [
    "displayName", "name", "ref", "comment", "tableType", "headerRowCount",
    "totalsRowCount", "totalsRowShown", "style", "ir", "table_kind",
    "native_table_supported", "native_table_skip_reason",
]


def _safe_get(obj: Any, attr: str) -> Any:
    try:
        return getattr(obj, attr)
    except Exception:
        return None


def _strip_none(x: Any) -> Any:
    if isinstance(x, dict):
        return {k: _strip_none(v) for k, v in x.items() if v is not None and v != {} and v != []}
    if isinstance(x, list):
        return [_strip_none(v) for v in x if v is not None]
    return x


def _attrs_to_dict(obj: Any, fields: List[str]) -> Dict[str, Any]:
    return _strip_none({f: _safe_get(obj, f) for f in fields})


def _omit_defaults_dict(data: Dict[str, Any], defaults: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in (data or {}).items() if v is not None and v != {} and v != [] and defaults.get(k, object()) != v}


def _sheet_filter(sheet_names: Optional[Iterable[str]]) -> Optional[set[str]]:
    if sheet_names is None:
        return None
    names = {str(s) for s in sheet_names if str(s)}
    return names or None


def _apply_attrs(obj: Any, data: Dict[str, Any], fields: List[str]) -> None:
    for f in fields:
        if f in data:
            try:
                setattr(obj, f, data[f])
            except Exception:
                pass


def dxf_to_dict(dxf: Optional[DifferentialStyle]) -> Dict[str, Any]:
    if not dxf:
        return {}
    data: Dict[str, Any] = {}
    if getattr(dxf, "font", None):
        data["font"] = excel_ir.font_to_dict(dxf.font)
    if getattr(dxf, "fill", None):
        data["fill"] = excel_ir.fill_to_dict(dxf.fill)
    if getattr(dxf, "border", None):
        data["border"] = excel_ir.border_to_dict(dxf.border)
    if getattr(dxf, "alignment", None):
        data["alignment"] = excel_ir.alignment_to_dict(dxf.alignment)
    if getattr(dxf, "protection", None):
        data["protection"] = excel_ir.protection_to_dict(dxf.protection)
    if getattr(dxf, "numFmt", None):
        data["numFmt"] = str(dxf.numFmt)
    return _strip_none(data)


def dxf_from_dict(data: Dict[str, Any]) -> Optional[DifferentialStyle]:
    if not data:
        return None
    kwargs: Dict[str, Any] = {}
    if "font" in data:
        kwargs["font"] = excel_ir.font_from_dict(data["font"])
    if "fill" in data:
        kwargs["fill"] = excel_ir.fill_from_dict(data["fill"])
    if "border" in data:
        kwargs["border"] = excel_ir.border_from_dict(data["border"])
    if "alignment" in data:
        kwargs["alignment"] = excel_ir.alignment_from_dict(data["alignment"])
    if "protection" in data:
        kwargs["protection"] = excel_ir.protection_from_dict(data["protection"])
    return DifferentialStyle(**kwargs)


def cfvo_to_dict(x: Any) -> Dict[str, Any]:
    return _strip_none({
        "type": _safe_get(x, "type"),
        "val": _safe_get(x, "val"),
        "gte": _safe_get(x, "gte"),
    })


def color_scale_to_dict(cs: Any) -> Dict[str, Any]:
    if not cs:
        return {}
    return _strip_none({
        "cfvo": [cfvo_to_dict(x) for x in getattr(cs, "cfvo", [])],
        "color": [excel_ir.color_to_dict(x) for x in getattr(cs, "color", [])],
    })


def rule_to_dict(rule: Rule) -> Dict[str, Any]:
    fields = [
        "type", "operator", "formula", "priority", "stopIfTrue", "aboveAverage",
        "percent", "bottom", "text", "timePeriod", "rank", "stdDev", "equalAverage",
    ]
    data = _attrs_to_dict(rule, fields)
    dxf = dxf_to_dict(getattr(rule, "dxf", None))
    if dxf:
        data["dxf"] = dxf
    cs = color_scale_to_dict(getattr(rule, "colorScale", None))
    if cs:
        data["colorScale"] = cs
    # MVP v0.2 records dataBar/iconSet metadata but does not recreate them generically yet.
    if getattr(rule, "dataBar", None):
        data["dataBar_present"] = True
    if getattr(rule, "iconSet", None):
        data["iconSet_present"] = True
    return _strip_none(data)


def rule_from_dict(data: Dict[str, Any]) -> Rule:
    kwargs: Dict[str, Any] = {"type": data.get("type")}
    for f in ["operator", "formula", "stopIfTrue", "aboveAverage", "percent", "bottom", "text", "timePeriod", "rank", "stdDev", "equalAverage"]:
        if f in data:
            kwargs[f] = data[f]
    dxf = dxf_from_dict(data.get("dxf", {}))
    if dxf:
        kwargs["dxf"] = dxf
    rule = Rule(**kwargs)
    if "priority" in data:
        try:
            rule.priority = data["priority"]
        except Exception:
            pass
    return rule


def parse_conditional_formatting(ws) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    try:
        cfs = list(ws.conditional_formatting)
    except Exception:
        cfs = []
    for cf in cfs:
        items.append(_strip_none({
            "sqref": str(getattr(cf, "sqref", "")),
            "rules": [rule_to_dict(r) for r in getattr(cf, "rules", [])],
        }))
    return items


def apply_conditional_formatting(ws, items: List[Dict[str, Any]]) -> None:
    for cf in items or []:
        sqref = cf.get("sqref")
        if not sqref:
            continue
        for rdata in cf.get("rules", []):
            if rdata.get("dataBar_present") or rdata.get("iconSet_present") or rdata.get("colorScale"):
                # Recorded for analysis; generic rebuild deferred to later versions.
                continue
            try:
                ws.conditional_formatting.add(sqref, rule_from_dict(rdata))
            except Exception:
                pass


def parse_data_validations(ws) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    dvs = getattr(getattr(ws, "data_validations", None), "dataValidation", []) or []
    fields = [
        "type", "operator", "formula1", "formula2", "allowBlank", "showDropDown",
        "showErrorMessage", "showInputMessage", "promptTitle", "errorStyle", "error",
        "prompt", "errorTitle", "imeMode",
    ]
    for dv in dvs:
        data = _attrs_to_dict(dv, fields)
        data["sqref"] = str(getattr(dv, "sqref", ""))
        result.append(_strip_none(data))
    return result


def apply_data_validations(ws, items: List[Dict[str, Any]]) -> None:
    fields = [
        "operator", "allowBlank", "showDropDown", "showErrorMessage", "showInputMessage",
        "promptTitle", "errorStyle", "error", "prompt", "errorTitle", "imeMode",
    ]
    for item in items or []:
        try:
            dv = DataValidation(type=item.get("type"), formula1=item.get("formula1"), formula2=item.get("formula2"))
            _apply_attrs(dv, item, fields)
            ws.add_data_validation(dv)
            if item.get("sqref"):
                dv.add(item["sqref"])
        except Exception:
            pass


def table_style_to_dict(ts: Optional[TableStyleInfo]) -> Dict[str, Any]:
    if not ts:
        return {}
    return _attrs_to_dict(ts, ["name", "showFirstColumn", "showLastColumn", "showRowStripes", "showColumnStripes"])


def table_style_from_dict(data: Dict[str, Any]) -> Optional[TableStyleInfo]:
    if not data:
        return None
    return TableStyleInfo(
        name=data.get("name"),
        showFirstColumn=data.get("showFirstColumn"),
        showLastColumn=data.get("showLastColumn"),
        showRowStripes=data.get("showRowStripes"),
        showColumnStripes=data.get("showColumnStripes"),
    )


def merged_parent_value(ws, row: int, col: int) -> Any:
    v = ws.cell(row, col).value
    if v not in (None, ""):
        return v
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return v


def guess_field_map_for_table(ws, ref: str, header_rows: Optional[int] = None) -> Dict[str, Any]:
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return {}
    if header_rows is None:
        header_rows = 2 if maxr - minr + 1 >= 3 else 1
    fmap: Dict[str, str] = {}
    confidence: Dict[str, float] = {}
    for c in range(minc, maxc + 1):
        parts = []
        for r in range(minr, min(minr + header_rows, maxr + 1)):
            v = merged_parent_value(ws, r, c)
            if v not in (None, ""):
                parts.append(str(v).strip())
        if parts:
            # Remove adjacent duplicates caused by vertical merged headers.
            dedup = []
            for p in parts:
                if not dedup or dedup[-1] != p:
                    dedup.append(p)
            parts = dedup
            letter = get_column_letter(c)
            full = "/".join(parts)
            fmap[full] = letter
            confidence[full] = 0.9
            leaf = parts[-1]
            if leaf not in fmap:
                fmap[leaf] = letter
                confidence[leaf] = 0.72 if len(parts) > 1 else 0.85
    return {"header_rows": header_rows, "field_map_candidates": fmap, "confidence": confidence}


def excel_table_native_status(ws, ref: str) -> Dict[str, Any]:
    """Return whether an IR table can be safely rebuilt as a native Excel Table.

    Excel native tables require a single, unmerged, non-empty string header row.
    Human reports often use merged / multi-row headers; those are kept as
    semantic IR tables but skipped during native OOXML table rebuild to avoid
    openpyxl's "column headings must be strings" warning and potentially
    unreadable files.
    """
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return {"native_table_supported": False, "native_table_skip_reason": "invalid_ref", "table_kind": "semantic"}
    for rng in ws.merged_cells.ranges:
        if not (rng.max_row < minr or rng.min_row > maxr or rng.max_col < minc or rng.min_col > maxc):
            return {"native_table_supported": False, "native_table_skip_reason": "merged_cells_intersect_table", "table_kind": "semantic"}
    headings: List[str] = []
    for c in range(minc, maxc + 1):
        v = ws.cell(minr, c).value
        if not isinstance(v, str) or not v.strip():
            return {"native_table_supported": False, "native_table_skip_reason": "non_string_or_blank_header", "table_kind": "semantic"}
        headings.append(v.strip())
    if len(set(headings)) != len(headings):
        return {"native_table_supported": False, "native_table_skip_reason": "duplicate_header", "table_kind": "semantic"}
    return {"native_table_supported": True, "table_kind": "native"}


def parse_tables(ws) -> List[Dict[str, Any]]:
    tables: List[Dict[str, Any]] = []
    try:
        values = list(ws.tables.values())
    except Exception:
        values = []
    for t in values:
        if not hasattr(t, "ref"):
            continue
        meta = excel_table_native_status(ws, getattr(t, "ref", None))
        tables.append(_strip_none({
            "displayName": getattr(t, "displayName", None),
            "name": getattr(t, "name", None),
            "ref": getattr(t, "ref", None),
            "comment": getattr(t, "comment", None),
            "tableType": getattr(t, "tableType", None),
            "headerRowCount": getattr(t, "headerRowCount", None),
            "totalsRowCount": getattr(t, "totalsRowCount", None),
            "totalsRowShown": getattr(t, "totalsRowShown", None),
            "style": table_style_to_dict(getattr(t, "tableStyleInfo", None)),
            "ir": guess_field_map_for_table(ws, getattr(t, "ref", None)),
            "native_table_supported": meta.get("native_table_supported"),
            "native_table_skip_reason": meta.get("native_table_skip_reason"),
            "table_kind": meta.get("table_kind"),
        }))
    return tables


def apply_tables(ws, items: List[Dict[str, Any]]) -> None:
    for item in items or []:
        if item.get("table_kind") == "semantic" or item.get("native_table_supported") is False:
            # Keep semantic table metadata in IR but do not rebuild as an Excel
            # native Table when the header violates OOXML table constraints.
            continue
        name = item.get("displayName") or item.get("name")
        ref = item.get("ref")
        if not name or not ref:
            continue
        try:
            tab = Table(displayName=name, ref=ref)
            for f in ["comment", "tableType", "headerRowCount", "totalsRowCount", "totalsRowShown"]:
                if f in item:
                    try:
                        setattr(tab, f, item[f])
                    except Exception:
                        pass
            tsi = table_style_from_dict(item.get("style", {}))
            if tsi:
                tab.tableStyleInfo = tsi
            ws.add_table(tab)
        except Exception:
            pass


def marker_to_dict(m: Any) -> Dict[str, Any]:
    if not m:
        return {}
    return _strip_none({
        "row": getattr(m, "row", None),
        "col": getattr(m, "col", None),
        "rowOff": getattr(m, "rowOff", None),
        "colOff": getattr(m, "colOff", None),
        "cell": f"{get_column_letter(getattr(m, 'col', 0) + 1)}{getattr(m, 'row', 0) + 1}",
    })


def parse_images(ws, include_binary: bool = True) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for idx, img in enumerate(getattr(ws, "_images", []) or [], start=1):
        data = b""
        if include_binary:
            try:
                data = img._data()
            except Exception:
                data = b""
        anchor = getattr(img, "anchor", None)
        anchor_data: Dict[str, Any] = {}
        if isinstance(anchor, str):
            anchor_data = {"type": "cell", "cell": anchor}
        elif anchor is not None:
            anchor_data = {
                "type": type(anchor).__name__,
                "from": marker_to_dict(getattr(anchor, "_from", None)),
                "to": marker_to_dict(getattr(anchor, "to", None)),
            }
            ext = getattr(anchor, "ext", None)
            if ext:
                anchor_data["ext"] = {"cx": getattr(ext, "cx", None), "cy": getattr(ext, "cy", None)}
        cell = anchor_data.get("cell") or anchor_data.get("from", {}).get("cell") or "A1"
        result.append(_strip_none({
            "name": f"image{idx}",
            "format": getattr(img, "format", None) or "png",
            "width": getattr(img, "width", None),
            "height": getattr(img, "height", None),
            "anchor": anchor_data,
            "anchor_cell": cell,
            "sha256": hashlib.sha256(data).hexdigest() if data else None,
            "data_b64": base64.b64encode(data).decode("ascii") if data else None,
        }))
    return result


def apply_images(ws, items: List[Dict[str, Any]], temp_paths: List[str]) -> None:
    for item in items or []:
        b64 = item.get("data_b64")
        if not b64:
            continue
        suffix = "." + (item.get("format") or "png")
        try:
            fd, tmp = tempfile.mkstemp(prefix="excel_ir_img_", suffix=suffix)
            with os.fdopen(fd, "wb") as f:
                f.write(base64.b64decode(b64))
            temp_paths.append(tmp)
            img = XLImage(tmp)
            if item.get("width"):
                img.width = item["width"]
            if item.get("height"):
                img.height = item["height"]
            cell = item.get("anchor_cell") or item.get("anchor", {}).get("cell") or item.get("anchor", {}).get("from", {}).get("cell") or "A1"
            ws.add_image(img, cell)
        except Exception:
            pass


def parse_header_footer(ws) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    for name in ["oddHeader", "oddFooter", "evenHeader", "evenFooter", "firstHeader", "firstFooter"]:
        obj = getattr(ws, name, None)
        if not obj:
            continue
        data = {}
        for pos in ["left", "center", "right"]:
            part = getattr(obj, pos, None)
            txt = getattr(part, "text", None) if part else None
            if txt:
                data[pos] = txt
        if data:
            result[name] = data
    return result


def apply_header_footer(ws, data: Dict[str, Any]) -> None:
    for name, parts in (data or {}).items():
        obj = getattr(ws, name, None)
        if not obj:
            continue
        for pos, txt in parts.items():
            part = getattr(obj, pos, None)
            if part:
                part.text = txt


def formula_ref_to_range(f: Optional[str]) -> Optional[Dict[str, Any]]:
    if not f:
        return None
    raw = str(f)
    if "!" in raw:
        sheet_part, ref = raw.split("!", 1)
        sheet = sheet_part.strip("'")
    else:
        sheet = None
        ref = raw
    ref = ref.replace("$", "")
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return {"formula": raw}
    return _strip_none({
        "sheet": sheet,
        "ref": f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{maxr}",
        "min_col": minc,
        "min_row": minr,
        "max_col": maxc,
        "max_row": maxr,
        "formula": raw,
    })


def chart_title_text(title: Any) -> Optional[str]:
    if not title:
        return None
    try:
        return title.tx.rich.p[0].r[0].t
    except Exception:
        return None


def parse_charts(ws) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []
    for idx, ch in enumerate(getattr(ws, "_charts", []) or [], start=1):
        series = []
        for s in getattr(ch, "ser", []) or []:
            val_f = None
            cat_f = None
            tx_f = None
            try:
                val_f = s.val.numRef.f if s.val and s.val.numRef else None
            except Exception:
                pass
            try:
                cat_f = s.cat.numRef.f if s.cat and s.cat.numRef else None
            except Exception:
                pass
            try:
                if not cat_f:
                    cat_f = s.cat.strRef.f if s.cat and s.cat.strRef else None
            except Exception:
                pass
            try:
                tx_f = s.tx.strRef.f if s.tx and s.tx.strRef else None
            except Exception:
                pass
            series.append(_strip_none({
                "title_ref": tx_f,
                "title": formula_ref_to_range(tx_f),
                "values": formula_ref_to_range(val_f),
                "categories": formula_ref_to_range(cat_f),
            }))
        anchor = getattr(ch, "anchor", None)
        if isinstance(anchor, str):
            anchor_cell = anchor
        else:
            anchor_cell = marker_to_dict(getattr(anchor, "_from", None)).get("cell", "A1")
        charts.append(_strip_none({
            "name": f"chart{idx}",
            "type": type(ch).__name__,
            "anchor": anchor_cell,
            "anchor_cell": anchor_cell,
            "title": chart_title_text(getattr(ch, "title", None)),
            "x_axis_title": chart_title_text(getattr(getattr(ch, "x_axis", None), "title", None)),
            "y_axis_title": chart_title_text(getattr(getattr(ch, "y_axis", None), "title", None)),
            "style": getattr(ch, "style", None),
            "height": getattr(ch, "height", None),
            "width": getattr(ch, "width", None),
            "series": series,
        }))
    return charts


def _range_to_reference(ws, r: Dict[str, Any], title: Dict[str, Any] | None = None) -> Optional[Reference]:
    if not r:
        return None
    try:
        min_col = int(r["min_col"])
        min_row = int(r["min_row"])
        max_col = int(r["max_col"])
        max_row = int(r["max_row"])
        # openpyxl expects the title cell to be included in the data reference
        # when titles_from_data=True.
        if title and title.get("min_col") == min_col and title.get("max_col") == max_col:
            trow = int(title.get("min_row"))
            if trow == min_row - 1:
                min_row = trow
        return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    except Exception:
        if r.get("ref"):
            try:
                minc, minr, maxc, maxr = range_boundaries(str(r["ref"]))
                return Reference(ws, min_col=minc, min_row=minr, max_col=maxc, max_row=maxr)
            except Exception:
                return None
    return None


def apply_charts(ws, items: List[Dict[str, Any]]) -> None:
    for item in items or []:
        typ = item.get("type")
        if typ == "LineChart":
            chart = LineChart()
        elif typ == "PieChart":
            chart = PieChart()
        else:
            chart = BarChart()
        if item.get("title"):
            chart.title = item["title"]
        if item.get("x_axis_title") and hasattr(chart, "x_axis"):
            chart.x_axis.title = item["x_axis_title"]
        if item.get("y_axis_title") and hasattr(chart, "y_axis"):
            chart.y_axis.title = item["y_axis_title"]
        if item.get("style") is not None:
            chart.style = item.get("style")
        if item.get("height"):
            chart.height = item["height"]
        if item.get("width"):
            chart.width = item["width"]
        # openpyxl may reset chart.height/width on load/save. Keep requested
        # values in IR for intent tracking; structural diff normalizes them.
        if item.get("height"):
            setattr(chart, "_excel_ir_height", item["height"])
        if item.get("width"):
            setattr(chart, "_excel_ir_width", item["width"])
        series = item.get("series", []) or []
        if series:
            # Group series with adjacent value columns where possible; if not,
            # add each one individually. This is sufficient for many generated
            # report charts.
            for s in series:
                val_ref = _range_to_reference(ws, s.get("values", {}), s.get("title", {}))
                if val_ref:
                    chart.add_data(val_ref, titles_from_data=bool(s.get("title_ref")))
            cat_ref = _range_to_reference(ws, series[0].get("categories", {}) if series else {})
            if cat_ref:
                try:
                    chart.set_categories(cat_ref)
                except Exception:
                    pass
        if item.get("anchor"):
            try:
                ws.add_chart(chart, item["anchor"])
            except Exception:
                ws.add_chart(chart, "A1")


def parse_sheet_extra(ws, *, include_images: bool = True, include_charts: bool = True, include_binary: bool = True) -> Dict[str, Any]:
    tab_color = excel_ir.color_to_dict(getattr(ws.sheet_properties, "tabColor", None))
    sheet_view = _omit_defaults_dict(_attrs_to_dict(ws.sheet_view, [
        "showFormulas", "showGridLines", "showRowColHeaders", "showZeros", "rightToLeft",
        "view", "topLeftCell", "zoomScale", "zoomScaleNormal",
    ]), {"showFormulas": False, "showGridLines": True, "showRowColHeaders": True, "showZeros": True, "rightToLeft": False, "view": "normal", "zoomScale": 100, "zoomScaleNormal": 100})
    sheet_format = _omit_defaults_dict(_attrs_to_dict(ws.sheet_format, [
        "baseColWidth", "defaultColWidth", "defaultRowHeight", "customHeight", "zeroHeight",
        "thickTop", "thickBottom", "outlineLevelRow", "outlineLevelCol",
    ]), {"baseColWidth": 8, "defaultRowHeight": 15, "customHeight": False, "zeroHeight": False, "thickTop": False, "thickBottom": False, "outlineLevelRow": 0, "outlineLevelCol": 0})
    page_margins = _attrs_to_dict(ws.page_margins, ["left", "right", "top", "bottom", "header", "footer"])
    page_setup = _attrs_to_dict(ws.page_setup, [
        "orientation", "paperSize", "scale", "fitToHeight", "fitToWidth", "firstPageNumber",
        "useFirstPageNumber", "paperHeight", "paperWidth", "pageOrder", "usePrinterDefaults",
        "blackAndWhite", "draft", "cellComments", "errors", "horizontalDpi", "verticalDpi", "copies",
    ])
    print_options = _omit_defaults_dict(_attrs_to_dict(ws.print_options, ["horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet"]), {"horizontalCentered": False, "verticalCentered": False, "headings": False, "gridLines": False, "gridLinesSet": True})
    protection = _attrs_to_dict(ws.protection, [
        "sheet", "objects", "scenarios", "formatCells", "formatColumns", "formatRows", "insertColumns",
        "insertRows", "insertHyperlinks", "deleteColumns", "deleteRows", "selectLockedCells",
        "selectUnlockedCells", "sort", "autoFilter", "pivotTables", "algorithmName", "saltValue",
        "spinCount", "hashValue",
    ])
    return _strip_none({
        "schema_version": SCHEMA_VERSION,
        "sheet_state": getattr(ws, "sheet_state", None) if getattr(ws, "sheet_state", None) != "visible" else None,
        "tab_color": tab_color,
        "sheet_view": sheet_view,
        "sheet_format": sheet_format,
        "page_margins": page_margins,
        "page_setup": page_setup,
        "print_options": print_options,
        "print_area": str(ws.print_area) if getattr(ws, "print_area", None) else None,
        "print_title_rows": str(ws.print_title_rows) if getattr(ws, "print_title_rows", None) else None,
        "print_title_cols": str(ws.print_title_cols) if getattr(ws, "print_title_cols", None) else None,
        "header_footer": parse_header_footer(ws),
        "protection": protection if protection.get("sheet") else {},
        "auto_filter": {"ref": ws.auto_filter.ref} if getattr(ws.auto_filter, "ref", None) else {},
        "data_validations": parse_data_validations(ws),
        "conditional_formatting": parse_conditional_formatting(ws),
        "tables": parse_tables(ws),
        "images": parse_images(ws, include_binary=include_binary) if include_images else [],
        "charts": parse_charts(ws) if include_charts else [],
    })


def apply_sheet_extra(ws, extra: Dict[str, Any], temp_paths: List[str]) -> None:
    if "sheet_state" in extra:
        ws.sheet_state = extra["sheet_state"]
    if extra.get("tab_color"):
        ws.sheet_properties.tabColor = excel_ir.color_from_dict(extra["tab_color"])
    _apply_attrs(ws.sheet_view, extra.get("sheet_view", {}), [
        "showFormulas", "showGridLines", "showRowColHeaders", "showZeros", "rightToLeft",
        "view", "topLeftCell", "zoomScale", "zoomScaleNormal",
    ])
    _apply_attrs(ws.sheet_format, extra.get("sheet_format", {}), [
        "baseColWidth", "defaultColWidth", "defaultRowHeight", "customHeight", "zeroHeight",
        "thickTop", "thickBottom", "outlineLevelRow", "outlineLevelCol",
    ])
    _apply_attrs(ws.page_margins, extra.get("page_margins", {}), ["left", "right", "top", "bottom", "header", "footer"])
    _apply_attrs(ws.page_setup, extra.get("page_setup", {}), [
        "orientation", "paperSize", "scale", "fitToHeight", "fitToWidth", "firstPageNumber",
        "useFirstPageNumber", "paperHeight", "paperWidth", "pageOrder", "usePrinterDefaults",
        "blackAndWhite", "draft", "cellComments", "errors", "horizontalDpi", "verticalDpi", "copies",
    ])
    _apply_attrs(ws.print_options, extra.get("print_options", {}), ["horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet"])
    if extra.get("print_area"):
        try:
            ws.print_area = extra["print_area"]
        except Exception:
            pass
    if extra.get("print_title_rows"):
        try:
            ws.print_title_rows = extra["print_title_rows"]
        except Exception:
            pass
    if extra.get("print_title_cols"):
        try:
            ws.print_title_cols = extra["print_title_cols"]
        except Exception:
            pass
    apply_header_footer(ws, extra.get("header_footer", {}))
    if extra.get("protection"):
        _apply_attrs(ws.protection, extra["protection"], list(extra["protection"].keys()))
    if extra.get("auto_filter", {}).get("ref"):
        ws.auto_filter.ref = extra["auto_filter"]["ref"]
    apply_data_validations(ws, extra.get("data_validations", []))
    apply_conditional_formatting(ws, extra.get("conditional_formatting", []))
    apply_tables(ws, extra.get("tables", []))
    apply_images(ws, extra.get("images", []), temp_paths)
    apply_charts(ws, extra.get("charts", []))


def _filter_ir_sheets(ir: Dict[str, Any], sheet_names: Optional[Iterable[str]]) -> Dict[str, Any]:
    requested = _sheet_filter(sheet_names)
    if requested is None:
        return ir
    out = copy.deepcopy(ir)
    out.get("workbook", {})["sheets"] = [s for s in out.get("workbook", {}).get("sheets", []) if s.get("name") in requested]
    out.get("workbook", {})["sheet_names"] = [s.get("name") for s in out.get("workbook", {}).get("sheets", [])]
    return out


def parse_workbook_plus(path: str, include_empty_styled: bool = True, infer_logic: bool = True, engine: str | None = None, sheet_names: Optional[Iterable[str]] = None, include_formula_cache: bool = True, include_extra: bool = True, include_images: bool = True, include_charts: bool = True, include_binary: bool = True, read_only: bool = False, sparse: bool = True, profile: str = "full") -> Dict[str, Any]:
    if profile not in {"full", "fast"}:
        raise ValueError("profile must be full or fast")
    if profile == "fast":
        include_empty_styled = False
        infer_logic = False
        include_formula_cache = False
        include_extra = False
        include_images = False
        include_charts = False
        include_binary = False
        read_only = True
        sparse = True
    ir = excel_ir.parse_workbook(path, include_empty_styled=include_empty_styled, infer_logic=infer_logic, engine=engine, sheet_names=sheet_names, include_formula_cache=include_formula_cache, sparse=sparse, read_only=read_only)
    ir["schema_version"] = SCHEMA_VERSION
    ir.get("workbook", {})["sheets"] = [
        s for s in ir.get("workbook", {}).get("sheets", [])
        if s.get("name") != METADATA_SHEET_NAME
    ]
    wb = _load_workbook(path, engine=engine, data_only=False) if include_extra else None
    requested_sheets = _sheet_filter(sheet_names)
    visible_worksheets = [ws for ws in wb.worksheets if ws.title != METADATA_SHEET_NAME and (requested_sheets is None or ws.title in requested_sheets)] if wb is not None else []
    for sheet_ir, ws in zip(ir["workbook"].get("sheets", []), visible_worksheets):
        sheet_ir["extra"] = parse_sheet_extra(ws, include_images=include_images, include_charts=include_charts, include_binary=include_binary)
    if wb is not None:
        apply_semantic_metadata(ir, read_semantic_metadata_sheet(wb))
    ir.setdefault("workbook", {}).setdefault("engine", _engine_info(engine))
    ir.setdefault("workbook", {})["parse_profile"] = profile
    if profile == "fast":
        ir.setdefault("workbook", {})["parse_warnings"] = ["fast profile skips formula cache, empty styled cells, logical inference, extended OOXML extras, images, charts and hidden semantic metadata; it uses read-only streaming mode where supported"]
    ir.setdefault("workbook", {})["sheet_names"] = [s.get("name") for s in ir.get("workbook", {}).get("sheets", [])]
    if requested_sheets is not None:
        ir.setdefault("workbook", {})["selected_sheets"] = [s for s in ir["workbook"]["sheet_names"]]
    return ir


def rebuild_workbook_plus(ir: Dict[str, Any], path: str, engine: str | None = None, sheet_names: Optional[Iterable[str]] = None) -> None:
    # First rebuild the reversible grid/style core using v0.1 engine.
    # Charts are applied later by v0.3; remove them from the first pass to avoid
    # duplicate charts when the temporary workbook is loaded again.
    requested_sheets = _sheet_filter(sheet_names)
    core_ir = _filter_ir_sheets(copy.deepcopy(ir), requested_sheets)
    core_ir.get("workbook", {})["sheets"] = [
        s for s in core_ir.get("workbook", {}).get("sheets", [])
        if s.get("name") != METADATA_SHEET_NAME
    ]
    for s in core_ir.get("workbook", {}).get("sheets", []):
        if s.get("extra"):
            s["extra"].pop("charts", None)
    excel_ir.rebuild_workbook(core_ir, path, engine=engine, sheet_names=requested_sheets)
    wb = _load_workbook(path, engine=engine, for_modify=True)
    temp_paths: List[str] = []
    try:
        selected_ir = _filter_ir_sheets(ir, requested_sheets)
        sheets_by_name = {ws.title: ws for ws in wb.worksheets}
        for sheet_ir in selected_ir["workbook"].get("sheets", []):
            if sheet_ir.get("name") == METADATA_SHEET_NAME:
                continue
            ws = sheets_by_name.get(sheet_ir.get("name"))
            if not ws:
                continue
            # v0.1 parsed comments but did not rebuild them; v0.2 does.
            for coord, cdata in sheet_ir.get("cells", {}).items():
                cmt = cdata.get("comment")
                if cmt:
                    ws[coord].comment = Comment(cmt.get("text", ""), cmt.get("author", ""))
            apply_sheet_extra(ws, sheet_ir.get("extra", {}), temp_paths)
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass
        write_semantic_metadata_sheet(wb, collect_semantic_metadata(selected_ir))
        wb.save(path)
    finally:
        for tmp in temp_paths:
            try:
                os.remove(tmp)
            except Exception:
                pass


def metadata_canonical_payload(metadata: Dict[str, Any]) -> Dict[str, Any]:
    data = copy.deepcopy(metadata or {})
    data.pop("checksum", None)
    return data


def metadata_checksum(metadata: Dict[str, Any]) -> str:
    raw = json.dumps(metadata_canonical_payload(metadata), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def attach_metadata_checksum(metadata: Dict[str, Any]) -> Dict[str, Any]:
    data = copy.deepcopy(metadata or {})
    data["checksum"] = metadata_checksum(data)
    return data


def verify_metadata_checksum(metadata: Dict[str, Any]) -> bool:
    if not metadata or not metadata.get("checksum"):
        return False
    return metadata.get("checksum") == metadata_checksum(metadata)


def verify_semantic_metadata(metadata: Dict[str, Any]) -> Dict[str, Any]:
    errors: List[Dict[str, Any]] = []
    if not isinstance(metadata, dict):
        return {"ok": False, "errors": [{"level": "error", "message": "metadata must be an object"}]}
    if metadata.get("kind") != METADATA_KIND:
        errors.append({"level": "error", "field": "kind", "message": f"expected {METADATA_KIND}"})
    if not metadata.get("version"):
        errors.append({"level": "error", "field": "version", "message": "missing metadata version"})
    if not isinstance(metadata.get("sheets", []), list):
        errors.append({"level": "error", "field": "sheets", "message": "sheets must be a list"})
    if not metadata.get("checksum"):
        errors.append({"level": "warning", "field": "checksum", "message": "checksum missing"})
    elif not verify_metadata_checksum(metadata):
        errors.append({"level": "error", "field": "checksum", "message": "checksum mismatch"})
    table_count = 0
    for si, sheet in enumerate(metadata.get("sheets", []) or []):
        if not sheet.get("name"):
            errors.append({"level": "warning", "field": f"sheets[{si}].name", "message": "sheet name missing"})
        for ti, table in enumerate(sheet.get("tables", []) or []):
            table_count += 1
            tk = table.get("table_kind")
            if tk and tk not in {"native", "semantic"}:
                errors.append({"level": "error", "field": f"sheets[{si}].tables[{ti}].table_kind", "message": "invalid table_kind"})
            if not (table.get("ref") or table.get("displayName") or table.get("name")):
                errors.append({"level": "warning", "field": f"sheets[{si}].tables[{ti}]", "message": "table identity missing"})
    return {"ok": not any(e.get("level") == "error" for e in errors), "errors": errors, "tables": table_count, "checksum_ok": bool(metadata.get("checksum") and verify_metadata_checksum(metadata))}


def verify_semantic_metadata_file(path: str) -> Dict[str, Any]:
    return verify_semantic_metadata(load_json(path))


def extract_semantic_metadata_from_xlsx(path: str, engine: str | None = None) -> Dict[str, Any]:
    wb = _load_workbook(path, engine=engine, data_only=False)
    metadata = read_semantic_metadata_sheet(wb)
    if metadata:
        return metadata
    return collect_semantic_metadata(parse_workbook_plus(path, engine=engine))


def verify_semantic_metadata_xlsx(path: str, engine: str | None = None) -> Dict[str, Any]:
    metadata = extract_semantic_metadata_from_xlsx(path, engine=engine)
    result = verify_semantic_metadata(metadata)
    result["source"] = "xlsx"
    return result


def metadata_status_xlsx(path: str, engine: str | None = None) -> Dict[str, Any]:
    """Return hidden metadata carrier status without mutating the workbook."""
    wb = _load_workbook(path, engine=engine, data_only=False)
    raw_present = METADATA_SHEET_NAME in wb.sheetnames
    raw = None
    sheet_state = None
    if raw_present:
        ws = wb[METADATA_SHEET_NAME]
        sheet_state = ws.sheet_state
        raw = ws[METADATA_CELL].value
    metadata = read_semantic_metadata_sheet(wb)
    parsed = bool(metadata)
    result = {
        "ok": True,
        "path": path,
        "engine": _engine_info(engine).get("engine"),
        "present": raw_present,
        "parsed": parsed,
        "sheet_state": sheet_state,
        "raw_bytes": len(raw.encode("utf-8")) if isinstance(raw, str) else 0,
        "checksum_ok": bool(metadata and verify_metadata_checksum(metadata)),
        "tables": sum(len(s.get("tables", [])) for s in metadata.get("sheets", [])) if metadata else 0,
    }
    if raw_present and not parsed:
        result["warning"] = "metadata sheet exists but payload is missing, invalid, or checksum-mismatched"
    return _strip_none(result)


def _anon_text(value: str, keep_formulas: bool = True) -> str:
    if keep_formulas and value.startswith("="):
        return value
    if re.fullmatch(r"[\w.+-]+@[\w.-]+", value):
        return "user@example.com"
    if re.fullmatch(r"\d{6,}", value):
        return "000000"
    # Preserve common structural labels that help table detection.
    keep = {"合计", "小计", "总计", "累计", "日期", "部门", "区域", "产品", "项目", "负责人", "评级"}
    if value in keep:
        return value
    return f"文本{len(value)}"


def anonymize_workbook_xlsx(src_path: str, out_path: str, keep_formulas: bool = True, engine: str | None = None) -> Dict[str, Any]:
    """Create a shareable workbook by replacing literal cell values.

    Styles, merges, dimensions, formulas (by default), validations, charts and
    other workbook structure are preserved as much as openpyxl allows. Hidden
    semantic metadata is stripped because it may contain source field labels.
    """
    wb = _load_workbook(src_path, engine=engine, for_modify=True, data_only=False)
    if METADATA_SHEET_NAME in wb.sheetnames:
        del wb[METADATA_SHEET_NAME]
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, bool):
                    continue
                if isinstance(value, (int, float)):
                    if not (isinstance(value, int) and value in (0, 1)):
                        cell.value = 0
                        changed += 1
                elif isinstance(value, str):
                    new_value = _anon_text(value, keep_formulas=keep_formulas)
                    if new_value != value:
                        cell.value = new_value
                        changed += 1
    wb.save(out_path)
    return {"ok": True, "source": src_path, "output": out_path, "engine": _engine_info(engine).get("engine"), "cells_changed": changed, "metadata_stripped": True}


def strip_semantic_metadata_xlsx(src_path: str, out_path: str, engine: str | None = None) -> Dict[str, Any]:
    """Copy an XLSX while removing the hidden semantic metadata carrier sheet."""
    wb = _load_workbook(src_path, engine=engine, for_modify=True, data_only=False)
    removed = False
    if METADATA_SHEET_NAME in wb.sheetnames:
        del wb[METADATA_SHEET_NAME]
        removed = True
    wb.save(out_path)
    return {"ok": True, "source": src_path, "output": out_path, "engine": _engine_info(engine).get("engine"), "removed": removed}


def repair_semantic_metadata_xlsx(src_path: str, out_path: str, engine: str | None = None) -> Dict[str, Any]:
    """Rewrite an XLSX with freshly collected semantic metadata embedded.

    This is intentionally conservative: it reparses the visible workbook, rebuilds
    it through the IR pipeline, and writes a new veryHidden metadata carrier with
    a fresh checksum. The source file is never modified in-place.
    """
    ir = parse_workbook_plus(src_path, engine=engine)
    rebuild_workbook_plus(ir, out_path, engine=engine)
    metadata = extract_semantic_metadata_from_xlsx(out_path, engine=engine)
    result = verify_semantic_metadata(metadata)
    result.update({
        "source": src_path,
        "output": out_path,
        "engine": _engine_info(engine).get("engine"),
        "repaired": bool(result.get("ok")),
    })
    return result


def inspect_workbook(path: str, engine: str | None = None) -> Dict[str, Any]:
    """Return a compact, JSON-serializable workbook overview for CLI/users."""
    wb = _load_workbook(path, engine=engine, data_only=False)
    hidden_metadata = read_semantic_metadata_sheet(wb)
    ir = parse_workbook_plus(path, engine=engine)
    sheets = []
    total_cells = 0
    total_tables = 0
    total_semantic = 0
    total_native = 0
    total_formulas = 0
    for sheet in ir.get("workbook", {}).get("sheets", []) or []:
        cells = sheet.get("cells", {}) or {}
        extra = sheet.get("extra", {}) or {}
        formulas = sum(1 for c in cells.values() if isinstance(c.get("value"), str) and c.get("value", "").startswith("="))
        tables = extra.get("tables", []) or []
        semantic = sum(1 for t in tables if t.get("table_kind") == "semantic")
        native = sum(1 for t in tables if t.get("table_kind") == "native")
        total_cells += len(cells)
        total_tables += len(tables)
        total_semantic += semantic
        total_native += native
        total_formulas += formulas
        sheets.append(_strip_none({
            "name": sheet.get("name"),
            "dimensions": sheet.get("dimensions"),
            "cell_count": len(cells),
            "formula_count": formulas,
            "merge_count": len(sheet.get("merged_cells", []) or []),
            "table_count": len(tables),
            "native_table_count": native,
            "semantic_table_count": semantic,
            "chart_count": len(extra.get("charts", []) or []),
            "image_count": len(extra.get("images", []) or []),
            "data_validation_count": len(extra.get("data_validations", []) or []),
            "conditional_format_count": len(extra.get("conditional_formatting", []) or []),
        }))
    metadata = collect_semantic_metadata(ir)
    return _strip_none({
        "ok": True,
        "path": path,
        "engine": _engine_info(engine).get("engine"),
        "sheet_count": len(sheets),
        "sheets": sheets,
        "totals": {
            "cells": total_cells,
            "formulas": total_formulas,
            "tables": total_tables,
            "native_tables": total_native,
            "semantic_tables": total_semantic,
        },
        "hidden_metadata": {
            "present": bool(hidden_metadata),
            "checksum_ok": bool(hidden_metadata and verify_metadata_checksum(hidden_metadata)),
            "tables": sum(len(s.get("tables", [])) for s in hidden_metadata.get("sheets", [])) if hidden_metadata else 0,
        },
        "collected_metadata": verify_semantic_metadata(metadata),
    })


def _metadata_table_item(table: Dict[str, Any]) -> Dict[str, Any]:
    item = {k: copy.deepcopy(table.get(k)) for k in TABLE_METADATA_KEYS if k in table}
    if "table_kind" not in item:
        ir_meta = item.get("ir", {}) or {}
        if int(ir_meta.get("header_rows", 1) or 1) > 1 or ir_meta.get("field_map") or ir_meta.get("field_map_candidates"):
            item["table_kind"] = "semantic"
    if item.get("table_kind") == "semantic" and "native_table_supported" not in item:
        item["native_table_supported"] = False
    return _strip_none(item)


def collect_semantic_metadata(ir: Dict[str, Any]) -> Dict[str, Any]:
    """Extract semantic metadata that should survive XLSX-only round-trips.

    The hidden sheet stores only semantic table intent / field maps, not the full
    reversible cell/style IR. This gives downstream systems a small metadata
    payload that can be exported, imported, or embedded in regenerated XLSX files.
    """
    sheets = []
    for sheet in ir.get("workbook", {}).get("sheets", []) or []:
        tables = []
        for table in sheet.get("extra", {}).get("tables", []) or []:
            item = _metadata_table_item(table)
            if item:
                tables.append(item)
        if tables:
            sheets.append({"name": sheet.get("name"), "tables": tables})
    metadata = _strip_none({
        "kind": METADATA_KIND,
        "version": METADATA_VERSION,
        "schema_version": SCHEMA_VERSION,
        "sheets": sheets,
    })
    return attach_metadata_checksum(metadata)


def apply_semantic_metadata(ir: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, Any]:
    """Merge exported / hidden semantic metadata into a workbook IR."""
    if not metadata or metadata.get("kind") != METADATA_KIND:
        return ir
    by_name = {s.get("name"): s for s in ir.get("workbook", {}).get("sheets", []) or []}
    for ms in metadata.get("sheets", []) or []:
        sheet = by_name.get(ms.get("name"))
        if not sheet:
            continue
        extra = sheet.setdefault("extra", {})
        tables = extra.setdefault("tables", [])
        by_ref = {t.get("ref"): t for t in tables if t.get("ref")}
        by_name_tbl = {t.get("displayName") or t.get("name"): t for t in tables if (t.get("displayName") or t.get("name"))}
        for mt in ms.get("tables", []) or []:
            target = by_ref.get(mt.get("ref")) or by_name_tbl.get(mt.get("displayName") or mt.get("name"))
            if target is None:
                target = {}
                tables.append(target)
            for k, v in mt.items():
                target[k] = copy.deepcopy(v)
    return ir


def write_semantic_metadata_sheet(wb, metadata: Dict[str, Any]) -> None:
    if METADATA_SHEET_NAME in wb.sheetnames:
        del wb[METADATA_SHEET_NAME]
    if not metadata.get("sheets"):
        return
    ws = wb.create_sheet(METADATA_SHEET_NAME)
    ws.sheet_state = "veryHidden"
    ws[METADATA_CELL] = json.dumps(metadata, ensure_ascii=False, separators=(",", ":"))


def read_semantic_metadata_sheet(wb) -> Dict[str, Any]:
    if METADATA_SHEET_NAME not in wb.sheetnames:
        return {}
    raw = wb[METADATA_SHEET_NAME][METADATA_CELL].value
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except Exception:
        return {}
    if data.get("kind") != METADATA_KIND:
        return {}
    if data.get("checksum") and not verify_metadata_checksum(data):
        return {}
    return data


def export_semantic_metadata_from_ir(ir: Dict[str, Any], path: str) -> Dict[str, Any]:
    metadata = collect_semantic_metadata(ir)
    save_json(metadata, path)
    return metadata


def import_semantic_metadata_to_ir(ir: Dict[str, Any], metadata: Dict[str, Any] | str) -> Dict[str, Any]:
    if isinstance(metadata, str):
        metadata = load_json(metadata)
    return apply_semantic_metadata(ir, metadata)


def semantic_metadata_diff(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
    aa = metadata_canonical_payload(a or {})
    bb = metadata_canonical_payload(b or {})
    diffs: List[Dict[str, Any]] = []
    if aa.get("kind") != bb.get("kind"):
        diffs.append({"type": "kind", "a": aa.get("kind"), "b": bb.get("kind")})
    if aa.get("version") != bb.get("version"):
        diffs.append({"type": "version", "a": aa.get("version"), "b": bb.get("version")})
    sheets_a = {s.get("name"): s for s in aa.get("sheets", []) or []}
    sheets_b = {s.get("name"): s for s in bb.get("sheets", []) or []}
    for sname in sorted(set(sheets_a) | set(sheets_b)):
        sa = sheets_a.get(sname)
        sb = sheets_b.get(sname)
        if sa is None or sb is None:
            diffs.append({"type": "sheet", "sheet": sname, "a": sa is not None, "b": sb is not None})
            continue
        ta = {t.get("ref") or t.get("displayName") or t.get("name") or str(i): t for i, t in enumerate(sa.get("tables", []) or [])}
        tb = {t.get("ref") or t.get("displayName") or t.get("name") or str(i): t for i, t in enumerate(sb.get("tables", []) or [])}
        for key in sorted(set(ta) | set(tb)):
            if ta.get(key) != tb.get(key):
                diffs.append({"type": "table", "sheet": sname, "table": key, "a": ta.get(key), "b": tb.get(key)})
                if len(diffs) >= 200:
                    return {"ok": False, "diff_count": 201, "diffs": diffs, "truncated": True}
    return {"ok": not diffs, "diff_count": len(diffs), "diffs": diffs[:200], "truncated": len(diffs) > 200}


def semantic_metadata_diff_files(a_path: str, b_path: str) -> Dict[str, Any]:
    return semantic_metadata_diff(load_json(a_path), load_json(b_path))


def _normalize_ir_for_diff(data: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize semantically equivalent openpyxl round-trip artifacts.

    Openpyxl sometimes changes empty inline strings to blank numeric cells, and
    header/footer text assigned through the API is not always preserved in a
    from-scratch intermediate load/save cycle. The diff tool should focus on
    report IR fidelity rather than these package-level incidental differences.
    """
    for s in data.get("workbook", {}).get("sheets", []):
        if s.get("name") == METADATA_SHEET_NAME:
            continue
        for coord, cell in list(s.get("cells", {}).items()):
            if cell.get("value") in (None, "") and cell.get("data_type") in ("inlineStr", "s", "str", "n"):
                cell.pop("data_type", None)
        extra = s.get("extra", {})
        extra.pop("header_footer", None)
        if extra.get("tables"):
            safe_tables = []
            for t in extra.get("tables", []) or []:
                if t.get("table_kind") == "semantic" or t.get("native_table_supported") is False:
                    # In v2.0 alpha, unsafe native tables (merged/multi-row headers)
                    # are preserved in parse IR for semantic patching but are not
                    # rebuilt as OOXML native tables. Exclude them from structural
                    # round-trip diff; the cell grid / styles / filters remain covered.
                    continue
                safe_tables.append(t)
            if safe_tables:
                extra["tables"] = safe_tables
            else:
                extra.pop("tables", None)
        for idx, img in enumerate(extra.get("images", []) or [], start=1):
            img.pop("data_b64", None)
            img["name"] = f"image{idx}"
        # Charts are now parsed and rebuilt in v0.3. Ignore generated internal
        # chart names only; compare chart type, anchor, titles, dimensions and ranges.
        for idx, ch in enumerate(extra.get("charts", []) or [], start=1):
            ch["name"] = f"chart{idx}"
            # openpyxl normalizes chart dimensions on reload; do not treat this
            # as a semantic diff in the MVP.
            ch.pop("height", None)
            ch.pop("width", None)
    return data


def canonical_for_diff(ir: Dict[str, Any]) -> Dict[str, Any]:
    data = copy.deepcopy(ir)
    data["schema_version"] = SCHEMA_VERSION
    data.get("workbook", {})["sheets"] = [
        s for s in data.get("workbook", {}).get("sheets", [])
        if s.get("name") != METADATA_SHEET_NAME
    ]
    # Logical inference is heuristic; exclude it from reversible structural diff.
    for s in data.get("workbook", {}).get("sheets", []):
        s.pop("logical", None)
    return _normalize_ir_for_diff(data)


def compare_ir(a: Dict[str, Any], b: Dict[str, Any], mode: str = "full") -> Dict[str, Any]:
    if mode not in {"full", "semantic", "structural"}:
        raise ValueError("mode must be full, semantic, or structural")
    if mode == "semantic":
        return semantic_metadata_diff(collect_semantic_metadata(a), collect_semantic_metadata(b))
    ia = canonical_for_diff(a)
    ib = canonical_for_diff(b)
    if mode == "structural":
        for data in (ia, ib):
            for sheet in data.get("workbook", {}).get("sheets", []) or []:
                extra = sheet.get("extra", {}) or {}
                extra.pop("tables", None)
    if ia == ib:
        return {"ok": True, "diff_count": 0, "diffs": [], "truncated": False}
    diffs: List[Dict[str, Any]] = []
    sheets_a = {s.get("name"): s for s in ia.get("workbook", {}).get("sheets", []) or []}
    sheets_b = {s.get("name"): s for s in ib.get("workbook", {}).get("sheets", []) or []}
    for sname in sorted(set(sheets_a) | set(sheets_b), key=lambda x: str(x)):
        sa = sheets_a.get(sname)
        sb = sheets_b.get(sname)
        if sa is None or sb is None:
            diffs.append({"sheet": sname, "type": "sheet_missing", "a": sa is not None, "b": sb is not None})
            continue
        for key in sorted(set(sa) | set(sb), key=str):
            if sa.get(key) != sb.get(key):
                if key == "cells":
                    ca = sa.get("cells", {}) or {}
                    cb = sb.get("cells", {}) or {}
                    for coord in sorted(set(ca) | set(cb)):
                        if ca.get(coord) != cb.get(coord):
                            diffs.append({"sheet": sname, "coord": coord, "type": "cell", "a": ca.get(coord), "b": cb.get(coord)})
                            if len(diffs) >= 200:
                                return {"ok": False, "diff_count": 201, "diffs": diffs, "truncated": True}
                else:
                    diffs.append({"sheet": sname, "type": key, "a": sa.get(key), "b": sb.get(key)})
                    if len(diffs) >= 200:
                        return {"ok": False, "diff_count": 201, "diffs": diffs, "truncated": True}
    return {"ok": not diffs, "diff_count": len(diffs), "diffs": diffs[:200], "truncated": len(diffs) > 200}


def compare_ir_files(a_path: str, b_path: str, mode: str = "full") -> Dict[str, Any]:
    return compare_ir(load_json(a_path), load_json(b_path), mode=mode)


def diff_workbooks_plus(a: str, b: str, engine: str | None = None) -> Dict[str, Any]:
    ia = canonical_for_diff(parse_workbook_plus(a, infer_logic=False, engine=engine))
    ib = canonical_for_diff(parse_workbook_plus(b, infer_logic=False, engine=engine))
    if ia == ib:
        return {"diff_count": 0, "diffs": [], "truncated": False}
    diffs: List[Dict[str, Any]] = []
    sheets_a = {s["name"]: s for s in ia["workbook"].get("sheets", [])}
    sheets_b = {s["name"]: s for s in ib["workbook"].get("sheets", [])}
    for sname in sorted(set(sheets_a) | set(sheets_b)):
        sa = sheets_a.get(sname)
        sb = sheets_b.get(sname)
        if sa is None or sb is None:
            diffs.append({"sheet": sname, "type": "sheet_missing", "a": sa is not None, "b": sb is not None})
            continue
        for key in sorted(set(sa) | set(sb)):
            va = copy.deepcopy(sa.get(key))
            vb = copy.deepcopy(sb.get(key))
            if key == "extra":
                va = _normalize_ir_for_diff({"workbook": {"sheets": [{"extra": va or {}, "cells": {}}]}})["workbook"]["sheets"][0]["extra"]
                vb = _normalize_ir_for_diff({"workbook": {"sheets": [{"extra": vb or {}, "cells": {}}]}})["workbook"]["sheets"][0]["extra"]
            if va != vb:
                if key == "cells":
                    ca = copy.deepcopy(sa.get("cells", {}))
                    cb = copy.deepcopy(sb.get("cells", {}))
                    _normalize_ir_for_diff({"workbook": {"sheets": [{"cells": ca, "extra": {}}]}})
                    _normalize_ir_for_diff({"workbook": {"sheets": [{"cells": cb, "extra": {}}]}})
                    for coord in sorted(set(ca) | set(cb)):
                        if ca.get(coord) != cb.get(coord):
                            diffs.append({"sheet": sname, "coord": coord, "type": "cell", "a": ca.get(coord), "b": cb.get(coord)})
                            if len(diffs) >= 200:
                                return {"diff_count": 201, "diffs": diffs, "truncated": True}
                else:
                    diffs.append({"sheet": sname, "type": key, "a": sa.get(key), "b": sb.get(key)})
                    if len(diffs) >= 200:
                        return {"diff_count": 201, "diffs": diffs, "truncated": True}
    return {"diff_count": len(diffs), "diffs": diffs[:200], "truncated": len(diffs) > 200}


def save_json(data: Dict[str, Any], path: str) -> None:
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def main() -> None:
    p = argparse.ArgumentParser(description="Excel report IR MVP v0.2: extended objects")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_parse = sub.add_parser("parse")
    p_parse.add_argument("input")
    p_parse.add_argument("output")
    p_parse.add_argument("--no-logic", action="store_true")
    p_parse.add_argument("--engine", default="openpyxl", choices=["openpyxl", "wolfxl", "auto"])
    p_parse.add_argument("--sheet", action="append", dest="sheets")
    p_parse.add_argument("--profile", choices=["full", "fast"], default="full")
    p_parse.add_argument("--fast", action="store_true")

    p_rebuild = sub.add_parser("rebuild")
    p_rebuild.add_argument("input_json")
    p_rebuild.add_argument("output_xlsx")
    p_rebuild.add_argument("--engine", default="openpyxl", choices=["openpyxl", "wolfxl", "auto"])
    p_rebuild.add_argument("--sheet", action="append", dest="sheets")

    p_diff = sub.add_parser("diff")
    p_diff.add_argument("a")
    p_diff.add_argument("b")
    p_diff.add_argument("output_json", nargs="?")
    p_diff.add_argument("--engine", default="openpyxl", choices=["openpyxl", "wolfxl", "auto"])

    p_logic = sub.add_parser("logic")
    p_logic.add_argument("input_json")

    args = p.parse_args()
    if args.cmd == "parse":
        profile = "fast" if args.fast else args.profile
        ir = parse_workbook_plus(args.input, infer_logic=not args.no_logic, engine=args.engine, sheet_names=args.sheets, profile=profile)
        save_json(ir, args.output)
        print(json.dumps({"ok": True, "schema_version": SCHEMA_VERSION, "sheets": len(ir["workbook"]["sheets"]), "styles": len(ir["workbook"]["styles"]), "output": args.output}, ensure_ascii=False))
    elif args.cmd == "rebuild":
        ir = load_json(args.input_json)
        rebuild_workbook_plus(ir, args.output_xlsx, engine=args.engine, sheet_names=args.sheets)
        print(json.dumps({"ok": True, "output": args.output_xlsx}, ensure_ascii=False))
    elif args.cmd == "diff":
        d = diff_workbooks_plus(args.a, args.b, engine=args.engine)
        if args.output_json:
            save_json(d, args.output_json)
        print(json.dumps(d, ensure_ascii=False, indent=2))
    elif args.cmd == "logic":
        ir = load_json(args.input_json)
        for s in ir["workbook"]["sheets"]:
            print(json.dumps({"sheet": s["name"], "logical": s.get("logical", {})}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

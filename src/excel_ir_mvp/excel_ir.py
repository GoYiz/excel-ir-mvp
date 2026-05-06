from __future__ import annotations

import argparse
import copy
import json
import math
import re
import shutil
from collections import Counter, deque
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, range_boundaries

SCHEMA_VERSION = "0.1"


def _safe_get(obj: Any, attr: str) -> Any:
    try:
        return getattr(obj, attr)
    except Exception:
        return None


def _strip_none(x: Any) -> Any:
    if isinstance(x, dict):
        return {k: _strip_none(v) for k, v in x.items() if v is not None and v != {} and v != []}
    if isinstance(x, list):
        return [_strip_none(v) for v in x if v is not None]
    return x


def color_to_dict(c: Optional[Color]) -> Optional[Dict[str, Any]]:
    if c is None:
        return None
    t = _safe_get(c, "type")
    d: Dict[str, Any] = {"type": t}
    if t == "rgb":
        d["rgb"] = _safe_get(c, "rgb")
    elif t == "theme":
        d["theme"] = _safe_get(c, "theme")
        tint = _safe_get(c, "tint")
        if tint not in (None, 0, 0.0):
            d["tint"] = tint
    elif t == "indexed":
        d["indexed"] = _safe_get(c, "indexed")
    elif t == "auto":
        d["auto"] = _safe_get(c, "auto")
    else:
        for a in ("rgb", "indexed", "theme", "tint", "auto"):
            v = _safe_get(c, a)
            if v is not None and not isinstance(v, str) or (isinstance(v, str) and "Values must be" not in v):
                d[a] = v
    return _strip_none(d)


def color_from_dict(d: Optional[Dict[str, Any]]) -> Optional[Color]:
    if not d:
        return None
    t = d.get("type")
    kwargs = {}
    if t == "rgb" and d.get("rgb"):
        kwargs["rgb"] = d.get("rgb")
    elif t == "theme" and d.get("theme") is not None:
        kwargs["theme"] = d.get("theme")
        kwargs["tint"] = d.get("tint", 0.0)
    elif t == "indexed" and d.get("indexed") is not None:
        kwargs["indexed"] = d.get("indexed")
    elif t == "auto":
        kwargs["auto"] = bool(d.get("auto", True))
    elif d.get("rgb"):
        kwargs["rgb"] = d.get("rgb")
    else:
        return None
    return Color(**kwargs)


def side_to_dict(s: Optional[Side]) -> Dict[str, Any]:
    if s is None:
        return {}
    return _strip_none({
        "style": _safe_get(s, "style"),
        "color": color_to_dict(_safe_get(s, "color")),
    })


def side_from_dict(d: Optional[Dict[str, Any]]) -> Side:
    if not d:
        return Side()
    return Side(style=d.get("style"), color=color_from_dict(d.get("color")))


def font_to_dict(f: Font) -> Dict[str, Any]:
    return _strip_none({
        "name": f.name,
        "sz": f.sz,
        "b": f.b,
        "i": f.i,
        "u": f.u,
        "strike": f.strike,
        "color": color_to_dict(f.color),
        "vertAlign": f.vertAlign,
        "charset": f.charset,
        "family": f.family,
        "scheme": f.scheme,
    })


def font_from_dict(d: Dict[str, Any]) -> Font:
    return Font(
        name=d.get("name"),
        sz=d.get("sz"),
        b=d.get("b"),
        i=d.get("i"),
        u=d.get("u"),
        strike=d.get("strike"),
        color=color_from_dict(d.get("color")),
        vertAlign=d.get("vertAlign"),
        charset=d.get("charset"),
        family=d.get("family"),
        scheme=d.get("scheme"),
    )


def fill_to_dict(f: PatternFill) -> Dict[str, Any]:
    return _strip_none({
        "fill_type": f.fill_type,
        "fgColor": color_to_dict(f.fgColor),
        "bgColor": color_to_dict(f.bgColor),
    })


def fill_from_dict(d: Dict[str, Any]) -> PatternFill:
    return PatternFill(
        fill_type=d.get("fill_type"),
        fgColor=color_from_dict(d.get("fgColor")),
        bgColor=color_from_dict(d.get("bgColor")),
    )


def border_to_dict(b: Border) -> Dict[str, Any]:
    return _strip_none({
        "left": side_to_dict(b.left),
        "right": side_to_dict(b.right),
        "top": side_to_dict(b.top),
        "bottom": side_to_dict(b.bottom),
        "diagonal": side_to_dict(b.diagonal),
        "diagonalUp": b.diagonalUp,
        "diagonalDown": b.diagonalDown,
        "outline": b.outline,
    })


def border_from_dict(d: Dict[str, Any]) -> Border:
    return Border(
        left=side_from_dict(d.get("left")),
        right=side_from_dict(d.get("right")),
        top=side_from_dict(d.get("top")),
        bottom=side_from_dict(d.get("bottom")),
        diagonal=side_from_dict(d.get("diagonal")),
        diagonalUp=d.get("diagonalUp", False),
        diagonalDown=d.get("diagonalDown", False),
        outline=d.get("outline", True),
    )


def alignment_to_dict(a: Alignment) -> Dict[str, Any]:
    return _strip_none({
        "horizontal": a.horizontal,
        "vertical": a.vertical,
        "text_rotation": a.textRotation,
        "wrap_text": a.wrap_text,
        "shrink_to_fit": a.shrink_to_fit,
        "indent": a.indent,
    })


def alignment_from_dict(d: Dict[str, Any]) -> Alignment:
    return Alignment(
        horizontal=d.get("horizontal"),
        vertical=d.get("vertical"),
        text_rotation=d.get("text_rotation", 0),
        wrap_text=d.get("wrap_text"),
        shrink_to_fit=d.get("shrink_to_fit"),
        indent=d.get("indent", 0),
    )


def protection_to_dict(p: Protection) -> Dict[str, Any]:
    return _strip_none({"locked": p.locked, "hidden": p.hidden})


def protection_from_dict(d: Dict[str, Any]) -> Protection:
    return Protection(locked=d.get("locked", True), hidden=d.get("hidden", False))


def style_to_dict(cell) -> Dict[str, Any]:
    return _strip_none({
        "font": font_to_dict(cell.font),
        "fill": fill_to_dict(cell.fill),
        "border": border_to_dict(cell.border),
        "alignment": alignment_to_dict(cell.alignment),
        "number_format": cell.number_format,
        "protection": protection_to_dict(cell.protection),
    })


def apply_style(cell, style: Dict[str, Any]) -> None:
    if "font" in style:
        cell.font = font_from_dict(style.get("font", {}))
    if "fill" in style:
        cell.fill = fill_from_dict(style.get("fill", {}))
    if "border" in style:
        cell.border = border_from_dict(style.get("border", {}))
    if "alignment" in style:
        cell.alignment = alignment_from_dict(style.get("alignment", {}))
    if "number_format" in style:
        cell.number_format = style.get("number_format") or "General"
    if "protection" in style:
        cell.protection = protection_from_dict(style.get("protection", {}))


def stable_json_key(d: Dict[str, Any]) -> str:
    return json.dumps(d, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


def normalized_value(v: Any) -> Any:
    if hasattr(v, "isoformat"):
        try:
            return {"__type__": type(v).__name__, "iso": v.isoformat()}
        except Exception:
            return str(v)
    return v


def denormalize_value(v: Any) -> Any:
    # MVP: keep date-like values as strings unless openpyxl already loads them from a rebuilt file.
    # This avoids guessing date/datetime classes incorrectly.
    if isinstance(v, dict) and "iso" in v:
        return v.get("iso")
    return v


def is_cell_material(cell) -> bool:
    if isinstance(cell, MergedCell):
        return False
    if cell.value is not None:
        return True
    # Empty cells with explicit style matter for templates.
    if cell.has_style:
        return True
    return False


def _stream_match(value: Any, match: Any, contains: bool = False, case_sensitive: bool = False) -> bool:
    if isinstance(value, dict):
        value = value.get("iso")
    if contains:
        left = "" if value is None else str(value)
        right = "" if match is None else str(match)
        if not case_sensitive:
            left, right = left.lower(), right.lower()
        return right in left
    if isinstance(value, str) and isinstance(match, str) and not case_sensitive:
        return value.lower() == match.lower()
    return value == match


def _stream_cell_info(cell: Any, *, match_value: Any = None, visited: Optional[int] = None) -> Dict[str, Any]:
    info = {
        "sheet": cell.parent.title,
        "coord": cell.coordinate,
        "row": cell.row,
        "col": cell.column,
        "value": normalized_value(cell.value),
    }
    if match_value is not None:
        info["match_value"] = normalized_value(match_value)
    if visited is not None:
        info["visited_cells"] = visited
    return info


def _stream_scan_cells(path: str, match: Any, *, sheet: Optional[str] = None, start: str = "left", contains: bool = False, case_sensitive: bool = False, max_cells: Optional[int] = None, stop_after_first: bool = True, on_match: Optional[Callable[[Any, int], Optional[Dict[str, Any]]]] = None) -> Dict[str, Any]:
    if start not in {"left", "right"}:
        raise ValueError("start must be left or right")
    wb = load_workbook(path, data_only=False)
    sheets = [wb[sheet]] if sheet else wb.worksheets
    visited = 0
    matches: List[Dict[str, Any]] = []
    for ws in sheets:
        for r in range(1, ws.max_row + 1):
            cols = range(1, ws.max_column + 1) if start == "left" else range(ws.max_column, 0, -1)
            for c in cols:
                visited += 1
                if max_cells is not None and visited > max_cells:
                    return {"ok": False, "found": bool(matches), "matches": matches, "visited_cells": visited - 1, "stopped_reason": "max_cells", "workbook": wb}
                cell = ws.cell(r, c)
                if _stream_match(cell.value, match, contains=contains, case_sensitive=case_sensitive):
                    info = _stream_cell_info(cell, visited=visited)
                    if on_match:
                        extra = on_match(cell, visited)
                        if extra:
                            info.update(extra)
                    matches.append(info)
                    if stop_after_first:
                        return {"ok": True, "found": True, "matches": matches, "visited_cells": visited, "stopped_reason": "found", "workbook": wb}
    return {"ok": True, "found": bool(matches), "matches": matches, "visited_cells": visited, "stopped_reason": "end", "workbook": wb}


def stream_find_cell_xlsx(path: str, match: Any, *, sheet: Optional[str] = None, start: str = "left", contains: bool = False, case_sensitive: bool = False, max_cells: Optional[int] = None, offset_row: int = 0, offset_col: int = 0) -> Dict[str, Any]:
    """Scan a workbook like a human and stop as soon as a cell matches.

    This intentionally avoids building the full IR. `start='left'` scans each row
    left-to-right; `start='right'` scans each row right-to-left. Optional offsets
    return the target cell relative to the matched anchor.
    """
    def target_info(cell: Any, visited: int) -> Dict[str, Any]:
        if not offset_row and not offset_col:
            return {}
        tr = cell.row + offset_row
        tc = cell.column + offset_col
        if tr < 1 or tc < 1:
            raise ValueError("offset target is outside worksheet bounds")
        target = cell.parent.cell(tr, tc)
        return {"anchor": _stream_cell_info(cell), "target": _stream_cell_info(target, match_value=cell.value)}

    result = _stream_scan_cells(path, match, sheet=sheet, start=start, contains=contains, case_sensitive=case_sensitive, max_cells=max_cells, stop_after_first=True, on_match=target_info)
    wb = result.pop("workbook")
    if not result.get("matches"):
        return result
    first = result.pop("matches")[0]
    result.update(first)
    return result


def stream_update_first_match_xlsx(src_path: str, out_path: str, match: Any, new_value: Any, *, sheet: Optional[str] = None, start: str = "left", contains: bool = False, case_sensitive: bool = False, max_cells: Optional[int] = None, offset_row: int = 0, offset_col: int = 0, preview: bool = False, update_all: bool = False) -> Dict[str, Any]:
    """Find matching cells using streaming scan semantics, then update target cells."""
    changes: List[Dict[str, Any]] = []

    def plan_change(cell: Any, visited: int) -> Dict[str, Any]:
        target = cell.parent.cell(cell.row + offset_row, cell.column + offset_col)
        before = target.value
        change = {
            "anchor": _stream_cell_info(cell, visited=visited),
            "target": _stream_cell_info(target, match_value=cell.value),
            "old_value": normalized_value(before),
            "new_value": normalized_value(new_value),
        }
        changes.append({"cell": target, "old_value": before, "change": change})
        return {"target": change["target"]} if (offset_row or offset_col) else {}

    result = _stream_scan_cells(src_path, match, sheet=sheet, start=start, contains=contains, case_sensitive=case_sensitive, max_cells=max_cells, stop_after_first=not update_all, on_match=plan_change)
    wb = result.pop("workbook")
    if not changes:
        if (not preview) and src_path != out_path:
            shutil.copyfile(src_path, out_path)
        result.update({"updated": False, "preview": preview, "output": out_path, "changes": []})
        return result
    if not preview:
        for item in changes:
            item["cell"].value = denormalize_value(new_value)
        wb.save(out_path)
    result.update({
        "updated": not preview,
        "preview": preview,
        "output": None if preview else out_path,
        "changed_count": len(changes),
        "changes": [item["change"] for item in changes],
    })
    if len(changes) == 1:
        change = changes[0]["change"]
        anchor = change["anchor"]
        target = change["target"]
        result.update({
            "sheet": anchor["sheet"],
            "coord": anchor["coord"],
            "row": anchor["row"],
            "col": anchor["col"],
            "value": anchor["value"],
            "target_sheet": target["sheet"],
            "target_coord": target["coord"],
            "old_value": change["old_value"],
            "new_value": change["new_value"],
        })
    return result


def parse_workbook(path: str, include_empty_styled: bool = True, infer_logic: bool = True) -> Dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    data_wb = load_workbook(path, data_only=True)
    styles: Dict[str, Dict[str, Any]] = {}
    style_ids: Dict[str, str] = {}
    sheets: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        data_ws = data_wb[ws.title] if ws.title in data_wb.sheetnames else None
        cells: Dict[str, Dict[str, Any]] = {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is None and not (include_empty_styled and cell.has_style):
                    continue
                sdict = style_to_dict(cell)
                skey = stable_json_key(sdict)
                if skey not in style_ids:
                    sid = f"s{len(style_ids) + 1:04d}"
                    style_ids[skey] = sid
                    styles[sid] = sdict
                entry = {
                    "row": cell.row,
                    "col": cell.column,
                    "value": normalized_value(cell.value),
                    "data_type": cell.data_type,
                    "style_id": style_ids[skey],
                }
                if cell.data_type == "f" and data_ws is not None:
                    entry["computed_value"] = normalized_value(data_ws[cell.coordinate].value)
                    entry["computed_value_source"] = "xlsx_cached_value"
                if cell.hyperlink:
                    entry["hyperlink"] = cell.hyperlink.target
                if cell.comment:
                    entry["comment"] = {"text": cell.comment.text, "author": cell.comment.author}
                out_entry = _strip_none(entry)
                if cell.data_type == "f" and "computed_value" not in out_entry:
                    out_entry["computed_value"] = entry.get("computed_value")
                if cell.data_type == "f" and "computed_value_source" not in out_entry:
                    out_entry["computed_value_source"] = entry.get("computed_value_source") or "xlsx_cached_value"
                cells[cell.coordinate] = out_entry

        rows: Dict[str, Dict[str, Any]] = {}
        for idx, dim in ws.row_dimensions.items():
            d = _strip_none({"height": dim.height, "hidden": dim.hidden, "outlineLevel": dim.outlineLevel})
            if d:
                rows[str(idx)] = d

        cols: Dict[str, Dict[str, Any]] = {}
        for key, dim in ws.column_dimensions.items():
            d = _strip_none({"width": dim.width, "hidden": dim.hidden, "outlineLevel": dim.outlineLevel})
            if d:
                cols[str(key)] = d

        sheet = {
            "name": ws.title,
            "dimensions": {"max_row": ws.max_row, "max_col": ws.max_column},
            "freeze_panes": ws.freeze_panes,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "rows": rows,
            "cols": cols,
            "cells": cells,
            "logical": {"regions": []},
        }
        if infer_logic:
            sheet["logical"] = infer_sheet_logic(sheet, styles)
        sheets.append(sheet)

    return {"schema_version": SCHEMA_VERSION, "workbook": {"sheets": sheets, "styles": styles}}


def rebuild_workbook(ir: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    styles = ir["workbook"].get("styles", {})

    for sheet in ir["workbook"].get("sheets", []):
        ws = wb.create_sheet(sheet["name"])

        # Layout before cells.
        for r, d in sheet.get("rows", {}).items():
            rd = ws.row_dimensions[int(r)]
            if "height" in d:
                rd.height = d["height"]
            if "hidden" in d:
                rd.hidden = d["hidden"]
            if "outlineLevel" in d:
                rd.outlineLevel = d["outlineLevel"]

        for c, d in sheet.get("cols", {}).items():
            cd = ws.column_dimensions[c]
            if "width" in d:
                cd.width = d["width"]
            if "hidden" in d:
                cd.hidden = d["hidden"]
            if "outlineLevel" in d:
                cd.outlineLevel = d["outlineLevel"]

        for coord, cdata in sheet.get("cells", {}).items():
            cell = ws[coord]
            cell.value = denormalize_value(cdata.get("value"))
            sid = cdata.get("style_id")
            if sid and sid in styles:
                apply_style(cell, styles[sid])
            if cdata.get("hyperlink"):
                cell.hyperlink = cdata["hyperlink"]
            # Comment intentionally skipped in MVP rebuild unless later needed.

        for mr in sheet.get("merged_ranges", []):
            ws.merge_cells(mr)

        if sheet.get("freeze_panes"):
            ws.freeze_panes = sheet.get("freeze_panes")

    wb.save(path)


def cell_text(sheet: Dict[str, Any], coord: str) -> str:
    v = sheet.get("cells", {}).get(coord, {}).get("value")
    if v is None:
        return ""
    if isinstance(v, dict):
        return str(v.get("iso", ""))
    return str(v)


def cell_style(sheet: Dict[str, Any], styles: Dict[str, Any], coord: str) -> Dict[str, Any]:
    sid = sheet.get("cells", {}).get(coord, {}).get("style_id")
    return styles.get(sid, {}) if sid else {}


def coord_of(r: int, c: int) -> str:
    return f"{get_column_letter(c)}{r}"


def non_empty_matrix(sheet: Dict[str, Any]) -> List[List[bool]]:
    max_row = sheet["dimensions"]["max_row"]
    max_col = sheet["dimensions"]["max_col"]
    m = [[False] * (max_col + 1) for _ in range(max_row + 1)]
    for coord, cdata in sheet.get("cells", {}).items():
        v = cdata.get("value")
        if v is not None and str(v) != "":
            m[cdata["row"]][cdata["col"]] = True
    return m


def find_components(sheet: Dict[str, Any]) -> List[Tuple[int, int, int, int, int]]:
    m = non_empty_matrix(sheet)
    max_row = sheet["dimensions"]["max_row"]
    max_col = sheet["dimensions"]["max_col"]
    seen = [[False] * (max_col + 1) for _ in range(max_row + 1)]
    comps = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if not m[r][c] or seen[r][c]:
                continue
            q = deque([(r, c)])
            seen[r][c] = True
            minr = maxr = r
            minc = maxc = c
            count = 0
            while q:
                rr, cc = q.popleft()
                count += 1
                minr, maxr = min(minr, rr), max(maxr, rr)
                minc, maxc = min(minc, cc), max(maxc, cc)
                for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    nr, nc = rr + dr, cc + dc
                    if 1 <= nr <= max_row and 1 <= nc <= max_col and m[nr][nc] and not seen[nr][nc]:
                        seen[nr][nc] = True
                        q.append((nr, nc))
            comps.append((minr, minc, maxr, maxc, count))
    return comps


def has_border(style: Dict[str, Any]) -> bool:
    b = style.get("border", {})
    for side in ("left", "right", "top", "bottom"):
        if b.get(side, {}).get("style"):
            return True
    return False


def side_border_count(style: Dict[str, Any]) -> int:
    b = style.get("border", {})
    return sum(1 for side in ("left", "right", "top", "bottom") if b.get(side, {}).get("style"))


def font_size(style: Dict[str, Any]) -> float:
    try:
        return float(style.get("font", {}).get("sz") or 11)
    except Exception:
        return 11.0


def is_bold(style: Dict[str, Any]) -> bool:
    return bool(style.get("font", {}).get("b"))


def horizontal(style: Dict[str, Any]) -> Optional[str]:
    return style.get("alignment", {}).get("horizontal")


def infer_title(sheet: Dict[str, Any], styles: Dict[str, Any]) -> List[Dict[str, Any]]:
    regions = []
    top_limit = min(sheet["dimensions"]["max_row"], 6)
    merged = sheet.get("merged_ranges", [])
    merged_by_top_left = {}
    for rng in merged:
        minc, minr, maxc, maxr = range_boundaries(rng)
        merged_by_top_left[(minr, minc)] = (minr, minc, maxr, maxc, rng)
    for r in range(1, top_limit + 1):
        row_values = []
        for c in range(1, sheet["dimensions"]["max_col"] + 1):
            coord = coord_of(r, c)
            txt = cell_text(sheet, coord).strip()
            if txt:
                st = cell_style(sheet, styles, coord)
                row_values.append((coord, txt, st, c))
        if not row_values:
            continue
        # Prefer a single prominent merged/bold/large centered text.
        if len(row_values) <= 2:
            coord, txt, st, c = row_values[0]
            merged_range = None
            if (r, c) in merged_by_top_left:
                merged_range = merged_by_top_left[(r, c)][4]
            score = 0
            if font_size(st) >= 14:
                score += 2
            if is_bold(st):
                score += 1
            if horizontal(st) in ("center", "centerContinuous"):
                score += 1
            if merged_range:
                score += 2
            if score >= 3:
                regions.append({
                    "type": "title",
                    "range": merged_range or coord,
                    "text": txt,
                    "confidence": min(0.99, 0.55 + 0.1 * score),
                    "evidence": {"font_size": font_size(st), "bold": is_bold(st), "merged": bool(merged_range), "horizontal": horizontal(st)},
                })
                break
    return regions


def infer_metadata(sheet: Dict[str, Any], styles: Dict[str, Any]) -> List[Dict[str, Any]]:
    regions = []
    keywords = ("单位", "日期", "制表", "编制", "审核", "部门", "期间", "报表", "币种")
    max_row = min(sheet["dimensions"]["max_row"], 10)
    max_col = sheet["dimensions"]["max_col"]
    for r in range(1, max_row + 1):
        row_hits = []
        for c in range(1, max_col + 1):
            coord = coord_of(r, c)
            txt = cell_text(sheet, coord).strip()
            if not txt:
                continue
            if any(k in txt for k in keywords) or ":" in txt or "：" in txt:
                row_hits.append({"coord": coord, "text": txt})
        if row_hits:
            cols = [int(re.findall(r"\d+", h["coord"])[0]) for h in row_hits]
            regions.append({
                "type": "metadata_row",
                "range": f"A{r}:{get_column_letter(max_col)}{r}",
                "items": row_hits,
                "confidence": 0.65,
            })
    return regions


def infer_tables(sheet: Dict[str, Any], styles: Dict[str, Any]) -> List[Dict[str, Any]]:
    # First find broad connected components by non-empty cells. Human section captions
    # often touch a bordered table in the next row, so a component may contain both a
    # section title and the actual grid. We therefore trim every component to the
    # strongest rectangular bordered grid before classifying it as a table.
    comps = find_components(sheet)
    regions = []
    for comp_minr, comp_minc, comp_maxr, comp_maxc, count in comps:
        if comp_maxr - comp_minr + 1 < 2 or comp_maxc - comp_minc + 1 < 2 or count < 4:
            continue

        row_border_counts: Dict[int, int] = {}
        col_border_counts: Dict[int, int] = {}
        for r in range(comp_minr, comp_maxr + 1):
            row_count = 0
            for c in range(comp_minc, comp_maxc + 1):
                coord = coord_of(r, c)
                txt = cell_text(sheet, coord).strip()
                st = cell_style(sheet, styles, coord)
                if has_border(st) or (not txt and side_border_count(st) >= 2):
                    row_count += 1
                    col_border_counts[c] = col_border_counts.get(c, 0) + 1
            row_border_counts[r] = row_count

        # Candidate table rows: rows with enough bordered cells. This removes section
        # captions such as “一、渠道销售概览” that are adjacent to the table.
        candidate_rows = [
            r for r, n in row_border_counts.items()
            if n >= max(2, min(3, comp_maxc - comp_minc + 1))
        ]
        if candidate_rows:
            # Keep the longest consecutive run of candidate rows.
            runs = []
            start = prev = candidate_rows[0]
            for r in candidate_rows[1:]:
                if r == prev + 1:
                    prev = r
                else:
                    runs.append((start, prev))
                    start = prev = r
            runs.append((start, prev))
            minr, maxr = max(runs, key=lambda x: (x[1] - x[0] + 1, -x[0]))
            candidate_cols = [
                c for c in range(comp_minc, comp_maxc + 1)
                if sum(1 for r in range(minr, maxr + 1) if has_border(cell_style(sheet, styles, coord_of(r, c))) or side_border_count(cell_style(sheet, styles, coord_of(r, c))) >= 2) >= max(2, int((maxr - minr + 1) * 0.35))
            ]
            if candidate_cols:
                minc, maxc = min(candidate_cols), max(candidate_cols)
            else:
                minc, maxc = comp_minc, comp_maxc
        else:
            minr, minc, maxr, maxc = comp_minr, comp_minc, comp_maxr, comp_maxc

        height = maxr - minr + 1
        width = maxc - minc + 1
        if height < 2 or width < 2:
            continue

        non_empty_count = 0
        density_denominator = height * width
        border_count = 0
        numeric_count = 0
        text_count = 0
        total_rows = []
        for r in range(minr, maxr + 1):
            row_has_total = False
            for c in range(minc, maxc + 1):
                coord = coord_of(r, c)
                txt = cell_text(sheet, coord).strip()
                st = cell_style(sheet, styles, coord)
                if has_border(st):
                    border_count += 1
                if not txt:
                    continue
                non_empty_count += 1
                v = sheet.get("cells", {}).get(coord, {}).get("value")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    numeric_count += 1
                else:
                    text_count += 1
                if txt in ("合计", "小计", "总计", "累计"):
                    row_has_total = True
            if row_has_total:
                total_rows.append(r)
        density = non_empty_count / float(density_denominator)
        border_density = border_count / max(density_denominator, 1)
        numeric_ratio = numeric_count / max(numeric_count + text_count, 1)
        score = 0.0
        if density >= 0.35:
            score += 0.2
        if border_density >= 0.4:
            score += 0.25
        if numeric_ratio >= 0.25:
            score += 0.2
        if height >= 4:
            score += 0.15
        if width >= 4:
            score += 0.1
        if total_rows:
            score += 0.1
        if score < 0.45:
            continue

        # Header heuristic: leading consecutive rows until first row with numeric majority.
        header_end = minr
        for r in range(minr, maxr + 1):
            nums = 0
            vals = 0
            for c in range(minc, maxc + 1):
                coord = coord_of(r, c)
                if not cell_text(sheet, coord).strip():
                    continue
                vals += 1
                v = sheet.get("cells", {}).get(coord, {}).get("value")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    nums += 1
            if vals and nums / vals >= 0.4:
                header_end = max(minr, r - 1)
                break
            header_end = r
        if header_end >= maxr:
            header_end = min(minr, maxr - 1)
        body_start = header_end + 1
        regions.append({
            "type": "table",
            "range": f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{maxr}",
            "header_range": f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{header_end}",
            "body_range": f"{get_column_letter(minc)}{body_start}:{get_column_letter(maxc)}{maxr}",
            "total_rows": total_rows,
            "metrics": {
                "height": height,
                "width": width,
                "non_empty": non_empty_count,
                "density": round(density, 3),
                "border_density": round(border_density, 3),
                "numeric_ratio": round(numeric_ratio, 3),
                "component_range": f"{get_column_letter(comp_minc)}{comp_minr}:{get_column_letter(comp_maxc)}{comp_maxr}",
            },
            "confidence": round(min(0.98, 0.45 + score), 3),
        })
    return regions


def infer_sheet_logic(sheet: Dict[str, Any], styles: Dict[str, Any]) -> Dict[str, Any]:
    regions: List[Dict[str, Any]] = []
    regions.extend(infer_title(sheet, styles))
    regions.extend(infer_metadata(sheet, styles))
    regions.extend(infer_tables(sheet, styles))
    return {"regions": regions}


def diff_workbooks(path_a: str, path_b: str) -> Dict[str, Any]:
    a = parse_workbook(path_a, include_empty_styled=True, infer_logic=False)
    b = parse_workbook(path_b, include_empty_styled=True, infer_logic=False)
    diffs: List[Dict[str, Any]] = []

    styles_a = a["workbook"]["styles"]
    styles_b = b["workbook"]["styles"]
    sheets_a = {s["name"]: s for s in a["workbook"]["sheets"]}
    sheets_b = {s["name"]: s for s in b["workbook"]["sheets"]}

    for sname in sorted(set(sheets_a) | set(sheets_b)):
        sa = sheets_a.get(sname)
        sb = sheets_b.get(sname)
        if sa is None or sb is None:
            diffs.append({"sheet": sname, "type": "sheet_missing", "a": sa is not None, "b": sb is not None})
            continue
        for field in ("freeze_panes", "merged_ranges", "rows", "cols"):
            va = sa.get(field)
            vb = sb.get(field)
            if va != vb:
                diffs.append({"sheet": sname, "type": field, "a": va, "b": vb})
        coords = sorted(set(sa.get("cells", {})) | set(sb.get("cells", {})))
        for coord in coords:
            ca = sa.get("cells", {}).get(coord)
            cb = sb.get("cells", {}).get(coord)
            if ca is None or cb is None:
                diffs.append({"sheet": sname, "coord": coord, "type": "cell_missing", "a": ca, "b": cb})
                continue
            for field in ("value", "data_type", "hyperlink"):
                if ca.get(field) != cb.get(field):
                    diffs.append({"sheet": sname, "coord": coord, "type": field, "a": ca.get(field), "b": cb.get(field)})
            sty_a = styles_a.get(ca.get("style_id"), {})
            sty_b = styles_b.get(cb.get("style_id"), {})
            if sty_a != sty_b:
                diffs.append({"sheet": sname, "coord": coord, "type": "style", "a": sty_a, "b": sty_b})

    return {"diff_count": len(diffs), "diffs": diffs[:200], "truncated": len(diffs) > 200}


def save_json(data: Dict[str, Any], path: str) -> None:
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def main() -> None:
    p = argparse.ArgumentParser(description="Excel report IR MVP")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_parse = sub.add_parser("parse")
    p_parse.add_argument("input")
    p_parse.add_argument("output")
    p_parse.add_argument("--no-logic", action="store_true")

    p_rebuild = sub.add_parser("rebuild")
    p_rebuild.add_argument("input_json")
    p_rebuild.add_argument("output_xlsx")

    p_diff = sub.add_parser("diff")
    p_diff.add_argument("a")
    p_diff.add_argument("b")
    p_diff.add_argument("output_json", nargs="?")

    p_logic = sub.add_parser("logic")
    p_logic.add_argument("input_json")

    args = p.parse_args()
    if args.cmd == "parse":
        ir = parse_workbook(args.input, infer_logic=not args.no_logic)
        save_json(ir, args.output)
        print(json.dumps({"ok": True, "sheets": len(ir["workbook"]["sheets"]), "styles": len(ir["workbook"]["styles"]), "output": args.output}, ensure_ascii=False))
    elif args.cmd == "rebuild":
        ir = load_json(args.input_json)
        rebuild_workbook(ir, args.output_xlsx)
        print(json.dumps({"ok": True, "output": args.output_xlsx}, ensure_ascii=False))
    elif args.cmd == "diff":
        d = diff_workbooks(args.a, args.b)
        if args.output_json:
            save_json(d, args.output_json)
        print(json.dumps(d, ensure_ascii=False, indent=2))
    elif args.cmd == "logic":
        ir = load_json(args.input_json)
        for s in ir["workbook"]["sheets"]:
            print(json.dumps({"sheet": s["name"], "logical": s.get("logical", {})}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

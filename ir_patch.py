from __future__ import annotations

import copy
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

import excel_ir
from formula_utils import extract_references, shift_formula_references, workbook_formula_dependencies


def load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def save_json(data: Dict[str, Any], path: str) -> None:
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def get_sheet(ir: Dict[str, Any], name: str | None = None, index: int | None = None) -> Dict[str, Any]:
    sheets = ir["workbook"]["sheets"]
    if name is not None:
        for s in sheets:
            if s["name"] == name:
                s.setdefault("name", name)
                return s
        raise KeyError(f"sheet not found: {name}")
    idx = index or 0
    sheets[idx].setdefault("name", sheets[idx].get("name", f"Sheet{idx+1}"))
    return sheets[idx]


def coord_to_rc(coord: str) -> Tuple[int, int]:
    m = re.fullmatch(r"([A-Z]+)(\d+)", coord.upper())
    if not m:
        raise ValueError(f"invalid coord: {coord}")
    return int(m.group(2)), column_index_from_string(m.group(1))


def rc_to_coord(r: int, c: int) -> str:
    return f"{get_column_letter(c)}{r}"


def iter_range(ref: str):
    minc, minr, maxc, maxr = range_boundaries(ref)
    for r in range(minr, maxr + 1):
        for c in range(minc, maxc + 1):
            yield rc_to_coord(r, c), r, c


def ensure_style(ir: Dict[str, Any], style: Dict[str, Any]) -> str:
    styles = ir["workbook"].setdefault("styles", {})
    key = json.dumps(style, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    for sid, s in styles.items():
        if json.dumps(s, ensure_ascii=False, sort_keys=True, separators=(",", ":")) == key:
            return sid
    sid = f"s{len(styles) + 1:04d}"
    while sid in styles:
        sid = f"s{len(styles) + 1:04d}_{len(styles)}"
    styles[sid] = style
    return sid


def default_cell(row: int, col: int, style_id: str | None = None) -> Dict[str, Any]:
    d = {"row": row, "col": col, "value": None, "data_type": "n"}
    if style_id:
        d["style_id"] = style_id
    return d


def ensure_cell(sheet: Dict[str, Any], coord: str, style_id: str | None = None) -> Dict[str, Any]:
    r, c = coord_to_rc(coord)
    sheet.setdefault("cells", {})
    if coord not in sheet["cells"]:
        sheet["cells"][coord] = default_cell(r, c, style_id)
    sheet["dimensions"]["max_row"] = max(sheet["dimensions"].get("max_row", 1), r)
    sheet["dimensions"]["max_col"] = max(sheet["dimensions"].get("max_col", 1), c)
    return sheet["cells"][coord]


def infer_data_type(value: Any) -> str:
    if value is None:
        return "n"
    if isinstance(value, str) and value.startswith("="):
        return "f"
    if isinstance(value, bool):
        return "b"
    if isinstance(value, (int, float)):
        return "n"
    return "s"


def shift_formula_text(formula: Any, row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> Any:
    return shift_formula_references(formula, row_delta=row_delta, col_delta=col_delta, row_at=row_at, col_at=col_at)


def shift_cells(sheet: Dict[str, Any], row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> None:
    new_cells: Dict[str, Dict[str, Any]] = {}
    for coord, cell in sheet.get("cells", {}).items():
        r, c = cell["row"], cell["col"]
        nr = r + row_delta if r >= row_at else r
        nc = c + col_delta if c >= col_at else c
        if nr < 1 or nc < 1:
            continue
        ncoord = rc_to_coord(nr, nc)
        ncell = copy.deepcopy(cell)
        ncell["row"] = nr
        ncell["col"] = nc
        if isinstance(ncell.get("value"), str) and ncell["value"].startswith("="):
            ncell["value"] = shift_formula_text(ncell["value"], row_delta, col_delta, row_at, col_at)
        new_cells[ncoord] = ncell
    sheet["cells"] = new_cells


def shift_range_ref(ref: str, row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> str:
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return ref
    if minr >= row_at:
        minr += row_delta
        maxr += row_delta
    elif maxr >= row_at:
        maxr += row_delta
    if minc >= col_at:
        minc += col_delta
        maxc += col_delta
    elif maxc >= col_at:
        maxc += col_delta
    return f"{get_column_letter(max(1, minc))}{max(1, minr)}:{get_column_letter(max(1, maxc))}{max(1, maxr)}"


def shift_refs(sheet: Dict[str, Any], row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> None:
    sheet["merged_ranges"] = [shift_range_ref(r, row_delta, col_delta, row_at, col_at) for r in sheet.get("merged_ranges", [])]
    rows = sheet.get("rows", {})
    if row_delta:
        new_rows = {}
        for k, v in rows.items():
            i = int(k)
            ni = i + row_delta if i >= row_at else i
            if ni >= 1:
                new_rows[str(ni)] = v
        sheet["rows"] = new_rows
    cols = sheet.get("cols", {})
    if col_delta:
        new_cols = {}
        for k, v in cols.items():
            i = column_index_from_string(k)
            ni = i + col_delta if i >= col_at else i
            if ni >= 1:
                new_cols[get_column_letter(ni)] = v
        sheet["cols"] = new_cols
    extra = sheet.get("extra", {})
    if extra.get("auto_filter", {}).get("ref"):
        extra["auto_filter"]["ref"] = shift_range_ref(extra["auto_filter"]["ref"], row_delta, col_delta, row_at, col_at)
    for t in extra.get("tables", []) or []:
        if t.get("ref"):
            t["ref"] = shift_range_ref(t["ref"], row_delta, col_delta, row_at, col_at)
    for dv in extra.get("data_validations", []) or []:
        if dv.get("sqref"):
            dv["sqref"] = " ".join(shift_range_ref(x, row_delta, col_delta, row_at, col_at) for x in str(dv["sqref"]).split())
    for cf in extra.get("conditional_formatting", []) or []:
        if cf.get("sqref"):
            cf["sqref"] = " ".join(shift_range_ref(x, row_delta, col_delta, row_at, col_at) for x in str(cf["sqref"]).split())
    for img in extra.get("images", []) or []:
        cell = img.get("anchor_cell")
        if cell:
            r, c = coord_to_rc(cell)
            if r >= row_at:
                r += row_delta
            if c >= col_at:
                c += col_delta
            img["anchor_cell"] = rc_to_coord(max(1, r), max(1, c))
    shift_chart_refs(sheet, row_delta, col_delta, row_at, col_at)


def shift_chart_range_obj(obj: Any, row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> None:
    if isinstance(obj, dict):
        for key in ("min_col", "max_col", "min_row", "max_row"):
            if key in obj:
                obj[key] = int(obj[key])
        if obj.get("min_row") is not None:
            if obj["min_row"] >= row_at:
                obj["min_row"] += row_delta
                obj["max_row"] += row_delta
            elif obj["max_row"] >= row_at:
                obj["max_row"] += row_delta
        if obj.get("min_col") is not None:
            if obj["min_col"] >= col_at:
                obj["min_col"] += col_delta
                obj["max_col"] += col_delta
            elif obj["max_col"] >= col_at:
                obj["max_col"] += col_delta
        if obj.get("min_col") is not None:
            obj["min_col"] = max(1, obj["min_col"])
            obj["max_col"] = max(1, obj["max_col"])
            obj["min_row"] = max(1, obj["min_row"])
            obj["max_row"] = max(1, obj["max_row"])
            obj["ref"] = f"{get_column_letter(obj['min_col'])}{obj['min_row']}:{get_column_letter(obj['max_col'])}{obj['max_row']}"
            sheet = obj.get("sheet")
            ref_abs = f"${get_column_letter(obj['min_col'])}${obj['min_row']}:${get_column_letter(obj['max_col'])}${obj['max_row']}"
            obj["formula"] = f"'{sheet}'!{ref_abs}" if sheet else ref_abs


def shift_chart_refs(sheet: Dict[str, Any], row_delta: int = 0, col_delta: int = 0, row_at: int = 1, col_at: int = 1) -> None:
    for ch in sheet.get("extra", {}).get("charts", []) or []:
        for key in ("anchor", "anchor_cell"):
            if ch.get(key):
                try:
                    r, c = coord_to_rc(ch[key])
                    if r >= row_at:
                        r += row_delta
                    if c >= col_at:
                        c += col_delta
                    ch[key] = rc_to_coord(max(1, r), max(1, c))
                except Exception:
                    pass
        for s in ch.get("series", []) or []:
            for key in ("title", "values", "categories"):
                shift_chart_range_obj(s.get(key), row_delta, col_delta, row_at, col_at)
            if s.get("title"):
                t = s["title"]
                s["title_ref"] = t.get("formula", s.get("title_ref"))


def insert_rows(sheet: Dict[str, Any], idx: int, amount: int) -> None:
    shift_cells(sheet, row_delta=amount, row_at=idx)
    shift_refs(sheet, row_delta=amount, row_at=idx)
    sheet["dimensions"]["max_row"] = sheet["dimensions"].get("max_row", 1) + amount


def delete_rows(sheet: Dict[str, Any], idx: int, amount: int) -> None:
    # Drop cells in deleted band, then shift cells after it.
    cells = {}
    for coord, cell in sheet.get("cells", {}).items():
        r = cell["row"]
        if idx <= r < idx + amount:
            continue
        cells[coord] = cell
    sheet["cells"] = cells
    shift_cells(sheet, row_delta=-amount, row_at=idx + amount)
    shift_refs(sheet, row_delta=-amount, row_at=idx + amount)
    sheet["dimensions"]["max_row"] = max(1, sheet["dimensions"].get("max_row", 1) - amount)


def insert_cols(sheet: Dict[str, Any], idx: int, amount: int) -> None:
    shift_cells(sheet, col_delta=amount, col_at=idx)
    shift_refs(sheet, col_delta=amount, col_at=idx)
    sheet["dimensions"]["max_col"] = sheet["dimensions"].get("max_col", 1) + amount


def delete_cols(sheet: Dict[str, Any], idx: int, amount: int) -> None:
    cells = {}
    for coord, cell in sheet.get("cells", {}).items():
        c = cell["col"]
        if idx <= c < idx + amount:
            continue
        cells[coord] = cell
    sheet["cells"] = cells
    shift_cells(sheet, col_delta=-amount, col_at=idx + amount)
    shift_refs(sheet, col_delta=-amount, col_at=idx + amount)
    sheet["dimensions"]["max_col"] = max(1, sheet["dimensions"].get("max_col", 1) - amount)


def copy_col_style(sheet: Dict[str, Any], source_col: int, target_col: int) -> None:
    max_row = sheet["dimensions"].get("max_row", 1)
    for r in range(1, max_row + 1):
        src = sheet.get("cells", {}).get(rc_to_coord(r, source_col))
        if src and src.get("style_id"):
            tgt = ensure_cell(sheet, rc_to_coord(r, target_col), src.get("style_id"))
            tgt["style_id"] = src["style_id"]
            tgt["value"] = None
            tgt["data_type"] = "n"
    src_letter = get_column_letter(source_col)
    tgt_letter = get_column_letter(target_col)
    if src_letter in sheet.get("cols", {}):
        sheet.setdefault("cols", {})[tgt_letter] = copy.deepcopy(sheet["cols"][src_letter])


def copy_row_style(sheet: Dict[str, Any], source_row: int, target_row: int) -> None:
    max_col = sheet["dimensions"].get("max_col", 1)
    for c in range(1, max_col + 1):
        src = sheet.get("cells", {}).get(rc_to_coord(source_row, c))
        if src and src.get("style_id"):
            tgt = ensure_cell(sheet, rc_to_coord(target_row, c), src.get("style_id"))
            tgt["style_id"] = src["style_id"]
            tgt["value"] = None
            tgt["data_type"] = "n"
    if str(source_row) in sheet.get("rows", {}):
        sheet.setdefault("rows", {})[str(target_row)] = copy.deepcopy(sheet["rows"][str(source_row)])


def set_cell(sheet: Dict[str, Any], coord: str, value: Any = None, style_id: str | None = None, data_type: str | None = None) -> None:
    cell = ensure_cell(sheet, coord, style_id)
    cell["value"] = value
    cell["data_type"] = data_type or infer_data_type(value)
    if style_id:
        cell["style_id"] = style_id


def set_range_values(sheet: Dict[str, Any], start: str, values: List[List[Any]], style_from: str | None = None) -> None:
    sr, sc = coord_to_rc(start)
    style_id = None
    if style_from:
        src = sheet.get("cells", {}).get(style_from)
        style_id = src.get("style_id") if src else None
    for i, row in enumerate(values):
        for j, value in enumerate(row):
            set_cell(sheet, rc_to_coord(sr + i, sc + j), value, style_id=style_id)


def apply_style_patch(ir: Dict[str, Any], sheet: Dict[str, Any], target: str, patch: Dict[str, Any]) -> str:
    if ":" in target:
        first = target.split(":", 1)[0]
    else:
        first = target
    cell = ensure_cell(sheet, first)
    current = copy.deepcopy(ir["workbook"].get("styles", {}).get(cell.get("style_id"), {}))
    # Deep merge patch.
    def merge(a, b):
        for k, v in b.items():
            if isinstance(v, dict) and isinstance(a.get(k), dict):
                merge(a[k], v)
            else:
                a[k] = v
        return a
    new_style = merge(current, patch)
    sid = ensure_style(ir, new_style)
    targets = [target] if ":" not in target else [coord for coord, _, _ in iter_range(target)]
    for coord in targets:
        ensure_cell(sheet, coord)["style_id"] = sid
    return sid


def table_bounds(table: Dict[str, Any]) -> Tuple[int, int, int, int]:
    minc, minr, maxc, maxr = range_boundaries(table["ref"])
    return minc, minr, maxc, maxr


def table_field_map(sheet: Dict[str, Any], table: Dict[str, Any], header_rows: int = 1, explicit: Dict[str, Any] | None = None) -> Dict[str, int]:
    fmap: Dict[str, int] = {}
    minc, minr, maxc, maxr = table_bounds(table)
    for c in range(minc, maxc + 1):
        parts = []
        for r in range(minr, min(minr + header_rows, maxr + 1)):
            v = sheet.get("cells", {}).get(rc_to_coord(r, c), {}).get("value")
            if v not in (None, ""):
                parts.append(str(v).strip())
        if parts:
            fmap["/".join(parts)] = c
            fmap[parts[-1]] = c
    for k, v in (explicit or {}).items():
        fmap[str(k)] = resolve_table_col(sheet, table, v, header_rows) if not isinstance(v, int) else v
    return fmap


def resolve_table_col(sheet: Dict[str, Any], table: Dict[str, Any], col: Any, header_rows: int = 1, field_map: Dict[str, Any] | None = None) -> int:
    minc, minr, maxc, maxr = table_bounds(table)
    if field_map and isinstance(col, str) and col in field_map:
        mapped = field_map[col]
        return mapped if isinstance(mapped, int) else resolve_table_col(sheet, table, mapped, header_rows, None)
    if isinstance(col, int):
        # Relative table index if small; absolute column if already in range.
        return col if minc <= col <= maxc else minc + col - 1
    if isinstance(col, str):
        s = col.strip()
        if re.fullmatch(r"[A-Za-z]+", s):
            return column_index_from_string(s.upper())
        # Header lookup across one or more header rows. Last non-empty duplicate wins
        # only if unique; otherwise users should specify a letter/index.
        hits = []
        for r in range(minr, min(minr + header_rows, maxr + 1)):
            for c in range(minc, maxc + 1):
                v = sheet.get("cells", {}).get(rc_to_coord(r, c), {}).get("value")
                if str(v).strip() == s:
                    hits.append(c)
        if len(hits) == 1:
            return hits[0]
        if len(hits) > 1:
            raise ValueError(f"ambiguous table column header {col!r}: {hits}")
    raise ValueError(f"cannot resolve table column: {col!r}")


def table_data_rows(sheet: Dict[str, Any], table: Dict[str, Any], header_rows: int = 1, include_total: bool = False) -> List[int]:
    minc, minr, maxc, maxr = table_bounds(table)
    start = minr + header_rows
    end = maxr if include_total else maxr - 1
    return list(range(start, max(start - 1, end) + 1))


def cell_value(sheet: Dict[str, Any], row: int, col: int) -> Any:
    return sheet.get("cells", {}).get(rc_to_coord(row, col), {}).get("value")


def compare_value(actual: Any, op: str, expected: Any) -> bool:
    if op in ("eq", "=="):
        return actual == expected
    if op in ("ne", "!="):
        return actual != expected
    if op == "contains":
        return str(expected) in str(actual)
    if op == "not_contains":
        return str(expected) not in str(actual)
    if op in ("gt", ">", "gte", ">=", "lt", "<", "lte", "<="):
        try:
            a = float(actual)
            b = float(expected)
        except Exception:
            return False
        if op in ("gt", ">"):
            return a > b
        if op in ("gte", ">="):
            return a >= b
        if op in ("lt", "<"):
            return a < b
        return a <= b
    if op == "regex":
        return re.search(str(expected), str(actual)) is not None
    if op == "in":
        return actual in expected
    raise ValueError(f"unknown condition op: {op}")


def row_matches(sheet: Dict[str, Any], table: Dict[str, Any], row: int, where: Any, header_rows: int = 1, field_map: Dict[str, Any] | None = None) -> bool:
    if not where:
        return True
    if isinstance(where, list):
        return all(row_matches(sheet, table, row, w, header_rows, field_map) for w in where)
    if "all" in where:
        return all(row_matches(sheet, table, row, w, header_rows, field_map) for w in where.get("all", []))
    if "and" in where:
        return all(row_matches(sheet, table, row, w, header_rows, field_map) for w in where.get("and", []))
    if "any" in where:
        return any(row_matches(sheet, table, row, w, header_rows, field_map) for w in where.get("any", []))
    if "or" in where:
        return any(row_matches(sheet, table, row, w, header_rows, field_map) for w in where.get("or", []))
    if "not" in where:
        return not row_matches(sheet, table, row, where.get("not"), header_rows, field_map)
    col = resolve_table_col(sheet, table, where.get("col") or where.get("column"), header_rows, field_map)
    actual = cell_value(sheet, row, col)
    return compare_value(actual, where.get("op", "eq"), where.get("value"))


def render_value_template(value: Any, row: int) -> Any:
    if isinstance(value, str):
        return value.replace("{row}", str(row))
    return value


def set_table_row_values(sheet: Dict[str, Any], table: Dict[str, Any], row: int, values: Any, header_rows: int = 1, field_map: Dict[str, Any] | None = None) -> None:
    minc, minr, maxc, maxr = table_bounds(table)
    if isinstance(values, dict):
        for col_spec, value in values.items():
            col = resolve_table_col(sheet, table, col_spec, header_rows, field_map)
            set_cell(sheet, rc_to_coord(row, col), render_value_template(value, row))
    else:
        for j, value in enumerate(values or []):
            if minc + j <= maxc:
                set_cell(sheet, rc_to_coord(row, minc + j), render_value_template(value, row))


def append_table_rows(ir: Dict[str, Any], sheet: Dict[str, Any], action: Dict[str, Any]) -> None:
    rows = action.get("rows", [])
    for values in rows:
        one = dict(action)
        one["values"] = values
        one.pop("rows", None)
        append_table_row(ir, sheet, one)


def update_rows_where(sheet: Dict[str, Any], action: Dict[str, Any]) -> int:
    table = find_table(sheet, action.get("table"), action.get("ref"))
    header_rows = int(action.get("header_rows", 1))
    fmap = get_action_field_map(sheet, table, action, header_rows)
    count = 0
    for r in table_data_rows(sheet, table, header_rows, include_total=bool(action.get("include_total", False))):
        if row_matches(sheet, table, r, action.get("where"), header_rows, fmap):
            set_table_row_values(sheet, table, r, action.get("updates", {}), header_rows, fmap)
            count += 1
    sheet.setdefault("logical", {}).setdefault("patch_stats", []).append({"op": "update_rows_where", "count": count})
    return count


def delete_rows_where(sheet: Dict[str, Any], action: Dict[str, Any]) -> int:
    table = find_table(sheet, action.get("table"), action.get("ref"))
    header_rows = int(action.get("header_rows", 1))
    fmap = get_action_field_map(sheet, table, action, header_rows)
    rows = [r for r in table_data_rows(sheet, table, header_rows, include_total=False) if row_matches(sheet, table, r, action.get("where"), header_rows, fmap)]
    for r in sorted(rows, reverse=True):
        delete_rows(sheet, r, 1)
    sheet.setdefault("logical", {}).setdefault("patch_stats", []).append({"op": "delete_rows_where", "count": len(rows)})
    return len(rows)


def recompute_totals(sheet: Dict[str, Any], action: Dict[str, Any]) -> None:
    table = find_table(sheet, action.get("table"), action.get("ref"))
    minc, minr, maxc, maxr = table_bounds(table)
    header_rows = int(action.get("header_rows", 1))
    total_row = int(action.get("total_row", maxr))
    data_start = int(action.get("data_start", minr + header_rows))
    data_end = int(action.get("data_end", total_row - 1))
    formulas = action.get("formulas", {})
    explicit_map = {}
    explicit_map.update(table.get("ir", {}).get("field_map", {}) or {})
    explicit_map.update(action.get("field_map", {}) or {})
    fmap = table_field_map(sheet, table, header_rows, explicit_map)
    for col_spec, spec in formulas.items():
        col = resolve_table_col(sheet, table, col_spec, header_rows, fmap)
        letter = get_column_letter(col)
        if isinstance(spec, str) and spec.upper() == "SUM":
            formula = f"=SUM({letter}{data_start}:{letter}{data_end})"
        elif isinstance(spec, str):
            formula = spec
            if not formula.startswith("="):
                formula = "=" + formula
            formula = formula.replace("{row}", str(total_row)).replace("{col}", letter).replace("{data_start}", str(data_start)).replace("{data_end}", str(data_end))
        else:
            continue
        set_cell(sheet, rc_to_coord(total_row, col), formula, data_type="f")


def find_table(sheet: Dict[str, Any], name: str | None = None, ref: str | None = None) -> Dict[str, Any]:
    tables = sheet.setdefault("extra", {}).setdefault("tables", [])
    if name:
        for t in tables:
            if t.get("displayName") == name or t.get("name") == name:
                return t
        raise KeyError(f"table not found: {name}")
    if ref:
        for t in tables:
            if t.get("ref") == ref:
                return t
    if tables:
        return tables[0]
    raise KeyError("no table in sheet")


def append_table_row(ir: Dict[str, Any], sheet: Dict[str, Any], action: Dict[str, Any]) -> None:
    table = find_table(sheet, action.get("table"), action.get("ref"))
    minc, minr, maxc, maxr = range_boundaries(table["ref"])
    insert_at = maxr + 1 if action.get("after_total") else maxr
    insert_rows(sheet, insert_at, 1)
    style_from = int(action.get("copy_style_from_row", maxr - 1 if maxr > minr else maxr))
    copy_row_style(sheet, style_from, insert_at)
    values = action.get("values", [])
    if isinstance(values, dict):
        # Map by header labels from the first table row.
        header_map = {}
        for c in range(minc, maxc + 1):
            txt = sheet.get("cells", {}).get(rc_to_coord(minr, c), {}).get("value")
            if txt:
                header_map[str(txt)] = c
        for k, v in values.items():
            if k in header_map:
                set_cell(sheet, rc_to_coord(insert_at, header_map[k]), v)
    else:
        for j, v in enumerate(values):
            if minc + j <= maxc:
                set_cell(sheet, rc_to_coord(insert_at, minc + j), v)
    # insert_rows already shifts table refs and auto filters through shift_refs.
    # Do not shift the table range a second time here.
    sheet.setdefault("extra", {}).setdefault("auto_filter", {})["ref"] = table["ref"]


def get_chart(sheet: Dict[str, Any], name: str | None = None, index: int | None = None) -> Dict[str, Any]:
    charts = sheet.setdefault("extra", {}).setdefault("charts", [])
    if name:
        for ch in charts:
            if ch.get("name") == name:
                return ch
        raise KeyError(f"chart not found: {name}")
    if index is not None:
        return charts[int(index)]
    if charts:
        return charts[0]
    raise KeyError("no chart in sheet")


def chart_range_from_ref(sheet_name: str, ref: str) -> Dict[str, Any]:
    minc, minr, maxc, maxr = range_boundaries(ref.replace("$", ""))
    return {
        "sheet": sheet_name,
        "ref": f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{maxr}",
        "min_col": minc,
        "min_row": minr,
        "max_col": maxc,
        "max_row": maxr,
        "formula": f"'{sheet_name}'!${get_column_letter(minc)}${minr}:${get_column_letter(maxc)}${maxr}",
    }


def apply_chart_action(sheet: Dict[str, Any], action: Dict[str, Any]) -> None:
    op = action.get("op")
    if op == "add_chart":
        charts = sheet.setdefault("extra", {}).setdefault("charts", [])
        charts.append({k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")})
        return
    ch = get_chart(sheet, action.get("chart"), action.get("chart_index"))
    if op == "set_chart_anchor":
        ch["anchor"] = action["anchor"]
        ch["anchor_cell"] = action["anchor"]
    elif op == "set_chart_title":
        ch["title"] = action.get("title")
        if "x_axis_title" in action:
            ch["x_axis_title"] = action.get("x_axis_title")
        if "y_axis_title" in action:
            ch["y_axis_title"] = action.get("y_axis_title")
    elif op == "set_chart_size":
        if "height" in action:
            ch["height"] = action["height"]
        if "width" in action:
            ch["width"] = action["width"]
    elif op == "set_chart_series_ranges":
        sheet_name = sheet.get("name", "Sheet")
        cats = action.get("categories")
        vals = action.get("values", [])
        titles = action.get("titles", [])
        series = []
        for i, vref in enumerate(vals):
            s = {"values": chart_range_from_ref(sheet_name, vref)}
            if cats:
                s["categories"] = chart_range_from_ref(sheet_name, cats)
            if i < len(titles):
                s["title"] = chart_range_from_ref(sheet_name, titles[i])
                s["title_ref"] = s["title"]["formula"]
            series.append(s)
        ch["series"] = series
    else:
        raise ValueError(f"unknown chart op: {op}")


def apply_action(ir: Dict[str, Any], action: Dict[str, Any]) -> None:
    op = action.get("op")
    sheet = get_sheet(ir, action.get("sheet"), action.get("sheet_index"))

    if op == "confirm_field_map":
        # Confirm parser-generated candidates, optionally with overrides.
        table = find_table(sheet, action.get("table"), action.get("ref"))
        candidates = table.get("ir", {}).get("field_map_candidates", {})
        overrides = action.get("field_map", {}) or {}
        action2 = dict(action)
        action2["field_map"] = {**candidates, **overrides}
        persist_table_field_map(sheet, action2)
    elif op == "persist_table_field_map":
        persist_table_field_map(sheet, action)
    elif op == "append_table_rows":
        append_table_rows(ir, sheet, action)
    elif op == "update_rows_where":
        update_rows_where(sheet, action)
    elif op == "delete_rows_where":
        delete_rows_where(sheet, action)
    elif op == "recompute_totals":
        recompute_totals(sheet, action)
    elif op in {"add_chart", "set_chart_anchor", "set_chart_title", "set_chart_size", "set_chart_series_ranges"}:
        apply_chart_action(sheet, action)
    elif op == "append_table_row":
        append_table_row(ir, sheet, action)
    elif op == "set_cell":
        set_cell(sheet, action["coord"], action.get("value"), action.get("style_id"), action.get("data_type"))
    elif op == "set_formula":
        formula = action["formula"]
        if not formula.startswith("="):
            formula = "=" + formula
        set_cell(sheet, action["coord"], formula, action.get("style_id"), "f")
    elif op == "set_range_values":
        set_range_values(sheet, action["start"], action.get("values", []), action.get("style_from"))
    elif op == "copy_cell":
        src = sheet.get("cells", {}).get(action["source"])
        if src:
            dst = copy.deepcopy(src)
            r, c = coord_to_rc(action["target"])
            dst["row"] = r
            dst["col"] = c
            if not action.get("copy_value", True):
                dst["value"] = None
                dst["data_type"] = "n"
            sheet.setdefault("cells", {})[action["target"]] = dst
    elif op == "copy_style":
        src = sheet.get("cells", {}).get(action["source"])
        if src and src.get("style_id"):
            target = action["target"]
            coords = [target] if ":" not in target else [coord for coord, _, _ in iter_range(target)]
            for coord in coords:
                ensure_cell(sheet, coord)["style_id"] = src["style_id"]
    elif op == "apply_style":
        apply_style_patch(ir, sheet, action["target"], action.get("style", {}))
    elif op == "merge":
        ref = action["range"]
        if ref not in sheet.setdefault("merged_ranges", []):
            sheet["merged_ranges"].append(ref)
    elif op == "unmerge":
        ref = action["range"]
        sheet["merged_ranges"] = [x for x in sheet.get("merged_ranges", []) if x != ref]
    elif op == "set_row_height":
        r = int(action["row"])
        sheet.setdefault("rows", {}).setdefault(str(r), {})["height"] = action["height"]
    elif op == "set_col_width":
        col = action["col"]
        if isinstance(col, int):
            col = get_column_letter(col)
        sheet.setdefault("cols", {}).setdefault(col, {})["width"] = action["width"]
    elif op == "insert_rows":
        insert_rows(sheet, int(action["idx"]), int(action.get("amount", 1)))
        if action.get("copy_style_from"):
            for offset in range(int(action.get("amount", 1))):
                copy_row_style(sheet, int(action["copy_style_from"]), int(action["idx"]) + offset)
    elif op == "insert_cols":
        col = action["idx"]
        idx = column_index_from_string(col) if isinstance(col, str) else int(col)
        insert_cols(sheet, idx, int(action.get("amount", 1)))
        if action.get("copy_style_from"):
            src = action["copy_style_from"]
            src_idx = column_index_from_string(src) if isinstance(src, str) else int(src)
            for offset in range(int(action.get("amount", 1))):
                copy_col_style(sheet, src_idx, idx + offset)
    elif op == "delete_cols":
        col = action["idx"]
        idx = column_index_from_string(col) if isinstance(col, str) else int(col)
        delete_cols(sheet, idx, int(action.get("amount", 1)))
    elif op == "delete_rows":
        delete_rows(sheet, int(action["idx"]), int(action.get("amount", 1)))
    elif op == "remove_data_validation":
        sqref = action.get("sqref")
        items = sheet.setdefault("extra", {}).setdefault("data_validations", [])
        if sqref:
            items[:] = [x for x in items if x.get("sqref") != sqref]
    elif op == "upsert_data_validation":
        sqref = action.get("sqref")
        items = sheet.setdefault("extra", {}).setdefault("data_validations", [])
        if sqref:
            items[:] = [x for x in items if x.get("sqref") != sqref]
        items.append({k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")})
    elif op == "add_data_validation":
        sheet.setdefault("extra", {}).setdefault("data_validations", []).append({k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")})
    elif op == "remove_conditional_formatting":
        sqref = action.get("sqref")
        items = sheet.setdefault("extra", {}).setdefault("conditional_formatting", [])
        if sqref:
            items[:] = [x for x in items if x.get("sqref") != sqref]
    elif op == "upsert_conditional_formatting":
        items = sheet.setdefault("extra", {}).setdefault("conditional_formatting", [])
        sqref = action["sqref"]
        rule = action["rule"]
        for cf in items:
            if cf.get("sqref") == sqref:
                cf.setdefault("rules", []).append(rule)
                break
        else:
            items.append({"sqref": sqref, "rules": [rule]})
    elif op == "add_conditional_formatting":
        sheet.setdefault("extra", {}).setdefault("conditional_formatting", []).append({
            "sqref": action["sqref"],
            "rules": [action["rule"]],
        })
    elif op == "set_auto_filter":
        sheet.setdefault("extra", {}).setdefault("auto_filter", {})["ref"] = action["ref"]
    elif op == "upsert_table":
        name = action.get("displayName") or action.get("name")
        items = sheet.setdefault("extra", {}).setdefault("tables", [])
        payload = {k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")}
        for i, t in enumerate(items):
            if name and (t.get("displayName") == name or t.get("name") == name):
                items[i] = payload
                break
        else:
            items.append(payload)
    elif op == "add_table":
        sheet.setdefault("extra", {}).setdefault("tables", []).append({k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")})
    elif op == "set_print_area":
        sheet.setdefault("extra", {})["print_area"] = action["ref"]
    elif op == "set_freeze_panes":
        sheet["freeze_panes"] = action.get("cell")
    elif op == "annotate":
        sheet.setdefault("logical", {}).setdefault("annotations", []).append({k: v for k, v in action.items() if k not in ("op", "sheet", "sheet_index")})
    else:
        raise ValueError(f"unknown op: {op}")


def persist_table_field_map(sheet: Dict[str, Any], action: Dict[str, Any]) -> Dict[str, Any]:
    table = find_table(sheet, action.get("table"), action.get("ref"))
    header_rows = int(action.get("header_rows", 1))
    fmap = get_action_field_map(sheet, table, action, header_rows)
    # Store as column letters for JSON readability/stability.
    stored = {k: get_column_letter(v) if isinstance(v, int) else v for k, v in fmap.items()}
    table.setdefault("ir", {})["header_rows"] = header_rows
    table.setdefault("ir", {})["field_map"] = stored
    sheet.setdefault("logical", {}).setdefault("patch_stats", []).append({"op": "persist_table_field_map", "fields": len(stored)})
    return stored


def get_action_field_map(sheet: Dict[str, Any], table: Dict[str, Any], action: Dict[str, Any], header_rows: int) -> Dict[str, Any]:
    explicit = {}
    explicit.update(table.get("ir", {}).get("field_map", {}) or {})
    explicit.update(action.get("field_map", {}) or {})
    return table_field_map(sheet, table, header_rows, explicit)


def validate_patch(ir: Dict[str, Any], patch: Dict[str, Any]) -> List[Dict[str, Any]]:
    errors: List[Dict[str, Any]] = []
    if not isinstance(patch.get("actions", []), list):
        return [{"level": "error", "message": "patch.actions must be a list"}]
    known = {
        "confirm_field_map", "persist_table_field_map",
        "append_table_rows", "update_rows_where", "delete_rows_where", "recompute_totals",
        "add_chart", "set_chart_anchor", "set_chart_title", "set_chart_size", "set_chart_series_ranges",
        "append_table_row", "set_cell", "set_formula", "set_range_values", "copy_cell", "copy_style",
        "apply_style", "merge", "unmerge", "set_row_height", "set_col_width", "insert_rows",
        "insert_cols", "delete_cols", "delete_rows", "remove_data_validation", "upsert_data_validation",
        "add_data_validation", "remove_conditional_formatting", "upsert_conditional_formatting",
        "add_conditional_formatting", "set_auto_filter", "upsert_table", "add_table", "set_print_area",
        "set_freeze_panes", "annotate",
    }
    for i, action in enumerate(patch.get("actions", [])):
        op = action.get("op")
        if op not in known:
            errors.append({"level": "error", "index": i, "op": op, "message": f"unknown op: {op}"})
            continue
        try:
            sheet = get_sheet(ir, action.get("sheet"), action.get("sheet_index"))
        except Exception as e:
            errors.append({"level": "error", "index": i, "op": op, "message": str(e)})
            continue
        try:
            if op in {"append_table_row", "append_table_rows", "update_rows_where", "delete_rows_where", "recompute_totals", "persist_table_field_map", "confirm_field_map"}:
                find_table(sheet, action.get("table"), action.get("ref"))
            if op in {"set_chart_anchor", "set_chart_title", "set_chart_size", "set_chart_series_ranges"}:
                get_chart(sheet, action.get("chart"), action.get("chart_index"))
            required = {
                "set_cell": ["coord"], "set_formula": ["coord", "formula"], "set_range_values": ["start"],
                "insert_rows": ["idx"], "delete_rows": ["idx"], "insert_cols": ["idx"], "delete_cols": ["idx"],
                "append_table_row": ["values"], "append_table_rows": ["rows"], "update_rows_where": ["where", "updates"],
                "delete_rows_where": ["where"], "recompute_totals": ["formulas"],
            }
            for key in required.get(op, []):
                if key not in action:
                    errors.append({"level": "error", "index": i, "op": op, "message": f"missing required key: {key}"})
        except Exception as e:
            errors.append({"level": "error", "index": i, "op": op, "message": str(e)})
    return errors


def preview_action(ir: Dict[str, Any], action: Dict[str, Any]) -> Dict[str, Any]:
    op = action.get("op")
    sheet = get_sheet(ir, action.get("sheet"), action.get("sheet_index"))
    item: Dict[str, Any] = {"op": op, "sheet": sheet.get("name")}
    try:
        if op in {"append_table_row", "append_table_rows"}:
            table = find_table(sheet, action.get("table"), action.get("ref"))
            minc, minr, maxc, maxr = table_bounds(table)
            n = len(action.get("rows", [])) if op == "append_table_rows" else 1
            item.update({"table": table.get("name") or table.get("displayName"), "insert_at": maxr, "rows_added": n, "old_ref": table.get("ref")})
        elif op in {"update_rows_where", "delete_rows_where"}:
            table = find_table(sheet, action.get("table"), action.get("ref"))
            header_rows = int(action.get("header_rows", 1))
            fmap = get_action_field_map(sheet, table, action, header_rows)
            rows = [r for r in table_data_rows(sheet, table, header_rows, include_total=False) if row_matches(sheet, table, r, action.get("where"), header_rows, fmap)]
            item.update({"table": table.get("name") or table.get("displayName"), "matched_rows": rows, "count": len(rows)})
        elif op == "recompute_totals":
            table = find_table(sheet, action.get("table"), action.get("ref"))
            minc, minr, maxc, maxr = table_bounds(table)
            item.update({"table": table.get("name") or table.get("displayName"), "total_row": action.get("total_row", maxr), "columns": list(action.get("formulas", {}).keys())})
        elif op in {"insert_rows", "delete_rows"}:
            item.update({"idx": action.get("idx"), "amount": action.get("amount", 1)})
        elif op in {"insert_cols", "delete_cols"}:
            item.update({"idx": action.get("idx"), "amount": action.get("amount", 1)})
        elif op.startswith("set_chart"):
            ch = get_chart(sheet, action.get("chart"), action.get("chart_index"))
            item.update({"chart": ch.get("name"), "anchor": ch.get("anchor"), "title": ch.get("title")})
        else:
            item.update({"target": action.get("coord") or action.get("target") or action.get("range") or action.get("ref")})
    except Exception as e:
        item.update({"error": str(e)})
    return item


def dry_run(ir: Dict[str, Any], patch: Dict[str, Any], simulate: bool = True) -> Dict[str, Any]:
    validation = validate_patch(ir, patch)
    work = copy.deepcopy(ir)
    actions = []
    for a in patch.get("actions", []):
        preview = preview_action(work, a)
        actions.append(preview)
        if simulate and not preview.get("error"):
            try:
                apply_action(work, a)
            except Exception as e:
                preview["apply_error"] = str(e)
    return {
        "ok": not any(e.get("level") == "error" for e in validation) and not any(a.get("apply_error") for a in actions),
        "mode": "sequential-simulated" if simulate else "static",
        "validation": validation,
        "actions": actions,
    }


def formula_dependencies(ir: Dict[str, Any]) -> List[Dict[str, Any]]:
    return workbook_formula_dependencies(ir)


def collect_cell_diffs(before_ir: Dict[str, Any], after_ir: Dict[str, Any], limit: int = 80) -> List[Dict[str, Any]]:
    diffs: List[Dict[str, Any]] = []
    bs = {s.get("name"): s for s in before_ir.get("workbook", {}).get("sheets", [])}
    ads = {s.get("name"): s for s in after_ir.get("workbook", {}).get("sheets", [])}
    for name in sorted(set(bs) | set(ads)):
        b = bs.get(name, {})
        a = ads.get(name, {})
        bc = b.get("cells", {})
        ac = a.get("cells", {})
        for coord in sorted(set(bc) | set(ac)):
            if bc.get(coord) != ac.get(coord):
                diffs.append({"sheet": name, "coord": coord, "before": bc.get(coord), "after": ac.get(coord)})
                if len(diffs) >= limit:
                    return diffs
    return diffs


def summarize_impact(before_ir: Dict[str, Any], after_ir: Dict[str, Any]) -> Dict[str, Any]:
    summary = {"cells_changed": 0, "cells_added": 0, "cells_removed": 0, "tables_changed": [], "charts_changed": []}
    bs = {s.get("name"): s for s in before_ir.get("workbook", {}).get("sheets", [])}
    ads = {s.get("name"): s for s in after_ir.get("workbook", {}).get("sheets", [])}
    for name in set(bs) | set(ads):
        b = bs.get(name, {})
        a = ads.get(name, {})
        bc = b.get("cells", {})
        ac = a.get("cells", {})
        for coord in set(bc) | set(ac):
            if coord not in bc:
                summary["cells_added"] += 1
            elif coord not in ac:
                summary["cells_removed"] += 1
            elif bc[coord] != ac[coord]:
                summary["cells_changed"] += 1
        bt = {t.get("name") or t.get("displayName"): t.get("ref") for t in b.get("extra", {}).get("tables", [])}
        at = {t.get("name") or t.get("displayName"): t.get("ref") for t in a.get("extra", {}).get("tables", [])}
        for tname in set(bt) | set(at):
            if bt.get(tname) != at.get(tname):
                summary["tables_changed"].append({"sheet": name, "table": tname, "before": bt.get(tname), "after": at.get(tname)})
        bch = {c.get("name"): {"title": c.get("title"), "anchor": c.get("anchor")} for c in b.get("extra", {}).get("charts", [])}
        ach = {c.get("name"): {"title": c.get("title"), "anchor": c.get("anchor")} for c in a.get("extra", {}).get("charts", [])}
        for cname in set(bch) | set(ach):
            if bch.get(cname) != ach.get(cname):
                summary["charts_changed"].append({"sheet": name, "chart": cname, "before": bch.get(cname), "after": ach.get(cname)})
    return summary


def append_patch_history(ir: Dict[str, Any], patch: Dict[str, Any], log: Dict[str, Any], impact: Dict[str, Any]) -> None:
    hist = ir.setdefault("workbook", {}).setdefault("patch_history", [])
    hist.append({
        "name": patch.get("name"),
        "actions": len(patch.get("actions", [])),
        "ok": log.get("ok"),
        "impact": impact,
    })


def apply_patch_with_log(ir: Dict[str, Any], patch: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    original = copy.deepcopy(ir)
    out = copy.deepcopy(ir)
    log = {"ok": True, "actions": []}
    try:
        for i, action in enumerate(patch.get("actions", [])):
            before_ir = copy.deepcopy(out)
            before = preview_action(out, action)
            apply_action(out, action)
            after = preview_action(out, action) if action.get("op") not in {"delete_rows_where"} else {"note": "post-delete preview skipped"}
            impact = summarize_impact(before_ir, out)
            cell_diffs = collect_cell_diffs(before_ir, out)
            log["actions"].append({"index": i, "op": action.get("op"), "before": before, "after": after, "impact": impact, "cell_diffs_sample": cell_diffs})
    except Exception as e:
        log["ok"] = False
        log["error"] = str(e)
        raise
    log["impact"] = summarize_impact(original, out)
    deps = formula_dependencies(out)
    log["impact_graph"] = {
        "nodes": {
            "cells": log["impact"].get("cells_changed", 0) + log["impact"].get("cells_added", 0) + log["impact"].get("cells_removed", 0),
            "tables": len(log["impact"].get("tables_changed", [])),
            "charts": len(log["impact"].get("charts_changed", [])),
            "formulas": len(deps),
        },
        "edges": ["cells->formulas", "formulas->charts", "cells->tables", "tables->autofilter"],
        "formula_dependencies_sample": deps[:50],
    }
    append_patch_history(out, patch, log, log["impact"])
    return out, log


def apply_patch(ir: Dict[str, Any], patch: Dict[str, Any]) -> Dict[str, Any]:
    out = copy.deepcopy(ir)
    for action in patch.get("actions", []):
        apply_action(out, action)
    return out


def main() -> None:
    import argparse
    p = argparse.ArgumentParser(description="Apply action patch to Excel IR JSON")
    p.add_argument("input_json")
    p.add_argument("patch_json")
    p.add_argument("output_json", nargs="?")
    p.add_argument("--dry-run", action="store_true", help="validate and preview without writing patched IR")
    p.add_argument("--plan", help="write dry-run plan JSON to this path")
    p.add_argument("--log", help="write transaction apply log JSON to this path")
    args = p.parse_args()
    ir = load_json(args.input_json)
    patch = load_json(args.patch_json)
    plan = dry_run(ir, patch)
    if args.dry_run:
        if args.plan:
            save_json(plan, args.plan)
        print(json.dumps(plan, ensure_ascii=False, indent=2))
        return
    errors = [e for e in plan.get("validation", []) if e.get("level") == "error"]
    if errors:
        print(json.dumps({"ok": False, "errors": errors}, ensure_ascii=False, indent=2))
        raise SystemExit(1)
    if not args.output_json:
        print("output_json is required unless --dry-run is used", file=sys.stderr)
        raise SystemExit(2)
    out, txlog = apply_patch_with_log(ir, patch)
    save_json(out, args.output_json)
    if args.plan:
        save_json(plan, args.plan)
    if args.log:
        save_json(txlog, args.log)
    print(json.dumps({"ok": True, "actions": len(patch.get("actions", [])), "output": args.output_json, "plan": args.plan, "log": args.log}, ensure_ascii=False))


if __name__ == "__main__":
    main()

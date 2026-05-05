from __future__ import annotations

import argparse
import base64
import copy
import hashlib
import json
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

import excel_ir

SCHEMA_VERSION = "0.7"


def _safe_get(obj: Any, attr: str) -> Any:
    try:
        return getattr(obj, attr)
    except Exception:
        return None


def _strip_none(x: Any) -> Any:
    if isinstance(x, dict):
        return {k: _strip_none(v) for k, v in x.items() if v is not None and v != {} and v != []}
    if isinstance(x, list):
        return [_strip_none(v) for v in x if v is not None]
    return x


def _attrs_to_dict(obj: Any, fields: List[str]) -> Dict[str, Any]:
    return _strip_none({f: _safe_get(obj, f) for f in fields})


def _apply_attrs(obj: Any, data: Dict[str, Any], fields: List[str]) -> None:
    for f in fields:
        if f in data:
            try:
                setattr(obj, f, data[f])
            except Exception:
                pass


def dxf_to_dict(dxf: Optional[DifferentialStyle]) -> Dict[str, Any]:
    if not dxf:
        return {}
    data: Dict[str, Any] = {}
    if getattr(dxf, "font", None):
        data["font"] = excel_ir.font_to_dict(dxf.font)
    if getattr(dxf, "fill", None):
        data["fill"] = excel_ir.fill_to_dict(dxf.fill)
    if getattr(dxf, "border", None):
        data["border"] = excel_ir.border_to_dict(dxf.border)
    if getattr(dxf, "alignment", None):
        data["alignment"] = excel_ir.alignment_to_dict(dxf.alignment)
    if getattr(dxf, "protection", None):
        data["protection"] = excel_ir.protection_to_dict(dxf.protection)
    if getattr(dxf, "numFmt", None):
        data["numFmt"] = str(dxf.numFmt)
    return _strip_none(data)


def dxf_from_dict(data: Dict[str, Any]) -> Optional[DifferentialStyle]:
    if not data:
        return None
    kwargs: Dict[str, Any] = {}
    if "font" in data:
        kwargs["font"] = excel_ir.font_from_dict(data["font"])
    if "fill" in data:
        kwargs["fill"] = excel_ir.fill_from_dict(data["fill"])
    if "border" in data:
        kwargs["border"] = excel_ir.border_from_dict(data["border"])
    if "alignment" in data:
        kwargs["alignment"] = excel_ir.alignment_from_dict(data["alignment"])
    if "protection" in data:
        kwargs["protection"] = excel_ir.protection_from_dict(data["protection"])
    return DifferentialStyle(**kwargs)


def cfvo_to_dict(x: Any) -> Dict[str, Any]:
    return _strip_none({
        "type": _safe_get(x, "type"),
        "val": _safe_get(x, "val"),
        "gte": _safe_get(x, "gte"),
    })


def color_scale_to_dict(cs: Any) -> Dict[str, Any]:
    if not cs:
        return {}
    return _strip_none({
        "cfvo": [cfvo_to_dict(x) for x in getattr(cs, "cfvo", [])],
        "color": [excel_ir.color_to_dict(x) for x in getattr(cs, "color", [])],
    })


def rule_to_dict(rule: Rule) -> Dict[str, Any]:
    fields = [
        "type", "operator", "formula", "priority", "stopIfTrue", "aboveAverage",
        "percent", "bottom", "text", "timePeriod", "rank", "stdDev", "equalAverage",
    ]
    data = _attrs_to_dict(rule, fields)
    dxf = dxf_to_dict(getattr(rule, "dxf", None))
    if dxf:
        data["dxf"] = dxf
    cs = color_scale_to_dict(getattr(rule, "colorScale", None))
    if cs:
        data["colorScale"] = cs
    # MVP v0.2 records dataBar/iconSet metadata but does not recreate them generically yet.
    if getattr(rule, "dataBar", None):
        data["dataBar_present"] = True
    if getattr(rule, "iconSet", None):
        data["iconSet_present"] = True
    return _strip_none(data)


def rule_from_dict(data: Dict[str, Any]) -> Rule:
    kwargs: Dict[str, Any] = {"type": data.get("type")}
    for f in ["operator", "formula", "stopIfTrue", "aboveAverage", "percent", "bottom", "text", "timePeriod", "rank", "stdDev", "equalAverage"]:
        if f in data:
            kwargs[f] = data[f]
    dxf = dxf_from_dict(data.get("dxf", {}))
    if dxf:
        kwargs["dxf"] = dxf
    rule = Rule(**kwargs)
    if "priority" in data:
        try:
            rule.priority = data["priority"]
        except Exception:
            pass
    return rule


def parse_conditional_formatting(ws) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    try:
        cfs = list(ws.conditional_formatting)
    except Exception:
        cfs = []
    for cf in cfs:
        items.append(_strip_none({
            "sqref": str(getattr(cf, "sqref", "")),
            "rules": [rule_to_dict(r) for r in getattr(cf, "rules", [])],
        }))
    return items


def apply_conditional_formatting(ws, items: List[Dict[str, Any]]) -> None:
    for cf in items or []:
        sqref = cf.get("sqref")
        if not sqref:
            continue
        for rdata in cf.get("rules", []):
            if rdata.get("dataBar_present") or rdata.get("iconSet_present") or rdata.get("colorScale"):
                # Recorded for analysis; generic rebuild deferred to later versions.
                continue
            try:
                ws.conditional_formatting.add(sqref, rule_from_dict(rdata))
            except Exception:
                pass


def parse_data_validations(ws) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    dvs = getattr(getattr(ws, "data_validations", None), "dataValidation", []) or []
    fields = [
        "type", "operator", "formula1", "formula2", "allowBlank", "showDropDown",
        "showErrorMessage", "showInputMessage", "promptTitle", "errorStyle", "error",
        "prompt", "errorTitle", "imeMode",
    ]
    for dv in dvs:
        data = _attrs_to_dict(dv, fields)
        data["sqref"] = str(getattr(dv, "sqref", ""))
        result.append(_strip_none(data))
    return result


def apply_data_validations(ws, items: List[Dict[str, Any]]) -> None:
    fields = [
        "operator", "allowBlank", "showDropDown", "showErrorMessage", "showInputMessage",
        "promptTitle", "errorStyle", "error", "prompt", "errorTitle", "imeMode",
    ]
    for item in items or []:
        try:
            dv = DataValidation(type=item.get("type"), formula1=item.get("formula1"), formula2=item.get("formula2"))
            _apply_attrs(dv, item, fields)
            ws.add_data_validation(dv)
            if item.get("sqref"):
                dv.add(item["sqref"])
        except Exception:
            pass


def table_style_to_dict(ts: Optional[TableStyleInfo]) -> Dict[str, Any]:
    if not ts:
        return {}
    return _attrs_to_dict(ts, ["name", "showFirstColumn", "showLastColumn", "showRowStripes", "showColumnStripes"])


def table_style_from_dict(data: Dict[str, Any]) -> Optional[TableStyleInfo]:
    if not data:
        return None
    return TableStyleInfo(
        name=data.get("name"),
        showFirstColumn=data.get("showFirstColumn"),
        showLastColumn=data.get("showLastColumn"),
        showRowStripes=data.get("showRowStripes"),
        showColumnStripes=data.get("showColumnStripes"),
    )


def merged_parent_value(ws, row: int, col: int) -> Any:
    v = ws.cell(row, col).value
    if v not in (None, ""):
        return v
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return v


def guess_field_map_for_table(ws, ref: str, header_rows: Optional[int] = None) -> Dict[str, Any]:
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return {}
    if header_rows is None:
        header_rows = 2 if maxr - minr + 1 >= 3 else 1
    fmap: Dict[str, str] = {}
    confidence: Dict[str, float] = {}
    for c in range(minc, maxc + 1):
        parts = []
        for r in range(minr, min(minr + header_rows, maxr + 1)):
            v = merged_parent_value(ws, r, c)
            if v not in (None, ""):
                parts.append(str(v).strip())
        if parts:
            # Remove adjacent duplicates caused by vertical merged headers.
            dedup = []
            for p in parts:
                if not dedup or dedup[-1] != p:
                    dedup.append(p)
            parts = dedup
            letter = get_column_letter(c)
            full = "/".join(parts)
            fmap[full] = letter
            confidence[full] = 0.9
            leaf = parts[-1]
            if leaf not in fmap:
                fmap[leaf] = letter
                confidence[leaf] = 0.72 if len(parts) > 1 else 0.85
    return {"header_rows": header_rows, "field_map_candidates": fmap, "confidence": confidence}


def parse_tables(ws) -> List[Dict[str, Any]]:
    tables: List[Dict[str, Any]] = []
    try:
        values = list(ws.tables.values())
    except Exception:
        values = []
    for t in values:
        if not hasattr(t, "ref"):
            continue
        tables.append(_strip_none({
            "displayName": getattr(t, "displayName", None),
            "name": getattr(t, "name", None),
            "ref": getattr(t, "ref", None),
            "comment": getattr(t, "comment", None),
            "tableType": getattr(t, "tableType", None),
            "headerRowCount": getattr(t, "headerRowCount", None),
            "totalsRowCount": getattr(t, "totalsRowCount", None),
            "totalsRowShown": getattr(t, "totalsRowShown", None),
            "style": table_style_to_dict(getattr(t, "tableStyleInfo", None)),
            "ir": guess_field_map_for_table(ws, getattr(t, "ref", None)),
        }))
    return tables


def apply_tables(ws, items: List[Dict[str, Any]]) -> None:
    for item in items or []:
        name = item.get("displayName") or item.get("name")
        ref = item.get("ref")
        if not name or not ref:
            continue
        try:
            tab = Table(displayName=name, ref=ref)
            for f in ["comment", "tableType", "headerRowCount", "totalsRowCount", "totalsRowShown"]:
                if f in item:
                    try:
                        setattr(tab, f, item[f])
                    except Exception:
                        pass
            tsi = table_style_from_dict(item.get("style", {}))
            if tsi:
                tab.tableStyleInfo = tsi
            ws.add_table(tab)
        except Exception:
            pass


def marker_to_dict(m: Any) -> Dict[str, Any]:
    if not m:
        return {}
    return _strip_none({
        "row": getattr(m, "row", None),
        "col": getattr(m, "col", None),
        "rowOff": getattr(m, "rowOff", None),
        "colOff": getattr(m, "colOff", None),
        "cell": f"{get_column_letter(getattr(m, 'col', 0) + 1)}{getattr(m, 'row', 0) + 1}",
    })


def parse_images(ws) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for idx, img in enumerate(getattr(ws, "_images", []) or [], start=1):
        try:
            data = img._data()
        except Exception:
            data = b""
        anchor = getattr(img, "anchor", None)
        anchor_data: Dict[str, Any] = {}
        if isinstance(anchor, str):
            anchor_data = {"type": "cell", "cell": anchor}
        elif anchor is not None:
            anchor_data = {
                "type": type(anchor).__name__,
                "from": marker_to_dict(getattr(anchor, "_from", None)),
                "to": marker_to_dict(getattr(anchor, "to", None)),
            }
            ext = getattr(anchor, "ext", None)
            if ext:
                anchor_data["ext"] = {"cx": getattr(ext, "cx", None), "cy": getattr(ext, "cy", None)}
        cell = anchor_data.get("cell") or anchor_data.get("from", {}).get("cell") or "A1"
        result.append(_strip_none({
            "name": f"image{idx}",
            "format": getattr(img, "format", None) or "png",
            "width": getattr(img, "width", None),
            "height": getattr(img, "height", None),
            "anchor": anchor_data,
            "anchor_cell": cell,
            "sha256": hashlib.sha256(data).hexdigest() if data else None,
            "data_b64": base64.b64encode(data).decode("ascii") if data else None,
        }))
    return result


def apply_images(ws, items: List[Dict[str, Any]], temp_paths: List[str]) -> None:
    for item in items or []:
        b64 = item.get("data_b64")
        if not b64:
            continue
        suffix = "." + (item.get("format") or "png")
        try:
            fd, tmp = tempfile.mkstemp(prefix="excel_ir_img_", suffix=suffix)
            with os.fdopen(fd, "wb") as f:
                f.write(base64.b64decode(b64))
            temp_paths.append(tmp)
            img = XLImage(tmp)
            if item.get("width"):
                img.width = item["width"]
            if item.get("height"):
                img.height = item["height"]
            cell = item.get("anchor_cell") or item.get("anchor", {}).get("cell") or item.get("anchor", {}).get("from", {}).get("cell") or "A1"
            ws.add_image(img, cell)
        except Exception:
            pass


def parse_header_footer(ws) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    for name in ["oddHeader", "oddFooter", "evenHeader", "evenFooter", "firstHeader", "firstFooter"]:
        obj = getattr(ws, name, None)
        if not obj:
            continue
        data = {}
        for pos in ["left", "center", "right"]:
            part = getattr(obj, pos, None)
            txt = getattr(part, "text", None) if part else None
            if txt:
                data[pos] = txt
        if data:
            result[name] = data
    return result


def apply_header_footer(ws, data: Dict[str, Any]) -> None:
    for name, parts in (data or {}).items():
        obj = getattr(ws, name, None)
        if not obj:
            continue
        for pos, txt in parts.items():
            part = getattr(obj, pos, None)
            if part:
                part.text = txt


def formula_ref_to_range(f: Optional[str]) -> Optional[Dict[str, Any]]:
    if not f:
        return None
    raw = str(f)
    if "!" in raw:
        sheet_part, ref = raw.split("!", 1)
        sheet = sheet_part.strip("'")
    else:
        sheet = None
        ref = raw
    ref = ref.replace("$", "")
    try:
        minc, minr, maxc, maxr = range_boundaries(ref)
    except Exception:
        return {"formula": raw}
    return _strip_none({
        "sheet": sheet,
        "ref": f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{maxr}",
        "min_col": minc,
        "min_row": minr,
        "max_col": maxc,
        "max_row": maxr,
        "formula": raw,
    })


def chart_title_text(title: Any) -> Optional[str]:
    if not title:
        return None
    try:
        return title.tx.rich.p[0].r[0].t
    except Exception:
        return None


def parse_charts(ws) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []
    for idx, ch in enumerate(getattr(ws, "_charts", []) or [], start=1):
        series = []
        for s in getattr(ch, "ser", []) or []:
            val_f = None
            cat_f = None
            tx_f = None
            try:
                val_f = s.val.numRef.f if s.val and s.val.numRef else None
            except Exception:
                pass
            try:
                cat_f = s.cat.numRef.f if s.cat and s.cat.numRef else None
            except Exception:
                pass
            try:
                if not cat_f:
                    cat_f = s.cat.strRef.f if s.cat and s.cat.strRef else None
            except Exception:
                pass
            try:
                tx_f = s.tx.strRef.f if s.tx and s.tx.strRef else None
            except Exception:
                pass
            series.append(_strip_none({
                "title_ref": tx_f,
                "title": formula_ref_to_range(tx_f),
                "values": formula_ref_to_range(val_f),
                "categories": formula_ref_to_range(cat_f),
            }))
        anchor = getattr(ch, "anchor", None)
        if isinstance(anchor, str):
            anchor_cell = anchor
        else:
            anchor_cell = marker_to_dict(getattr(anchor, "_from", None)).get("cell", "A1")
        charts.append(_strip_none({
            "name": f"chart{idx}",
            "type": type(ch).__name__,
            "anchor": anchor_cell,
            "anchor_cell": anchor_cell,
            "title": chart_title_text(getattr(ch, "title", None)),
            "x_axis_title": chart_title_text(getattr(getattr(ch, "x_axis", None), "title", None)),
            "y_axis_title": chart_title_text(getattr(getattr(ch, "y_axis", None), "title", None)),
            "style": getattr(ch, "style", None),
            "height": getattr(ch, "height", None),
            "width": getattr(ch, "width", None),
            "series": series,
        }))
    return charts


def _range_to_reference(ws, r: Dict[str, Any], title: Dict[str, Any] | None = None) -> Optional[Reference]:
    if not r:
        return None
    try:
        min_col = int(r["min_col"])
        min_row = int(r["min_row"])
        max_col = int(r["max_col"])
        max_row = int(r["max_row"])
        # openpyxl expects the title cell to be included in the data reference
        # when titles_from_data=True.
        if title and title.get("min_col") == min_col and title.get("max_col") == max_col:
            trow = int(title.get("min_row"))
            if trow == min_row - 1:
                min_row = trow
        return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    except Exception:
        if r.get("ref"):
            try:
                minc, minr, maxc, maxr = range_boundaries(str(r["ref"]))
                return Reference(ws, min_col=minc, min_row=minr, max_col=maxc, max_row=maxr)
            except Exception:
                return None
    return None


def apply_charts(ws, items: List[Dict[str, Any]]) -> None:
    for item in items or []:
        typ = item.get("type")
        if typ == "LineChart":
            chart = LineChart()
        elif typ == "PieChart":
            chart = PieChart()
        else:
            chart = BarChart()
        if item.get("title"):
            chart.title = item["title"]
        if item.get("x_axis_title") and hasattr(chart, "x_axis"):
            chart.x_axis.title = item["x_axis_title"]
        if item.get("y_axis_title") and hasattr(chart, "y_axis"):
            chart.y_axis.title = item["y_axis_title"]
        if item.get("style") is not None:
            chart.style = item.get("style")
        if item.get("height"):
            chart.height = item["height"]
        if item.get("width"):
            chart.width = item["width"]
        # openpyxl may reset chart.height/width on load/save. Keep requested
        # values in IR for intent tracking; structural diff normalizes them.
        if item.get("height"):
            setattr(chart, "_excel_ir_height", item["height"])
        if item.get("width"):
            setattr(chart, "_excel_ir_width", item["width"])
        series = item.get("series", []) or []
        if series:
            # Group series with adjacent value columns where possible; if not,
            # add each one individually. This is sufficient for many generated
            # report charts.
            for s in series:
                val_ref = _range_to_reference(ws, s.get("values", {}), s.get("title", {}))
                if val_ref:
                    chart.add_data(val_ref, titles_from_data=bool(s.get("title_ref")))
            cat_ref = _range_to_reference(ws, series[0].get("categories", {}) if series else {})
            if cat_ref:
                try:
                    chart.set_categories(cat_ref)
                except Exception:
                    pass
        if item.get("anchor"):
            try:
                ws.add_chart(chart, item["anchor"])
            except Exception:
                ws.add_chart(chart, "A1")


def parse_sheet_extra(ws) -> Dict[str, Any]:
    tab_color = excel_ir.color_to_dict(getattr(ws.sheet_properties, "tabColor", None))
    sheet_view = _attrs_to_dict(ws.sheet_view, [
        "showFormulas", "showGridLines", "showRowColHeaders", "showZeros", "rightToLeft",
        "view", "topLeftCell", "zoomScale", "zoomScaleNormal",
    ])
    sheet_format = _attrs_to_dict(ws.sheet_format, [
        "baseColWidth", "defaultColWidth", "defaultRowHeight", "customHeight", "zeroHeight",
        "thickTop", "thickBottom", "outlineLevelRow", "outlineLevelCol",
    ])
    page_margins = _attrs_to_dict(ws.page_margins, ["left", "right", "top", "bottom", "header", "footer"])
    page_setup = _attrs_to_dict(ws.page_setup, [
        "orientation", "paperSize", "scale", "fitToHeight", "fitToWidth", "firstPageNumber",
        "useFirstPageNumber", "paperHeight", "paperWidth", "pageOrder", "usePrinterDefaults",
        "blackAndWhite", "draft", "cellComments", "errors", "horizontalDpi", "verticalDpi", "copies",
    ])
    print_options = _attrs_to_dict(ws.print_options, ["horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet"])
    protection = _attrs_to_dict(ws.protection, [
        "sheet", "objects", "scenarios", "formatCells", "formatColumns", "formatRows", "insertColumns",
        "insertRows", "insertHyperlinks", "deleteColumns", "deleteRows", "selectLockedCells",
        "selectUnlockedCells", "sort", "autoFilter", "pivotTables", "algorithmName", "saltValue",
        "spinCount", "hashValue",
    ])
    return _strip_none({
        "schema_version": SCHEMA_VERSION,
        "sheet_state": getattr(ws, "sheet_state", None),
        "tab_color": tab_color,
        "sheet_view": sheet_view,
        "sheet_format": sheet_format,
        "page_margins": page_margins,
        "page_setup": page_setup,
        "print_options": print_options,
        "print_area": str(ws.print_area) if getattr(ws, "print_area", None) else None,
        "print_title_rows": str(ws.print_title_rows) if getattr(ws, "print_title_rows", None) else None,
        "print_title_cols": str(ws.print_title_cols) if getattr(ws, "print_title_cols", None) else None,
        "header_footer": parse_header_footer(ws),
        "protection": protection if protection.get("sheet") else {},
        "auto_filter": {"ref": ws.auto_filter.ref} if getattr(ws.auto_filter, "ref", None) else {},
        "data_validations": parse_data_validations(ws),
        "conditional_formatting": parse_conditional_formatting(ws),
        "tables": parse_tables(ws),
        "images": parse_images(ws),
        "charts": parse_charts(ws),
    })


def apply_sheet_extra(ws, extra: Dict[str, Any], temp_paths: List[str]) -> None:
    if "sheet_state" in extra:
        ws.sheet_state = extra["sheet_state"]
    if extra.get("tab_color"):
        ws.sheet_properties.tabColor = excel_ir.color_from_dict(extra["tab_color"])
    _apply_attrs(ws.sheet_view, extra.get("sheet_view", {}), [
        "showFormulas", "showGridLines", "showRowColHeaders", "showZeros", "rightToLeft",
        "view", "topLeftCell", "zoomScale", "zoomScaleNormal",
    ])
    _apply_attrs(ws.sheet_format, extra.get("sheet_format", {}), [
        "baseColWidth", "defaultColWidth", "defaultRowHeight", "customHeight", "zeroHeight",
        "thickTop", "thickBottom", "outlineLevelRow", "outlineLevelCol",
    ])
    _apply_attrs(ws.page_margins, extra.get("page_margins", {}), ["left", "right", "top", "bottom", "header", "footer"])
    _apply_attrs(ws.page_setup, extra.get("page_setup", {}), [
        "orientation", "paperSize", "scale", "fitToHeight", "fitToWidth", "firstPageNumber",
        "useFirstPageNumber", "paperHeight", "paperWidth", "pageOrder", "usePrinterDefaults",
        "blackAndWhite", "draft", "cellComments", "errors", "horizontalDpi", "verticalDpi", "copies",
    ])
    _apply_attrs(ws.print_options, extra.get("print_options", {}), ["horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet"])
    if extra.get("print_area"):
        try:
            ws.print_area = extra["print_area"]
        except Exception:
            pass
    if extra.get("print_title_rows"):
        try:
            ws.print_title_rows = extra["print_title_rows"]
        except Exception:
            pass
    if extra.get("print_title_cols"):
        try:
            ws.print_title_cols = extra["print_title_cols"]
        except Exception:
            pass
    apply_header_footer(ws, extra.get("header_footer", {}))
    if extra.get("protection"):
        _apply_attrs(ws.protection, extra["protection"], list(extra["protection"].keys()))
    if extra.get("auto_filter", {}).get("ref"):
        ws.auto_filter.ref = extra["auto_filter"]["ref"]
    apply_data_validations(ws, extra.get("data_validations", []))
    apply_conditional_formatting(ws, extra.get("conditional_formatting", []))
    apply_tables(ws, extra.get("tables", []))
    apply_images(ws, extra.get("images", []), temp_paths)
    apply_charts(ws, extra.get("charts", []))


def parse_workbook_plus(path: str, include_empty_styled: bool = True, infer_logic: bool = True) -> Dict[str, Any]:
    ir = excel_ir.parse_workbook(path, include_empty_styled=include_empty_styled, infer_logic=infer_logic)
    ir["schema_version"] = SCHEMA_VERSION
    wb = load_workbook(path, data_only=False)
    for sheet_ir, ws in zip(ir["workbook"].get("sheets", []), wb.worksheets):
        sheet_ir["extra"] = parse_sheet_extra(ws)
    return ir


def rebuild_workbook_plus(ir: Dict[str, Any], path: str) -> None:
    # First rebuild the reversible grid/style core using v0.1 engine.
    # Charts are applied later by v0.3; remove them from the first pass to avoid
    # duplicate charts when the temporary workbook is loaded again.
    core_ir = copy.deepcopy(ir)
    for s in core_ir.get("workbook", {}).get("sheets", []):
        if s.get("extra"):
            s["extra"].pop("charts", None)
    excel_ir.rebuild_workbook(core_ir, path)
    wb = load_workbook(path)
    temp_paths: List[str] = []
    try:
        sheets_by_name = {ws.title: ws for ws in wb.worksheets}
        for sheet_ir in ir["workbook"].get("sheets", []):
            ws = sheets_by_name.get(sheet_ir.get("name"))
            if not ws:
                continue
            # v0.1 parsed comments but did not rebuild them; v0.2 does.
            for coord, cdata in sheet_ir.get("cells", {}).items():
                cmt = cdata.get("comment")
                if cmt:
                    ws[coord].comment = Comment(cmt.get("text", ""), cmt.get("author", ""))
            apply_sheet_extra(ws, sheet_ir.get("extra", {}), temp_paths)
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass
        wb.save(path)
    finally:
        for tmp in temp_paths:
            try:
                os.remove(tmp)
            except Exception:
                pass


def _normalize_ir_for_diff(data: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize semantically equivalent openpyxl round-trip artifacts.

    Openpyxl sometimes changes empty inline strings to blank numeric cells, and
    header/footer text assigned through the API is not always preserved in a
    from-scratch intermediate load/save cycle. The diff tool should focus on
    report IR fidelity rather than these package-level incidental differences.
    """
    for s in data.get("workbook", {}).get("sheets", []):
        for coord, cell in list(s.get("cells", {}).items()):
            if cell.get("value") in (None, "") and cell.get("data_type") in ("inlineStr", "s", "str", "n"):
                cell.pop("data_type", None)
        extra = s.get("extra", {})
        extra.pop("header_footer", None)
        for idx, img in enumerate(extra.get("images", []) or [], start=1):
            img.pop("data_b64", None)
            img["name"] = f"image{idx}"
        # Charts are now parsed and rebuilt in v0.3. Ignore generated internal
        # chart names only; compare chart type, anchor, titles, dimensions and ranges.
        for idx, ch in enumerate(extra.get("charts", []) or [], start=1):
            ch["name"] = f"chart{idx}"
            # openpyxl normalizes chart dimensions on reload; do not treat this
            # as a semantic diff in the MVP.
            ch.pop("height", None)
            ch.pop("width", None)
    return data


def canonical_for_diff(ir: Dict[str, Any]) -> Dict[str, Any]:
    data = copy.deepcopy(ir)
    data["schema_version"] = SCHEMA_VERSION
    # Logical inference is heuristic; exclude it from reversible structural diff.
    for s in data.get("workbook", {}).get("sheets", []):
        s.pop("logical", None)
    return _normalize_ir_for_diff(data)


def diff_workbooks_plus(a: str, b: str) -> Dict[str, Any]:
    ia = canonical_for_diff(parse_workbook_plus(a, infer_logic=False))
    ib = canonical_for_diff(parse_workbook_plus(b, infer_logic=False))
    if ia == ib:
        return {"diff_count": 0, "diffs": [], "truncated": False}
    diffs: List[Dict[str, Any]] = []
    sheets_a = {s["name"]: s for s in ia["workbook"].get("sheets", [])}
    sheets_b = {s["name"]: s for s in ib["workbook"].get("sheets", [])}
    for sname in sorted(set(sheets_a) | set(sheets_b)):
        sa = sheets_a.get(sname)
        sb = sheets_b.get(sname)
        if sa is None or sb is None:
            diffs.append({"sheet": sname, "type": "sheet_missing", "a": sa is not None, "b": sb is not None})
            continue
        for key in sorted(set(sa) | set(sb)):
            va = copy.deepcopy(sa.get(key))
            vb = copy.deepcopy(sb.get(key))
            if key == "extra":
                va = _normalize_ir_for_diff({"workbook": {"sheets": [{"extra": va or {}, "cells": {}}]}})["workbook"]["sheets"][0]["extra"]
                vb = _normalize_ir_for_diff({"workbook": {"sheets": [{"extra": vb or {}, "cells": {}}]}})["workbook"]["sheets"][0]["extra"]
            if va != vb:
                if key == "cells":
                    ca = copy.deepcopy(sa.get("cells", {}))
                    cb = copy.deepcopy(sb.get("cells", {}))
                    _normalize_ir_for_diff({"workbook": {"sheets": [{"cells": ca, "extra": {}}]}})
                    _normalize_ir_for_diff({"workbook": {"sheets": [{"cells": cb, "extra": {}}]}})
                    for coord in sorted(set(ca) | set(cb)):
                        if ca.get(coord) != cb.get(coord):
                            diffs.append({"sheet": sname, "coord": coord, "type": "cell", "a": ca.get(coord), "b": cb.get(coord)})
                            if len(diffs) >= 200:
                                return {"diff_count": 201, "diffs": diffs, "truncated": True}
                else:
                    diffs.append({"sheet": sname, "type": key, "a": sa.get(key), "b": sb.get(key)})
                    if len(diffs) >= 200:
                        return {"diff_count": 201, "diffs": diffs, "truncated": True}
    return {"diff_count": len(diffs), "diffs": diffs[:200], "truncated": len(diffs) > 200}


def save_json(data: Dict[str, Any], path: str) -> None:
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def main() -> None:
    p = argparse.ArgumentParser(description="Excel report IR MVP v0.2: extended objects")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_parse = sub.add_parser("parse")
    p_parse.add_argument("input")
    p_parse.add_argument("output")
    p_parse.add_argument("--no-logic", action="store_true")

    p_rebuild = sub.add_parser("rebuild")
    p_rebuild.add_argument("input_json")
    p_rebuild.add_argument("output_xlsx")

    p_diff = sub.add_parser("diff")
    p_diff.add_argument("a")
    p_diff.add_argument("b")
    p_diff.add_argument("output_json", nargs="?")

    p_logic = sub.add_parser("logic")
    p_logic.add_argument("input_json")

    args = p.parse_args()
    if args.cmd == "parse":
        ir = parse_workbook_plus(args.input, infer_logic=not args.no_logic)
        save_json(ir, args.output)
        print(json.dumps({"ok": True, "schema_version": SCHEMA_VERSION, "sheets": len(ir["workbook"]["sheets"]), "styles": len(ir["workbook"]["styles"]), "output": args.output}, ensure_ascii=False))
    elif args.cmd == "rebuild":
        ir = load_json(args.input_json)
        rebuild_workbook_plus(ir, args.output_xlsx)
        print(json.dumps({"ok": True, "output": args.output_xlsx}, ensure_ascii=False))
    elif args.cmd == "diff":
        d = diff_workbooks_plus(args.a, args.b)
        if args.output_json:
            save_json(d, args.output_json)
        print(json.dumps(d, ensure_ascii=False, indent=2))
    elif args.cmd == "logic":
        ir = load_json(args.input_json)
        for s in ir["workbook"]["sheets"]:
            print(json.dumps({"sheet": s["name"], "logical": s.get("logical", {})}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

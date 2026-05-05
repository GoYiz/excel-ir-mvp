from __future__ import annotations

from pathlib import Path

from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def make_logo(path: Path) -> None:
    img = Image.new("RGB", (140, 42), "#1F4E79")
    d = ImageDraw.Draw(img)
    d.rectangle((4, 4, 136, 38), outline="white", width=2)
    d.text((18, 13), "ACME REPORT", fill="white")
    img.save(path)


def make_complex_sample(path: str) -> None:
    root = Path(path).parent
    root.mkdir(parents=True, exist_ok=True)
    logo = root / "complex_logo.png"
    make_logo(logo)

    wb = Workbook()
    ws = wb.active
    ws.title = "经营驾驶舱"
    ws.sheet_properties.tabColor = "FF1F4E79"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_format.defaultRowHeight = 18
    ws.freeze_panes = "A8"
    ws.print_area = "A1:L24"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.oddHeader.center.text = "&B经营驾驶舱&B"
    ws.oddFooter.right.text = "第 &P 页 / 共 &N 页"

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col not in (1, 2) else 14
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 24

    # Styles
    title_font = Font(name="微软雅黑", size=20, bold=True, color="FFFFFFFF")
    white_font = Font(name="微软雅黑", size=10, color="FFFFFFFF", bold=True)
    normal = Font(name="微软雅黑", size=10)
    gray = Font(name="微软雅黑", size=9, color="FF666666")
    red = Font(name="微软雅黑", size=10, color="FFC00000", bold=True)
    blue_fill = PatternFill("solid", fgColor="FF1F4E79")
    light_blue = PatternFill("solid", fgColor="FFD9EAF7")
    orange_fill = PatternFill("solid", fgColor="FFFCE4D6")
    green_fill = PatternFill("solid", fgColor="FFE2F0D9")
    gray_fill = PatternFill("solid", fgColor="FFF2F2F2")
    thin = Side(style="thin", color="FFBFBFBF")
    medium = Side(style="medium", color="FF1F4E79")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_border = Border(left=thin, right=thin, top=medium, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:L1")
    ws["A1"] = "集团经营驾驶舱 - 月度经营分析"
    ws["A1"].font = title_font
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = center

    ws.add_image(XLImage(str(logo)), "A2")

    ws.merge_cells("C2:E2")
    ws["C2"] = "期间：2026-05"
    ws.merge_cells("F2:H2")
    ws["F2"] = "币种：人民币万元"
    ws.merge_cells("I2:L2")
    ws["I2"] = "版本：V1.0 / 未审计"
    for c in ["C2", "F2", "I2"]:
        ws[c].font = gray
        ws[c].alignment = center

    ws.merge_cells("A4:L4")
    ws["A4"] = "一、核心 KPI"
    ws["A4"].font = Font(name="微软雅黑", size=13, bold=True, color="FF1F4E79")

    headers_top = {
        "A5": "业务线", "B5": "区域", "C5": "收入", "F5": "利润", "I5": "效率", "K5": "状态"
    }
    merges = ["A5:A6", "B5:B6", "C5:E5", "F5:H5", "I5:J5", "K5:L5"]
    for m in merges:
        ws.merge_cells(m)
    for coord, txt in headers_top.items():
        ws[coord] = txt
    headers_bottom = {
        "C6": "本月", "D6": "预算", "E6": "达成率",
        "F6": "本月", "G6": "预算", "H6": "达成率",
        "I6": "人效", "J6": "周转天数", "K6": "评级", "L6": "备注",
    }
    for coord, txt in headers_bottom.items():
        ws[coord] = txt
    for row in ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=12):
        for cell in row:
            cell.font = white_font
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = border

    data = [
        ["云业务", "华东", 3200, 3000, None, 680, 650, None, 42.1, 31, "A", "增长稳健"],
        ["云业务", "华南", 2800, 3100, None, 520, 600, None, 38.5, 35, "B", "需关注预算缺口"],
        ["软件", "华东", 1900, 1800, None, 410, 390, None, 55.0, 22, "A", ""],
        ["软件", "华北", 1500, 1700, None, 260, 330, None, 46.8, 28, "C", "利润承压"],
        ["服务", "西南", 900, 850, None, 180, 160, None, 33.2, 41, "B", ""],
    ]
    for r, row in enumerate(data, start=7):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.font = normal
            cell.border = border
            cell.alignment = right if c in (3, 4, 5, 6, 7, 8, 9, 10) else center
            if c in (3, 4, 6, 7, 9, 10):
                cell.number_format = '#,##0.0'
            if c in (5, 8):
                cell.number_format = '0.0%'
        ws.cell(r, 5).value = f"=C{r}/D{r}"
        ws.cell(r, 8).value = f"=F{r}/G{r}"
    total = 12
    ws.merge_cells(start_row=total, start_column=1, end_row=total, end_column=2)
    ws.cell(total, 1, "总计")
    for c in range(1, 13):
        cell = ws.cell(total, c)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = green_fill
        cell.border = top_border
        cell.alignment = right if c >= 3 else center
    for c in [3, 4, 6, 7, 9, 10]:
        letter = get_column_letter(c)
        ws.cell(total, c).value = f"=SUM({letter}7:{letter}11)"
        ws.cell(total, c).number_format = '#,##0.0'
    ws.cell(total, 5).value = "=C12/D12"
    ws.cell(total, 8).value = "=F12/G12"
    ws.cell(total, 5).number_format = '0.0%'
    ws.cell(total, 8).number_format = '0.0%'

    # Excel table object and auto filter
    tab = Table(displayName="KPI_Table", ref="A5:L12")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.auto_filter.ref = "A5:L12"

    # Data validation and conditional formatting
    dv = DataValidation(type="list", formula1='"A,B,C,D"', allowBlank=False, showErrorMessage=True, errorTitle="评级错误", error="只能选择 A/B/C/D")
    ws.add_data_validation(dv)
    dv.add("K7:K11")
    ws.conditional_formatting.add("E7:E12", CellIsRule(operator="lessThan", formula=["0.9"], fill=orange_fill, font=red))
    ws.conditional_formatting.add("H7:H12", CellIsRule(operator="lessThan", formula=["0.9"], fill=orange_fill, font=red))

    ws["L9"].comment = Comment("该单元格用于记录特殊事项。", "Minis")

    # Secondary matrix
    start = 15
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=12)
    ws.cell(start, 1, "二、区域 × 业务线收入矩阵")
    ws.cell(start, 1).font = Font(name="微软雅黑", size=13, bold=True, color="FF1F4E79")
    matrix_headers = ["区域", "云业务", "软件", "服务", "合计"]
    for c, h in enumerate(matrix_headers, start=1):
        cell = ws.cell(start + 1, c, h)
        cell.font = white_font
        cell.fill = blue_fill
        cell.border = border
        cell.alignment = center
    matrix = [
        ["华东", 3200, 1900, 0],
        ["华南", 2800, 0, 0],
        ["华北", 0, 1500, 0],
        ["西南", 0, 0, 900],
    ]
    for r, row in enumerate(matrix, start=start + 2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.font = normal
            cell.border = border
            cell.alignment = right if c > 1 else center
            if c > 1:
                cell.number_format = '#,##0.0'
        ws.cell(r, 5).value = f"=SUM(B{r}:D{r})"
        ws.cell(r, 5).fill = gray_fill
    total2 = start + 6
    ws.cell(total2, 1, "合计")
    for c in range(1, 6):
        cell = ws.cell(total2, c)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = green_fill
        cell.border = top_border
        cell.alignment = right if c > 1 else center
    for c in range(2, 6):
        letter = get_column_letter(c)
        ws.cell(total2, c).value = f"=SUM({letter}{start+2}:{letter}{total2-1})"
        ws.cell(total2, c).number_format = '#,##0.0'

    # Chart: v0.2 records chart metadata, not full recreate.
    chart = BarChart()
    chart.title = "收入矩阵"
    chart.y_axis.title = "万元"
    chart.x_axis.title = "区域"
    data_ref = Reference(ws, min_col=2, max_col=4, min_row=start + 1, max_row=start + 5)
    cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 5)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "G16")

    wb.save(path)


if __name__ == "__main__":
    make_complex_sample("/var/minis/workspace/excel_ir_mvp/complex_report.xlsx")
    print("created /var/minis/workspace/excel_ir_mvp/complex_report.xlsx")

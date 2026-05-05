from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw
from pathlib import Path
p=Path('/var/minis/workspace/excel_ir_mvp/tmp_logo.png')
img=Image.new('RGB',(80,30),'#4472C4')
d=ImageDraw.Draw(img); d.text((8,8),'LOGO',fill='white')
img.save(p)
wb=Workbook(); ws=wb.active; ws.add_image(XLImage(str(p)), 'C3'); out='/var/minis/workspace/excel_ir_mvp/tmp_img.xlsx'; wb.save(out)
wb2=load_workbook(out); ws2=wb2.active; im=ws2._images[0]
print(type(im), im.width, im.height, type(im.anchor), im.anchor)
print(im.anchor.__dict__)
print(im.anchor._from.__dict__)
print(len(im._data()))

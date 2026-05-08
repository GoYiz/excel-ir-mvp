from openpyxl import Workbook
from pathlib import Path

path = Path('/var/minis/workspace/excel_ir_mvp/tests/fixtures/multi_header_dates.xlsx')
wb = Workbook()
ws = wb.active
ws.title = '日期表'
ws['A1'] = '项目'
ws.merge_cells('B1:D1'); ws['B1'] = 2026
ws.merge_cells('E1:G1'); ws['E1'] = 2027
ws.merge_cells('B2:D2'); ws['B2'] = 5
ws.merge_cells('E2:G2'); ws['E2'] = 5
for idx, day in enumerate([7, 8, 9, 7, 8, 9], start=2):
    ws.cell(3, idx).value = day
ws['A4'] = '门店A'; ws['B4'] = 10; ws['C4'] = 20; ws['D4'] = 30; ws['E4'] = 40; ws['F4'] = 50; ws['G4'] = 60
ws['A5'] = '门店B'; ws['B5'] = 11; ws['C5'] = 21; ws['D5'] = 31; ws['E5'] = 41; ws['F5'] = 51; ws['G5'] = 61
wb.save(path)
print(path)

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / 'tests' / 'fixtures'


def style_header(ws, ref):
    thin = Side(style='thin', color='FF999999')
    for row in ws[ref]:
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FFD9EAF7')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def make_native(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'NativeTable'
    ws.append(['Region', 'Revenue', 'Cost'])
    ws.append(['North', 100, 70])
    ws.append(['South', 120, 82])
    tab = Table(displayName='NativeSales', ref='A1:C3')
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tab)
    wb.save(path)


def make_semantic(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'SemanticTable'
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws['A1'] = '业务线'
    ws['B1'] = '收入'
    ws['B2'] = '本月'
    ws['C2'] = '预算'
    rows = [['硬件', 100, 90], ['软件', 120, 110], ['合计', '=SUM(B3:B4)', '=SUM(C3:C4)']]
    for row in rows:
        ws.append(row)
    style_header(ws, 'A1:C2')
    thin = Side(style='thin', color='FF999999')
    for row in ws['A3:C5']:
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    tab = Table(displayName='SemanticSales', ref='A1:C5')
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tab)
    wb.save(path)


def main():
    FIXTURES.mkdir(parents=True, exist_ok=True)
    make_native(FIXTURES / 'native_table.xlsx')
    make_semantic(FIXTURES / 'semantic_table.xlsx')
    print({'ok': True, 'fixtures': ['native_table.xlsx', 'semantic_table.xlsx']})


if __name__ == '__main__':
    main()

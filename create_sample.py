from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def make_sample(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "销售日报"

    # Layout
    widths = {
        "A": 10,
        "B": 14,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A7"

    # Styles
    title_font = Font(name="微软雅黑", size=18, bold=True, color="FF1F4E79")
    meta_font = Font(name="微软雅黑", size=10, italic=True, color="FF666666")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
    normal_font = Font(name="微软雅黑", size=10)
    total_font = Font(name="微软雅黑", size=10, bold=True)

    dark_fill = PatternFill("solid", fgColor="FF4472C4")
    sub_fill = PatternFill("solid", fgColor="FF5B9BD5")
    total_fill = PatternFill("solid", fgColor="FFD9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFFFF2CC")

    thin_gray = Side(style="thin", color="FFBFBFBF")
    medium_blue = Side(style="medium", color="FF4472C4")
    table_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    top_border = Border(left=thin_gray, right=thin_gray, top=medium_blue, bottom=thin_gray)

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title and metadata
    ws.merge_cells("A1:J1")
    ws["A1"] = "华东大区销售日报"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:C2")
    ws["A2"] = "报表日期：2026-05-04"
    ws["A2"].font = meta_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D2:F2")
    ws["D2"] = "单位：万元"
    ws["D2"].font = meta_font
    ws["D2"].alignment = center

    ws.merge_cells("G2:J2")
    ws["G2"] = "制表人：运营分析组"
    ws["G2"].font = meta_font
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")

    ws["A4"] = "一、渠道销售概览"
    ws["A4"].font = Font(name="微软雅黑", size=12, bold=True, color="FF333333")

    # Main table header
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:B6")
    ws.merge_cells("C5:E5")
    ws.merge_cells("F5:H5")
    ws.merge_cells("I5:J5")
    headers = {
        "A5": "区域",
        "B5": "渠道",
        "C5": "本日",
        "F5": "本月累计",
        "I5": "达成情况",
        "C6": "销售额",
        "D6": "毛利",
        "E6": "订单数",
        "F6": "销售额",
        "G6": "毛利",
        "H6": "订单数",
        "I6": "目标",
        "J6": "达成率",
    }
    for coord, value in headers.items():
        cell = ws[coord]
        cell.value = value
        cell.font = header_font
        cell.fill = dark_fill if coord.endswith("5") else sub_fill
        cell.alignment = wrap_center
        cell.border = table_border

    # Style merged member cells too for border continuity
    for row in ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=10):
        for cell in row:
            cell.border = table_border
            if cell.value is None:
                cell.fill = dark_fill if cell.row == 5 else sub_fill
                cell.alignment = wrap_center

    data = [
        ["上海", "直营", 128.5, 31.2, 342, 1850.2, 421.0, 4520, 2200, None],
        ["上海", "经销", 96.3, 21.7, 210, 1420.7, 300.5, 3311, 1800, None],
        ["江苏", "直营", 155.0, 40.2, 410, 2011.4, 530.1, 5020, 2400, None],
        ["江苏", "经销", 88.8, 18.6, 198, 1310.0, 260.4, 2902, 1600, None],
        ["浙江", "直营", 173.2, 45.8, 433, 2300.9, 610.2, 5800, 2600, None],
        ["浙江", "经销", 101.5, 20.4, 221, 1502.1, 288.2, 3109, 1700, None],
    ]
    start = 7
    for idx, row in enumerate(data, start=start):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(idx, col, value)
            cell.font = normal_font
            cell.border = table_border
            cell.alignment = right if col >= 3 else center
            if col in (3, 4, 6, 7, 9, 10):
                cell.number_format = '#,##0.0'
            elif col in (5, 8):
                cell.number_format = '#,##0'
        ws.cell(idx, 10).value = f"=F{idx}/I{idx}"
        ws.cell(idx, 10).number_format = "0.0%"

    total_row = start + len(data)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(total_row, 1, "合计")
    for col in range(1, 11):
        cell = ws.cell(total_row, col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = top_border
        cell.alignment = center if col <= 2 else right
        if col in (3, 4, 6, 7, 9, 10):
            cell.number_format = '#,##0.0'
        elif col in (5, 8):
            cell.number_format = '#,##0'
    for col in [3, 4, 5, 6, 7, 8, 9]:
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}{start}:{letter}{total_row-1})"
    ws.cell(total_row, 10).value = f"=F{total_row}/I{total_row}"
    ws.cell(total_row, 10).number_format = "0.0%"

    # Secondary small table separated by blank row
    sec_top = total_row + 3
    ws[sec_top][0].value = "二、异常提示"
    ws.cell(sec_top, 1).font = Font(name="微软雅黑", size=12, bold=True, color="FFC00000")
    sec_header = sec_top + 1
    for col, val in enumerate(["级别", "区域", "问题", "负责人", "截止日期"], start=1):
        cell = ws.cell(sec_header, col, val)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="FFC00000")
        cell.alignment = center
        cell.border = table_border
    alerts = [
        ["高", "江苏", "经销渠道订单数低于预警线", "张三", "2026-05-05"],
        ["中", "上海", "直营毛利率波动", "李四", "2026-05-06"],
    ]
    for r_idx, row in enumerate(alerts, start=sec_header + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = normal_font
            cell.border = table_border
            cell.alignment = center if c_idx in (1, 2, 4, 5) else Alignment(wrap_text=True, vertical="center")
            if c_idx == 1:
                cell.fill = yellow_fill

    # Notes
    note_row = sec_header + len(alerts) + 3
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    ws.cell(note_row, 1).value = "注：本报表由系统初步生成，经人工调整格式后发布。"
    ws.cell(note_row, 1).font = Font(name="微软雅黑", size=9, color="FF808080")
    ws.cell(note_row, 1).alignment = Alignment(horizontal="left")

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    make_sample("/var/minis/workspace/excel_ir_mvp/sample_report.xlsx")
    print("created /var/minis/workspace/excel_ir_mvp/sample_report.xlsx")
